from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from accounts.permissions import (
    app_access_required, permission_required, finance_staff_required
)
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from datetime import date, datetime, timedelta
from .models import Invoice, InvoiceLineItem, Payment, InsuranceClaim, PaymentPlan, ServicePriceGroup, BillingAuditLog
import csv
from .forms import InvoiceForm, InvoiceLineItemFormSet, PaymentForm, InsuranceClaimForm, PaymentPlanForm
from patients.models import Patient
from appointments.models import Appointment, Service
from laboratory.models import LabTest, LabTestPriceGroup
from clinic_system.pagination_utils import paginate_queryset
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import io
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.conf import settings
from django.urls import reverse

@login_required
@app_access_required('billing')
def billing_dashboard(request):
    from django.utils import timezone
    from django.db.models import Count
    from .models import GroupInvoice, GroupPayment
    
    # Calculate key metrics
    today = timezone.now().date()
    current_month = today.replace(day=1)
    
    # Regular invoice metrics
    monthly_revenue = Payment.objects.filter(
        status='completed', 
        payment_date__gte=current_month
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Add group payments to monthly revenue
    group_monthly_revenue = GroupPayment.objects.filter(
        status='completed',
        payment_date__gte=current_month
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    total_monthly_revenue = monthly_revenue + group_monthly_revenue
    
    # Outstanding amount (sent + overdue invoices) - regular invoices
    outstanding_amount = Invoice.objects.filter(
        status__in=['sent', 'overdue']
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Add group invoices outstanding amount
    group_outstanding = GroupInvoice.objects.filter(
        status__in=['sent', 'overdue']
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_outstanding = outstanding_amount + group_outstanding
    
    # Daily payments - regular
    daily_payments = Payment.objects.filter(
        status='completed',
        payment_date=today
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Daily group payments
    group_daily_payments = GroupPayment.objects.filter(
        status='completed',
        payment_date__date=today
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    total_daily_payments = daily_payments + group_daily_payments
    
    # Overdue count - regular + group
    overdue_count = Invoice.objects.filter(status='overdue').count()
    group_overdue_count = GroupInvoice.objects.filter(status='overdue').count()
    total_overdue_count = overdue_count + group_overdue_count
    
    # Combined recent and draft invoices (last 15) - include group invoices
    all_invoices = Invoice.objects.select_related('patient').order_by('-created_at')[:15]
    
    # Recent group invoices
    recent_group_invoices = GroupInvoice.objects.select_related('patient_group', 'created_by').order_by('-created_at')[:5]
    
    # Payment methods summary for current month - combine regular and group
    payment_methods_data = Payment.objects.filter(
        status='completed',
        payment_date__gte=current_month
    ).values('payment_method').annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    # Group payment methods (mapped to similar categories)
    group_payment_methods = GroupPayment.objects.filter(
        status='completed',
        payment_date__gte=current_month
    ).values('payment_method').annotate(
        total=Sum('amount')
    )
    
    # Overdue invoices - combine regular and group
    overdue_invoices = Invoice.objects.filter(status='overdue').select_related('patient')[:5]
    overdue_group_invoices = GroupInvoice.objects.filter(status='overdue').select_related('patient_group')[:5]
    
    # Insurance claims summary
    insurance_claims_summary = {
        'pending': InsuranceClaim.objects.filter(status='pending').count(),
        'approved': InsuranceClaim.objects.filter(status='approved').count(),
        'processing': InsuranceClaim.objects.filter(status='processing').count(),
        'denied': InsuranceClaim.objects.filter(status='denied').count(),
    }
    
    # Revenue chart data (last 6 months) - include both regular and group payments
    import json
    revenue_chart_labels = []
    revenue_chart_data = []
    for i in range(5, -1, -1):
        month_date = (current_month - timedelta(days=32*i)).replace(day=1)
        next_month = (month_date + timedelta(days=32)).replace(day=1)
        
        # Regular payments
        month_revenue = Payment.objects.filter(
            status='completed',
            payment_date__gte=month_date,
            payment_date__lt=next_month
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        
        # Group payments
        group_month_revenue = GroupPayment.objects.filter(
            status='completed',
            payment_date__gte=month_date,
            payment_date__lt=next_month
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        
        total_month_revenue = month_revenue + group_month_revenue
        
        revenue_chart_labels.append(month_date.strftime('%b %Y'))
        revenue_chart_data.append(float(total_month_revenue))

    # Aging buckets chart data - include both regular and group invoices
    aging_labels = ['Current', '1-30 Days', '31-60 Days', '61-90 Days', '90+ Days']
    
    # Regular invoice aging
    regular_aging = [
        Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__gte=today).aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today, due_date__gte=today - timedelta(days=30)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=30), due_date__gte=today - timedelta(days=60)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=60), due_date__gte=today - timedelta(days=90)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=90)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
    ]
    
    # Group invoice aging
    group_aging = [
        GroupInvoice.objects.filter(status__in=['sent', 'overdue'], due_date__gte=today).aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        GroupInvoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today, due_date__gte=today - timedelta(days=30)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        GroupInvoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=30), due_date__gte=today - timedelta(days=60)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        GroupInvoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=60), due_date__gte=today - timedelta(days=90)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        GroupInvoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=90)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
    ]
    
    # Combine aging data
    aging_data = [float(regular_aging[i] + group_aging[i]) for i in range(5)]

    # Payment method doughnut chart data
    method_labels = []
    method_data = []
    for entry in payment_methods_data:
        label = dict(Payment.PAYMENT_METHODS).get(entry['payment_method'], entry['payment_method'])
        method_labels.append(label)
        method_data.append(float(entry['total']))
    
    # Add group payment methods
    for entry in group_payment_methods:
        label = dict(GroupPayment.PAYMENT_METHODS).get(entry['payment_method'], entry['payment_method'])
        if label not in method_labels:
            method_labels.append(label + ' (Group)')
            method_data.append(float(entry['total']))
        else:
            idx = method_labels.index(label)
            method_data[idx] += float(entry['total'])

    # Recent audit log entries
    recent_audit = BillingAuditLog.objects.select_related('performed_by', 'invoice', 'payment').order_by('-timestamp')[:10]

    # ---- SERVICE STATISTICS ----
    from django.db.models import Avg, Max, Min

    # All-time top services by total revenue (includes both regular and group invoices)
    svc_alltime = (
        InvoiceLineItem.objects
        .values('service__name', 'description')
        .annotate(
            total_revenue=Sum('total_amount'),
            times_billed=Count('id'),
            avg_price=Avg('unit_price'),
            max_price=Max('unit_price'),
            min_price=Min('unit_price'),
        )
        .order_by('-total_revenue')[:12]
    )

    # This month's top services by revenue
    svc_thismonth = (
        InvoiceLineItem.objects
        .filter(invoice__issue_date__gte=current_month)
        .values('service__name', 'description')
        .annotate(
            total_revenue=Sum('total_amount'),
            times_billed=Count('id'),
        )
        .order_by('-total_revenue')[:10]
    )

    # Grand total for % share calculation
    svc_grand_total = sum(float(s['total_revenue'] or 0) for s in svc_alltime)

    # Service stats with % share
    svc_stats = []
    for s in svc_alltime:
        rev = float(s['total_revenue'] or 0)
        svc_stats.append({
            'name': s['service__name'] or s['description'] or '—',
            'total_revenue': rev,
            'times_billed': s['times_billed'],
            'avg_price': float(s['avg_price'] or 0),
            'max_price': float(s['max_price'] or 0),
            'min_price': float(s['min_price'] or 0),
            'pct': round((rev / svc_grand_total * 100), 1) if svc_grand_total else 0,
        })

    # Top 3 highlights
    top_revenue_svc = svc_stats[0] if svc_stats else None
    top_billed_svc = sorted(svc_stats, key=lambda x: x['times_billed'], reverse=True)[0] if svc_stats else None
    top_avgval_svc = sorted(svc_stats, key=lambda x: x['avg_price'], reverse=True)[0] if svc_stats else None

    # Service revenue horizontal bar chart data (top 10 all-time)
    svc_bar_labels = json.dumps([s['name'][:30] for s in svc_stats[:10]])
    svc_bar_data = json.dumps([s['total_revenue'] for s in svc_stats[:10]])
    svc_bar_counts = json.dumps([s['times_billed'] for s in svc_stats[:10]])

    # Top 5 services trend over last 6 months (multi-line chart)
    top5_names = [s['name'] for s in svc_stats[:5]]
    svc_trend_labels = []
    svc_trend_datasets = {n: [] for n in top5_names}

    for i in range(5, -1, -1):
        month_date = (current_month - timedelta(days=32 * i)).replace(day=1)
        next_month_date = (month_date + timedelta(days=32)).replace(day=1)
        svc_trend_labels.append(month_date.strftime('%b %Y'))
        for svc_name in top5_names:
            rev = (
                InvoiceLineItem.objects
                .filter(
                    invoice__issue_date__gte=month_date,
                    invoice__issue_date__lt=next_month_date,
                )
                .filter(
                    Q(service__name=svc_name) | Q(description=svc_name)
                )
                .aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            )
            svc_trend_datasets[svc_name].append(float(rev))

    trend_colors = ['#4e73df', '#1cc88a', '#36b9cc', '#f6c23e', '#e74a3b']
    svc_trend_chart_datasets = json.dumps([
        {
            'label': name,
            'data': svc_trend_datasets[name],
            'borderColor': trend_colors[i % len(trend_colors)],
            'backgroundColor': trend_colors[i % len(trend_colors)] + '22',
            'borderWidth': 2,
            'fill': False,
            'tension': 0.3,
        }
        for i, name in enumerate(top5_names)
    ])
    svc_trend_labels_json = json.dumps(svc_trend_labels)

    # Monthly total billed units per service (this month)
    svc_thismonth_display = []
    for s in svc_thismonth:
        svc_thismonth_display.append({
            'name': s['service__name'] or s['description'] or '—',
            'total_revenue': float(s['total_revenue'] or 0),
            'times_billed': s['times_billed'],
        })

    # Active patients for invoice creation
    active_patients = Patient.objects.filter(is_active=True).order_by('first_name', 'last_name')
    
    # Group invoice statistics for dashboard
    group_stats = {
        'total_group_invoices': GroupInvoice.objects.count(),
        'group_outstanding': group_outstanding,
        'group_overdue_count': group_overdue_count,
        'group_monthly_revenue': group_monthly_revenue,
    }
    
    context = {
        'monthly_revenue': total_monthly_revenue,
        'outstanding_amount': total_outstanding,
        'daily_payments': total_daily_payments,
        'overdue_count': total_overdue_count,
        'recent_invoices': all_invoices,
        'recent_group_invoices': recent_group_invoices,
        'payment_methods_data': payment_methods_data,
        'overdue_invoices': overdue_invoices,
        'overdue_group_invoices': overdue_group_invoices,
        'insurance_claims_summary': insurance_claims_summary,
        'revenue_chart_labels': json.dumps(revenue_chart_labels),
        'revenue_chart_data': json.dumps(revenue_chart_data),
        'aging_labels': json.dumps(aging_labels),
        'aging_data': json.dumps(aging_data),
        'method_labels': json.dumps(method_labels),
        'method_data': json.dumps(method_data),
        'recent_audit': recent_audit,
        'svc_stats': svc_stats,
        'svc_thismonth': svc_thismonth_display,
        'svc_grand_total': svc_grand_total,
        'top_revenue_svc': top_revenue_svc,
        'top_billed_svc': top_billed_svc,
        'top_avgval_svc': top_avgval_svc,
        'svc_bar_labels': svc_bar_labels,
        'svc_bar_data': svc_bar_data,
        'svc_bar_counts': svc_bar_counts,
        'svc_trend_labels': svc_trend_labels_json,
        'svc_trend_datasets': svc_trend_chart_datasets,
        'active_patients': active_patients,
        'modal_active_patients': active_patients,
        'group_stats': group_stats,
    }
    return render(request, 'billing/billing_dashboard.html', context)

@login_required
def invoice_create_for_patient(request):
    """Create a new invoice directly for a selected patient"""
    if request.method == 'POST':
        patient_id = request.POST.get('patient_id')
        if patient_id:
            patient = get_object_or_404(Patient, pk=patient_id)
            
            # Generate invoice number
            last_invoice = Invoice.objects.order_by('-id').first()
            if last_invoice and last_invoice.invoice_number:
                try:
                    last_number = int(last_invoice.invoice_number.split('-')[1])
                    invoice_number = f"INV-{last_number + 1:06d}"
                except (ValueError, IndexError):
                    invoice_count = Invoice.objects.count()
                    invoice_number = f"INV-{invoice_count + 1:06d}"
            else:
                invoice_number = "INV-000001"
            
            # Create draft invoice
            from django.utils import timezone
            due_date = timezone.now().date() + timedelta(days=30)
            
            invoice = Invoice.objects.create(
                invoice_number=invoice_number,
                patient=patient,
                due_date=due_date,
                status='draft',
                subtotal=0,
                tax_rate=0,
                tax_amount=0,
                discount_amount=0,
                total_amount=0,
                notes=f'Invoice created directly for {patient.get_full_name()} on {timezone.now().date()}',
                created_by=request.user,
            )
            
            messages.success(request, f'Draft invoice {invoice_number} created for {patient.get_full_name()}. You can now add services.')
            
            # Handle AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Draft invoice {invoice_number} created successfully!',
                    'invoice_number': invoice_number,
                    'invoice_id': invoice.pk,
                    'redirect_url': f'/billing/invoices/{invoice.pk}/edit/'
                })
            
            return redirect('billing:invoice_edit', pk=invoice.pk)
        else:
            messages.error(request, 'Please select a patient.')
            
            # Handle AJAX request for error
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'Please select a patient.'
                })
    
    return redirect('billing:billing_dashboard')

@login_required
def invoice_create_ajax(request):
    """AJAX-only endpoint to create a new invoice header via modal.
    Returns JSON and rejects non-AJAX GETs."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    post_data = request.POST.copy()
    if not post_data.get('tax_rate'):
        post_data['tax_rate'] = '0'
    if not post_data.get('discount_amount'):
        post_data['discount_amount'] = '0'
    form = InvoiceForm(post_data)
    if form.is_valid():
        try:
            with transaction.atomic():
                invoice = form.save(commit=False)
                invoice.created_by = request.user

                # Generate invoice number
                last_invoice = Invoice.objects.order_by('-id').first()
                if last_invoice and last_invoice.invoice_number:
                    try:
                        last_number = int(last_invoice.invoice_number.split('-')[1])
                        invoice.invoice_number = f"INV-{last_number + 1:06d}"
                    except (ValueError, IndexError):
                        invoice.invoice_number = f"INV-{Invoice.objects.count() + 1:06d}"
                else:
                    invoice.invoice_number = "INV-000001"

                invoice.subtotal = invoice.subtotal or 0
                invoice.tax_rate = invoice.tax_rate or 0
                invoice.tax_amount = invoice.tax_amount or 0
                invoice.discount_amount = invoice.discount_amount or 0
                invoice.total_amount = invoice.total_amount or 0
                invoice.status = invoice.status or 'draft'

                invoice.save()

                # No line items handled here; follow-up happens on edit page

                # Respond based on request type
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'invoice_id': invoice.pk,
                        'invoice_number': invoice.invoice_number,
                        'redirect_url': "/billing/invoices/"
                    })
                else:
                    return redirect('billing:invoice_edit', pk=invoice.pk)
        except Exception as e:
            return JsonResponse({'success': False, 'errors': {'__all__': [str(e)]}}, status=400)

    # Return validation errors
    return JsonResponse({'success': False, 'errors': form.errors}, status=400)

@login_required
def invoice_create_full_ajax(request):
    """Create invoice header and multiple line items in one AJAX POST."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    post_data = request.POST.copy()
    if not post_data.get('tax_rate'):
        post_data['tax_rate'] = '0'
    if not post_data.get('discount_amount'):
        post_data['discount_amount'] = '0'

    header_form = InvoiceForm(post_data)
    if not header_form.is_valid():
        return JsonResponse({'success': False, 'errors': header_form.errors}, status=400)

    # Parse dynamic line items from fields: items-<idx>-service, description, quantity, unit_price
    items = []
    idx = 0
    while True:
        svc = post_data.get(f'items-{idx}-service')
        desc = post_data.get(f'items-{idx}-description')
        qty = post_data.get(f'items-{idx}-quantity')
        price = post_data.get(f'items-{idx}-unit_price')
        if svc is None and desc is None and qty is None and price is None:
            break
        # Only add if service or description is provided
        if (svc or desc) and qty and price:
            items.append({
                'service': svc,
                'description': desc or '',
                'quantity': qty,
                'unit_price': price,
            })
        idx += 1

    if not items:
        return JsonResponse({'success': False, 'errors': {'__all__': ['Please add at least one line item.']}}, status=400)

    try:
        with transaction.atomic():
            invoice = header_form.save(commit=False)
            invoice.created_by = request.user

            # Generate invoice number
            last_invoice = Invoice.objects.order_by('-id').first()
            if last_invoice and last_invoice.invoice_number:
                try:
                    last_number = int(last_invoice.invoice_number.split('-')[1])
                    invoice.invoice_number = f"INV-{last_number + 1:06d}"
                except (ValueError, IndexError):
                    invoice.invoice_number = f"INV-{Invoice.objects.count() + 1:06d}"
            else:
                invoice.invoice_number = "INV-000001"

            # Ensure numeric defaults
            invoice.subtotal = 0
            invoice.tax_amount = 0
            invoice.total_amount = 0
            if not invoice.status:
                invoice.status = 'draft'
            invoice.save()

            # Create line items with group-based pricing
            patient = invoice.patient
            for it in items:
                service = None
                lab_test = None
                unit_price = float(it['unit_price'])
                
                # Check if it's a service
                if it.get('service'):
                    try:
                        service = Service.objects.get(pk=it['service'])
                        # Apply group pricing if available
                        group_price = float(ServicePriceGroup.get_price_for_patient(service, patient))
                        # Use group price only if the submitted price matches the default
                        # (i.e. user didn't manually override)
                        if abs(unit_price - float(service.base_price)) < 0.01:
                            unit_price = group_price
                    except Service.DoesNotExist:
                        service = None
                
                # Check if it's a lab test (by matching description or checking for lab_test_id)
                if it.get('lab_test_id'):
                    try:
                        lab_test = LabTest.objects.get(pk=it['lab_test_id'])
                        # Apply group pricing if available
                        group_price = float(LabTestPriceGroup.get_price_for_patient(lab_test, patient))
                        # Use group price only if the submitted price matches the default
                        if abs(unit_price - float(lab_test.price)) < 0.01:
                            unit_price = group_price
                    except LabTest.DoesNotExist:
                        lab_test = None
                
                InvoiceLineItem.objects.create(
                    invoice=invoice,
                    service=service,
                    description=it['description'],
                    quantity=int(float(it['quantity'])),
                    unit_price=unit_price
                )

            # Calculate totals
            invoice.calculate_totals()

            return JsonResponse({
                'success': True,
                'invoice_id': invoice.pk,
                'invoice_number': invoice.invoice_number,
                'redirect_url': "/billing/invoices/"
            })
    except Exception as e:
        return JsonResponse({'success': False, 'errors': {'__all__': [str(e)]}}, status=400)

@login_required
def invoice_list(request):
    from .models import GroupInvoice
    
    # Get all invoices for statistics (before filtering)
    all_invoices = Invoice.objects.select_related('patient').all()
    
    # Calculate invoice statistics
    from django.db.models import Sum, Count, Q
    from datetime import date
    
    stats = {
        'total_invoices': all_invoices.count(),
        'total_amount': all_invoices.aggregate(total=Sum('total_amount'))['total'] or 0,
        'draft_count': all_invoices.filter(status='draft').count(),
        'sent_count': all_invoices.filter(status='sent').count(),
        'paid_count': all_invoices.filter(status='paid').count(),
        'overdue_count': all_invoices.filter(status='overdue').count(),
        'cancelled_count': all_invoices.filter(status='cancelled').count(),
    }
    
    # Calculate paid and outstanding amounts
    paid_amount = 0
    outstanding_amount = 0
    
    for invoice in all_invoices:
        if invoice.status == 'paid':
            paid_amount += invoice.total_amount
        elif invoice.status in ['sent', 'overdue']:
            balance_due = invoice.get_balance_due()
            outstanding_amount += balance_due
    
    stats['paid_amount'] = paid_amount
    stats['outstanding_amount'] = outstanding_amount
    
    # Group invoice statistics
    group_invoices = GroupInvoice.objects.all()
    group_stats = {
        'total_group_invoices': group_invoices.count(),
        'group_total_amount': group_invoices.aggregate(total=Sum('total_amount'))['total'] or 0,
        'group_draft_count': group_invoices.filter(status='draft').count(),
        'group_sent_count': group_invoices.filter(status='sent').count(),
        'group_paid_count': group_invoices.filter(status='paid').count(),
        'group_overdue_count': group_invoices.filter(status='overdue').count(),
        'group_outstanding': group_invoices.filter(status__in=['sent', 'overdue']).aggregate(
            total=Sum('total_amount'))['total'] or 0,
    }
    
    # Combined statistics
    combined_stats = {
        'total_all_invoices': stats['total_invoices'] + group_stats['total_group_invoices'],
        'total_all_amount': stats['total_amount'] + group_stats['group_total_amount'],
        'total_outstanding': stats['outstanding_amount'] + group_stats['group_outstanding'],
        'total_overdue': stats['overdue_count'] + group_stats['group_overdue_count'],
    }
    
    # Now filter invoices for the list display
    invoices = all_invoices
    
    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        invoices = invoices.filter(status=status_filter)
    
    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        invoices = invoices.filter(
            Q(invoice_number__icontains=search_query) |
            Q(patient__first_name__icontains=search_query) |
            Q(patient__last_name__icontains=search_query)
        )

    # Date range filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        invoices = invoices.filter(issue_date__gte=date_from)
    if date_to:
        invoices = invoices.filter(issue_date__lte=date_to)

    # Amount filter
    amount_min = request.GET.get('amount_min')
    amount_max = request.GET.get('amount_max')
    if amount_min:
        try:
            invoices = invoices.filter(total_amount__gte=float(amount_min))
        except ValueError:
            pass
    if amount_max:
        try:
            invoices = invoices.filter(total_amount__lte=float(amount_max))
        except ValueError:
            pass

    # Paginate with dynamic page size
    pagination_data = paginate_queryset(request, invoices, default_page_size=25)
    
    context = {
        'page_obj': pagination_data['page_obj'],
        'invoices': pagination_data['items'],
        'status_choices': Invoice.STATUS_CHOICES,
        'selected_status': status_filter,
        'search_query': search_query,
        'date_from': date_from or '',
        'date_to': date_to or '',
        'amount_min': amount_min or '',
        'amount_max': amount_max or '',
        'page_size': pagination_data['page_size'],
        'query_string': pagination_data['query_string'],
        'stats': stats,
        'group_stats': group_stats,
        'combined_stats': combined_stats,
        'total_amount': stats['total_amount'],
        'total_paid': stats['paid_amount'],
        'total_outstanding': stats['outstanding_amount'],
        'active_patients': Patient.objects.filter(is_active=True).order_by('first_name', 'last_name'),
        'modal_active_patients': Patient.objects.filter(is_active=True).order_by('first_name', 'last_name'),
        'services': Service.objects.all().order_by('name'),
    }
    return render(request, 'billing/invoice_list.html', context)

@login_required
def invoice_create(request):
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        formset = InvoiceLineItemFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                invoice = form.save(commit=False)
                invoice.created_by = request.user
                
                # Generate invoice number
                last_invoice = Invoice.objects.order_by('-id').first()
                if last_invoice:
                    last_number = int(last_invoice.invoice_number.split('-')[1])
                    invoice.invoice_number = f"INV-{last_number + 1:06d}"
                else:
                    invoice.invoice_number = "INV-000001"
                
                invoice.save()
                
                # Save line items
                formset.instance = invoice
                formset.save()
                
                # Calculate totals
                invoice.calculate_totals()
                
                messages.success(request, f'Invoice {invoice.invoice_number} created successfully!')
                return redirect('billing:invoice_detail', pk=invoice.pk)
    else:
        form = InvoiceForm()
        formset = InvoiceLineItemFormSet()
    
    context = {
        'form': form,
        'formset': formset,
        'services': Service.objects.all(),
    }
    return render(request, 'billing/invoice_create.html', context)

@login_required
def invoice_detail(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    line_items = invoice.line_items.all()
    payments = invoice.payments.all()
    
    context = {
        'invoice': invoice,
        'line_items': line_items,
        'payments': payments,
    }
    return render(request, 'billing/invoice_detail.html', context)

@login_required
def invoice_pdf(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    line_items = invoice.line_items.all()
    payments = invoice.payments.all()

    import qrcode
    from io import BytesIO
    import base64

    from clinic_settings.models import ClinicSettings
    try:
        clinic_settings = ClinicSettings.objects.first()
    except Exception:
        clinic_settings = None

    # Build R2 public URL for QR code (same pattern as publish view)
    import re
    from django.conf import settings as _settings
    patient_name_raw = invoice.patient.get_full_name() or str(invoice.patient)
    safe_name = re.sub(r'[^\w]', '_', patient_name_raw).strip('_') or f'Patient_{pk}'
    r2_public = getattr(_settings, 'R2_PUBLIC_URL', '').rstrip('/')
    r2_qr_url = f'{r2_public}/invoices/{safe_name}_invoice_{invoice.invoice_number}.pdf' if r2_public else ''

    qr_url = (r2_qr_url or invoice.gdrive_pdf_url or invoice.invoice_pdf_url
              or request.build_absolute_uri(request.path))
    qr = qrcode.QRCode(version=1, box_size=10, border=2)
    qr.add_data(qr_url)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    qr_img.save(buffer, format='PNG')
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()
    
    context = {
        'invoice': invoice,
        'line_items': line_items,
        'payments': payments,
        'clinic_settings': clinic_settings,
        'qr_code': qr_base64,
        'cloud_url': invoice.invoice_pdf_url,
        'gdrive_url': invoice.gdrive_pdf_url,
    }
    return render(request, 'billing/invoice_pdf.html', context)

@login_required
def payment_create(request, invoice_pk=None):
    invoice = None
    if invoice_pk:
        invoice = get_object_or_404(Invoice, pk=invoice_pk)
    elif request.GET.get('invoice'):
        # Handle query parameter case
        invoice = get_object_or_404(Invoice, pk=request.GET.get('invoice'))
    
    # Check if invoice is already fully paid
    if invoice and invoice.status == 'paid':
        messages.warning(request, f'Invoice {invoice.invoice_number} is already fully paid. No additional payment needed.')
        
        # Handle AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'already_paid': True,
                'message': f'Invoice {invoice.invoice_number} is already fully paid.',
                'invoice_number': invoice.invoice_number,
                'total_amount': float(invoice.total_amount),
                'redirect_url': f'/billing/invoices/{invoice.pk}/'
            })
        
        return redirect('billing:invoice_detail', pk=invoice.pk)
    
    if request.method == 'POST':
        # Create a mutable copy of POST data and ensure invoice/patient are set
        post_data = request.POST.copy()
        if invoice:
            post_data['patient'] = str(invoice.patient.pk)
            post_data['invoice'] = str(invoice.pk)
        
        form = PaymentForm(post_data, invoice=invoice)
        if form.is_valid():
            try:
                with transaction.atomic():
                    payment = form.save(commit=False)
                    payment.processed_by = request.user
                    
                    # Generate payment ID
                    last_payment = Payment.objects.order_by('-id').first()
                    if last_payment and last_payment.payment_id:
                        try:
                            last_number = int(last_payment.payment_id.split('-')[1])
                            payment.payment_id = f"PAY-{last_number + 1:06d}"
                        except (ValueError, IndexError):
                            payment_count = Payment.objects.count()
                            payment.payment_id = f"PAY-{payment_count + 1:06d}"
                    else:
                        payment.payment_id = "PAY-000001"
                    
                    payment.save()
                    
                    # Update invoice status if fully paid and invoice exists
                    if payment.invoice:
                        total_payments = payment.invoice.payments.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
                        if total_payments >= payment.invoice.total_amount:
                            payment.invoice.status = 'paid'
                            payment.invoice.save()
                            msg = f'Payment {payment.payment_id} recorded successfully! Invoice is now fully paid.'
                            messages.success(request, msg)
                        else:
                            remaining = payment.invoice.total_amount - total_payments
                            msg = f'Payment {payment.payment_id} recorded. Remaining balance: UGX {remaining:,.0f}'
                            messages.success(request, msg)

                        # Overpayment warning
                        if total_payments > payment.invoice.total_amount:
                            excess = total_payments - payment.invoice.total_amount
                            messages.warning(request, f'Overpayment detected! UGX {excess:,.0f} excess paid for invoice {payment.invoice.invoice_number}.')

                        # Audit log
                        BillingAuditLog.objects.create(
                            action='payment_created',
                            performed_by=request.user,
                            invoice=payment.invoice,
                            payment=payment,
                            details=f'Payment {payment.payment_id} of UGX {payment.amount:,} recorded via {payment.get_payment_method_display()}.',
                        )

                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({
                                'success': True,
                                'message': msg,
                                'payment_id': payment.payment_id,
                                'amount': float(payment.amount),
                                'balance_due': float(payment.invoice.get_balance_due()),
                                'invoice_status': payment.invoice.status,
                            })
                        return redirect('billing:invoice_detail', pk=payment.invoice.pk)
                    else:
                        messages.success(request, f'Payment {payment.payment_id} recorded successfully!')
                        # Standalone payment audit log
                        BillingAuditLog.objects.create(
                            action='payment_created',
                            performed_by=request.user,
                            payment=payment,
                            details=f'Standalone payment {payment.payment_id} of UGX {payment.amount:,} recorded.',
                        )
                        return redirect('billing:payment_list')
            except Exception as e:
                messages.error(request, f'Error creating payment: {str(e)}')
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'errors': {'__all__': [str(e)]}
                    })
        else:
            # Add debug information for form errors
            messages.error(request, 'Please correct the errors below.')
            # Handle AJAX form errors
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
    else:
        initial_data = {}
        if invoice:
            balance_due = invoice.get_balance_due()
            initial_data['amount'] = balance_due if balance_due > 0 else invoice.total_amount
            initial_data['status'] = 'completed'
        form = PaymentForm(initial=initial_data, invoice=invoice)
    
    context = {
        'form': form,
        'invoice': invoice,
    }
    return render(request, 'billing/payment_create.html', context)

@login_required
def payment_detail(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    
    context = {
        'payment': payment,
    }
    return render(request, 'billing/payment_detail.html', context)

@login_required
def payment_receipt(request, pk):
    payment = get_object_or_404(Payment, pk=pk)

    from clinic_settings.models import ClinicSettings
    import qrcode
    from io import BytesIO
    import base64

    try:
        clinic_settings = ClinicSettings.objects.first()
    except Exception:
        clinic_settings = None

    # Build R2 public URL for QR code (same pattern as publish view)
    import re
    from django.conf import settings as _settings
    patient_name_raw = payment.patient.get_full_name() or str(payment.patient)
    safe_name = re.sub(r'[^\w]', '_', patient_name_raw).strip('_') or f'Patient_{pk}'
    r2_public = getattr(_settings, 'R2_PUBLIC_URL', '').rstrip('/')
    r2_qr_url = f'{r2_public}/receipts/{safe_name}_receipt_{payment.payment_id}.pdf' if r2_public else ''

    qr_url = (r2_qr_url or payment.gdrive_pdf_url or payment.receipt_pdf_url
              or request.build_absolute_uri(request.path))
    qr = qrcode.QRCode(version=1, box_size=10, border=2)
    qr.add_data(qr_url)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    qr_img.save(buffer, format='PNG')
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()
    
    context = {
        'payment': payment,
        'clinic_settings': clinic_settings,
        'qr_code': qr_base64,
        'cloud_url': payment.receipt_pdf_url,
        'gdrive_url': payment.gdrive_pdf_url,
    }
    return render(request, 'billing/payment_receipt.html', context)

@login_required
def payment_list(request):
    payments = Payment.objects.select_related('patient', 'invoice').all()
    
    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        payments = payments.filter(status=status_filter)
    
    # Filter by payment method
    method_filter = request.GET.get('method')
    if method_filter:
        payments = payments.filter(payment_method=method_filter)

    # Date range filter
    pay_date_from = request.GET.get('date_from')
    pay_date_to = request.GET.get('date_to')
    if pay_date_from:
        payments = payments.filter(payment_date__date__gte=pay_date_from)
    if pay_date_to:
        payments = payments.filter(payment_date__date__lte=pay_date_to)

    # Patient search filter
    pay_patient = request.GET.get('patient_search')
    if pay_patient:
        payments = payments.filter(
            Q(patient__first_name__icontains=pay_patient) |
            Q(patient__last_name__icontains=pay_patient)
        )

    # Amount filter
    pay_amount_min = request.GET.get('amount_min')
    if pay_amount_min:
        try:
            payments = payments.filter(amount__gte=float(pay_amount_min))
        except ValueError:
            pass

    # Summary stats
    from django.db.models import Sum
    payment_total = payments.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0

    # Paginate with dynamic page size
    pagination_data = paginate_queryset(request, payments, default_page_size=25)
    
    context = {
        'page_obj': pagination_data['page_obj'],
        'payments': pagination_data['items'],
        'status_choices': Payment.PAYMENT_STATUS,
        'method_choices': Payment.PAYMENT_METHODS,
        'selected_status': status_filter,
        'selected_method': method_filter,
        'date_from': pay_date_from or '',
        'date_to': pay_date_to or '',
        'patient_search': pay_patient or '',
        'amount_min': pay_amount_min or '',
        'payment_total': payment_total,
        'page_size': pagination_data['page_size'],
        'query_string': pagination_data['query_string'],
        'modal_active_patients': Patient.objects.filter(is_active=True).order_by('first_name', 'last_name'),
    }
    return render(request, 'billing/payment_list.html', context)

@login_required
def insurance_claim_list(request):
    claims = InsuranceClaim.objects.select_related('patient', 'invoice').all()
    
    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        claims = claims.filter(status=status_filter)
    
    # Filter by provider
    provider_filter = request.GET.get('provider')
    if provider_filter:
        claims = claims.filter(insurance_provider__icontains=provider_filter)
    
    # Filter by date range
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        claims = claims.filter(submission_date__gte=date_from)
    if date_to:
        claims = claims.filter(submission_date__lte=date_to)
    
    # Calculate summary statistics
    approved_count = claims.filter(status='approved').count()
    pending_count = claims.filter(status__in=['submitted', 'pending']).count()
    total_claim_amount = claims.aggregate(Sum('claim_amount'))['claim_amount__sum'] or 0
    
    paginator = Paginator(claims, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'status_choices': InsuranceClaim.CLAIM_STATUS,
        'selected_status': status_filter,
        'approved_count': approved_count,
        'pending_count': pending_count,
        'total_claim_amount': total_claim_amount,
    }
    return render(request, 'billing/insurance_claim_list.html', context)

@login_required
def insurance_claim_create(request, invoice_pk=None):
    invoice = None
    if invoice_pk:
        invoice = get_object_or_404(Invoice, pk=invoice_pk)
    
    if request.method == 'POST':
        form = InsuranceClaimForm(request.POST)
        if form.is_valid():
            claim = form.save(commit=False)
            if invoice:
                claim.invoice = invoice
                claim.patient = invoice.patient
            claim.submitted_by = request.user
            
            # Generate claim number
            last_claim = InsuranceClaim.objects.order_by('-id').first()
            if last_claim:
                last_number = int(last_claim.claim_number.split('-')[1])
                claim.claim_number = f"CLM-{last_number + 1:06d}"
            else:
                claim.claim_number = "CLM-000001"
            
            claim.save()
            messages.success(request, f'Insurance claim {claim.claim_number} submitted successfully!')
            return redirect('billing:insurance_claim_list')
    else:
        initial_data = {}
        if invoice:
            # Pre-populate with patient's insurance information
            initial_data = {
                'insurance_provider': invoice.patient.insurance_provider,
                'policy_number': invoice.patient.insurance_policy_number,
                'group_number': invoice.patient.insurance_group_number,
                'claim_amount': invoice.total_amount,
            }
        form = InsuranceClaimForm(initial=initial_data)
    
    context = {
        'form': form,
        'invoice': invoice,
    }
    return render(request, 'billing/insurance_claim_create.html', context)

@login_required
def insurance_claim_print(request, pk):
    """Print view for insurance claim"""
    from clinic_settings.models import ClinicSettings
    
    claim = get_object_or_404(InsuranceClaim, pk=pk)
    
    # Get clinic settings for logo
    try:
        clinic_settings = ClinicSettings.objects.first()
    except:
        clinic_settings = None
    
    context = {
        'claim': claim,
        'patient': claim.patient,
        'invoice': claim.invoice,
        'clinic_settings': clinic_settings,
    }
    return render(request, 'billing/insurance_claim_print.html', context)

@login_required
def payment_plan_list(request):
    payment_plans = PaymentPlan.objects.select_related('patient', 'invoice').all()
    
    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        payment_plans = payment_plans.filter(status=status_filter)
    
    paginator = Paginator(payment_plans, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'status_choices': PaymentPlan.PLAN_STATUS,
        'selected_status': status_filter,
    }
    return render(request, 'billing/payment_plan_list.html', context)

@login_required
def payment_plan_create(request, invoice_pk):
    invoice = get_object_or_404(Invoice, pk=invoice_pk)
    
    if request.method == 'POST':
        form = PaymentPlanForm(request.POST)
        if form.is_valid():
            payment_plan = form.save(commit=False)
            payment_plan.invoice = invoice
            payment_plan.patient = invoice.patient
            payment_plan.created_by = request.user
            
            # Generate payment plan ID
            last_plan = PaymentPlan.objects.order_by('-id').first()
            if last_plan:
                last_number = int(last_plan.plan_id.split('-')[1])
                payment_plan.plan_id = f"PP-{last_number + 1:06d}"
            else:
                payment_plan.plan_id = "PP-000001"
            
            payment_plan.save()
            messages.success(request, f'Payment plan {payment_plan.plan_id} created successfully!')
            return redirect('billing:payment_plan_detail', pk=payment_plan.pk)
    else:
        form = PaymentPlanForm(initial={'total_amount': invoice.total_amount})
    
    context = {
        'form': form,
        'invoice': invoice,
    }
    return render(request, 'billing/payment_plan_create.html', context)

@login_required
def payment_plan_detail(request, pk):
    payment_plan = get_object_or_404(PaymentPlan, pk=pk)
    
    context = {
        'payment_plan': payment_plan,
    }
    return render(request, 'billing/payment_plan_detail.html', context)

@login_required
def invoice_status_update(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(Invoice.STATUS_CHOICES):
            invoice.status = new_status
            invoice.save()
            messages.success(request, f'Invoice status updated to {invoice.get_status_display()}')
        else:
            messages.error(request, 'Invalid status selected')
    
    return redirect('billing:invoice_detail', pk=pk)

@login_required
def bulk_invoice_action(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        invoice_ids = request.POST.getlist('invoice_ids')
        
        if not invoice_ids:
            messages.error(request, 'No invoices selected')
            return redirect('billing:invoice_list')
        
        invoices = Invoice.objects.filter(id__in=invoice_ids)
        
        if action == 'mark_sent':
            invoices.update(status='sent')
            messages.success(request, f'{len(invoice_ids)} invoices marked as sent')
        elif action == 'mark_paid':
            invoices.update(status='paid')
            messages.success(request, f'{len(invoice_ids)} invoices marked as paid')
        elif action == 'mark_overdue':
            invoices.update(status='overdue')
            messages.success(request, f'{len(invoice_ids)} invoices marked as overdue')
        else:
            messages.error(request, 'Invalid action selected')
    
    return redirect('billing:invoice_list')

@login_required
def payment_receipt_pdf(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    
    # Header
    title = Paragraph(f"PAYMENT RECEIPT", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Receipt details
    receipt_data = [
        ['Payment ID:', payment.payment_id],
        ['Payment Date:', payment.payment_date.strftime('%Y-%m-%d %H:%M')],
        ['Amount:', f"${payment.amount}"],
        ['Method:', payment.get_payment_method_display()],
        ['Status:', payment.get_status_display()],
    ]
    
    if payment.reference_number:
        receipt_data.append(['Reference:', payment.reference_number])
    
    receipt_table = Table(receipt_data, colWidths=[2*72, 3*72])
    receipt_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    story.append(receipt_table)
    story.append(Spacer(1, 12))
    
    # Patient and invoice information
    patient_info = Paragraph(f"<b>Patient:</b><br/>{payment.patient.get_full_name()}<br/><b>Invoice:</b> {payment.invoice.invoice_number}", styles['Normal'])
    story.append(patient_info)
    
    doc.build(story)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="receipt_{payment.payment_id}.pdf"'
    return response

@login_required
def invoice_aging_report(request):
    from .models import GroupInvoice
    
    # Calculate aging buckets
    today = date.today()
    
    # Regular invoices
    current_invoices = Invoice.objects.filter(
        status__in=['sent', 'overdue'],
        due_date__gte=today
    )
    
    overdue_1_30 = Invoice.objects.filter(
        status__in=['sent', 'overdue'],
        due_date__lt=today,
        due_date__gte=today - timedelta(days=30)
    )
    
    overdue_31_60 = Invoice.objects.filter(
        status__in=['sent', 'overdue'],
        due_date__lt=today - timedelta(days=30),
        due_date__gte=today - timedelta(days=60)
    )
    
    overdue_61_90 = Invoice.objects.filter(
        status__in=['sent', 'overdue'],
        due_date__lt=today - timedelta(days=60),
        due_date__gte=today - timedelta(days=90)
    )
    
    overdue_90_plus = Invoice.objects.filter(
        status__in=['sent', 'overdue'],
        due_date__lt=today - timedelta(days=90)
    )
    
    # Group invoices
    group_current = GroupInvoice.objects.filter(
        status__in=['sent', 'overdue'],
        due_date__gte=today
    )
    
    group_1_30 = GroupInvoice.objects.filter(
        status__in=['sent', 'overdue'],
        due_date__lt=today,
        due_date__gte=today - timedelta(days=30)
    )
    
    group_31_60 = GroupInvoice.objects.filter(
        status__in=['sent', 'overdue'],
        due_date__lt=today - timedelta(days=30),
        due_date__gte=today - timedelta(days=60)
    )
    
    group_61_90 = GroupInvoice.objects.filter(
        status__in=['sent', 'overdue'],
        due_date__lt=today - timedelta(days=60),
        due_date__gte=today - timedelta(days=90)
    )
    
    group_90_plus = GroupInvoice.objects.filter(
        status__in=['sent', 'overdue'],
        due_date__lt=today - timedelta(days=90)
    )
    
    # Combined totals
    current_total = (current_invoices.aggregate(Sum('total_amount'))['total_amount__sum'] or 0) + \
                    (group_current.aggregate(Sum('total_amount'))['total_amount__sum'] or 0)
    overdue_1_30_total = (overdue_1_30.aggregate(Sum('total_amount'))['total_amount__sum'] or 0) + \
                         (group_1_30.aggregate(Sum('total_amount'))['total_amount__sum'] or 0)
    overdue_31_60_total = (overdue_31_60.aggregate(Sum('total_amount'))['total_amount__sum'] or 0) + \
                          (group_31_60.aggregate(Sum('total_amount'))['total_amount__sum'] or 0)
    overdue_61_90_total = (overdue_61_90.aggregate(Sum('total_amount'))['total_amount__sum'] or 0) + \
                          (group_61_90.aggregate(Sum('total_amount'))['total_amount__sum'] or 0)
    overdue_90_plus_total = (overdue_90_plus.aggregate(Sum('total_amount'))['total_amount__sum'] or 0) + \
                            (group_90_plus.aggregate(Sum('total_amount'))['total_amount__sum'] or 0)
    
    context = {
        'current_invoices': current_invoices,
        'overdue_1_30': overdue_1_30,
        'overdue_31_60': overdue_31_60,
        'overdue_61_90': overdue_61_90,
        'overdue_90_plus': overdue_90_plus,
        'group_current': group_current,
        'group_1_30': group_1_30,
        'group_31_60': group_31_60,
        'group_61_90': group_61_90,
        'group_90_plus': group_90_plus,
        'current_total': current_total,
        'overdue_1_30_total': overdue_1_30_total,
        'overdue_31_60_total': overdue_31_60_total,
        'overdue_61_90_total': overdue_61_90_total,
        'overdue_90_plus_total': overdue_90_plus_total,
    }
    return render(request, 'billing/invoice_aging_report.html', context)

@login_required
def get_service_price(request):
    """AJAX endpoint to get service price, with group-based pricing support.
    Pass ?service_id=X&patient_id=Y to get the group-specific price."""
    service_id = request.GET.get('service_id')
    patient_id = request.GET.get('patient_id')
    if service_id:
        try:
            service = Service.objects.get(id=service_id)
            # Try group-based pricing if patient_id is provided
            price = float(service.base_price)
            if patient_id:
                try:
                    patient = Patient.objects.get(pk=patient_id)
                    price = float(ServicePriceGroup.get_price_for_patient(service, patient))
                except Patient.DoesNotExist:
                    pass
            return JsonResponse({
                'price': price,
                'description': service.name
            })
        except Service.DoesNotExist:
            pass
    
    return JsonResponse({'price': 0, 'description': ''})

@login_required
def payment_debug(request):
    """Debug view to test payment creation"""
    from django.http import HttpResponse
    import json
    
    debug_info = {
        'user': str(request.user),
        'user_authenticated': request.user.is_authenticated,
        'payment_count': Payment.objects.count(),
        'invoice_count': Invoice.objects.count(),
        'patient_count': Patient.objects.count(),
        'available_invoices': list(Invoice.objects.filter(status__in=['draft', 'sent', 'overdue']).values('id', 'invoice_number', 'status')),
        'recent_payments': list(Payment.objects.order_by('-id')[:5].values('payment_id', 'amount', 'status')),
    }
    
    return HttpResponse(json.dumps(debug_info, indent=2), content_type='application/json')

@login_required
def payment_test(request):
    """Test view for payment creation debugging"""
    context = {
        'patients': Patient.objects.filter(is_active=True)[:10],
        'invoices': Invoice.objects.filter(status__in=['draft', 'sent', 'overdue'])[:10],
        'payments': Payment.objects.order_by('-id')[:5],
    }
    return render(request, 'billing/payment_test.html', context)

@login_required
def invoice_edit(request, pk):
    """Edit an existing invoice (especially useful for draft invoices)"""
    invoice = get_object_or_404(Invoice, pk=pk)
    
    if request.method == 'POST':
        form = InvoiceForm(request.POST, instance=invoice)
        formset = InvoiceLineItemFormSet(request.POST, instance=invoice)
        
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                invoice = form.save()
                formset.save()
                
                # Calculate totals
                invoice.calculate_totals()
                
                messages.success(request, f'Invoice {invoice.invoice_number} updated successfully!')
                return redirect('billing:invoice_detail', pk=invoice.pk)
    else:
        form = InvoiceForm(instance=invoice)
        formset = InvoiceLineItemFormSet(instance=invoice)
    
    context = {
        'form': form,
        'formset': formset,
        'invoice': invoice,
        'services': Service.objects.all(),
        'is_edit': True,
    }
    return render(request, 'billing/invoice_create.html', context)

@login_required
def patient_draft_invoices(request, patient_id):
    """View draft invoices for a specific patient"""
    patient = get_object_or_404(Patient, patient_id=patient_id)
    draft_invoices = Invoice.objects.filter(patient=patient, status='draft').order_by('-created_at')
    
    context = {
        'patient': patient,
        'draft_invoices': draft_invoices,
    }
    return render(request, 'billing/patient_draft_invoices.html', context)

@login_required
def send_invoice_email(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    
    if not invoice.patient.email:
        return JsonResponse({'success': False, 'error': 'Patient has no email address'})
    
    try:
        # Build absolute link to the printable invoice page (can be saved as PDF by recipient)
        pdf_link = request.build_absolute_uri(
            reverse('billing:invoice_pdf', kwargs={'pk': invoice.pk})
        )

        # Prepare email body with link
        clinic_name = getattr(settings, 'CLINIC_NAME', 'PhysioNutrition Clinic')
        subject = f"Invoice #{invoice.invoice_number} from {clinic_name}"
        message = render_to_string('billing/email/invoice_email.txt', {
            'invoice': invoice,
            'clinic_name': clinic_name,
            'pdf_link': pdf_link,
        })

        email = EmailMessage(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [invoice.patient.email],
        )
        email.send()
        
        # Update invoice status
        invoice.status = 'sent'
        invoice.save()
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ==================== AJAX-ONLY VIEWS ====================

@login_required
def payment_record_ajax(request):
    """AJAX-only payment recording view"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    # Accept both AJAX and regular form POSTs
    
    # Get invoice if provided
    invoice = None
    invoice_pk = request.POST.get('invoice') or request.GET.get('invoice')
    if invoice_pk:
        try:
            invoice = get_object_or_404(Invoice, pk=invoice_pk)
            
            # Check if invoice is already fully paid
            if invoice.status == 'paid':
                return JsonResponse({
                    'success': False,
                    'already_paid': True,
                    'message': f'Invoice {invoice.invoice_number} is already fully paid.',
                    'invoice_number': invoice.invoice_number,
                    'total_amount': float(invoice.total_amount),
                    'redirect_url': f'/billing/invoices/{invoice.pk}/'
                }, status=400)
        except:
            pass
    
    # Create a mutable copy of POST data
    post_data = request.POST.copy()
    if invoice:
        post_data['patient'] = str(invoice.patient.pk)
        post_data['invoice'] = str(invoice.pk)
    
    form = PaymentForm(post_data, invoice=invoice)
    if form.is_valid():
        try:
            with transaction.atomic():
                payment = form.save(commit=False)
                payment.processed_by = request.user
                
                # Generate payment ID
                last_payment = Payment.objects.order_by('-id').first()
                if last_payment and last_payment.payment_id:
                    try:
                        last_number = int(last_payment.payment_id.split('-')[1])
                        payment.payment_id = f"PAY-{last_number + 1:06d}"
                    except (ValueError, IndexError):
                        payment_count = Payment.objects.count()
                        payment.payment_id = f"PAY-{payment_count + 1:06d}"
                else:
                    payment.payment_id = "PAY-000001"
                
                payment.save()
                
                # Update invoice status if fully paid and invoice exists
                if payment.invoice:
                    total_payments = payment.invoice.payments.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
                    if total_payments >= payment.invoice.total_amount:
                        payment.invoice.status = 'paid'
                        payment.invoice.save()
                        message = f'Payment {payment.payment_id} recorded successfully! Invoice is now fully paid.'
                    else:
                        remaining = payment.invoice.total_amount - total_payments
                        message = f'Payment {payment.payment_id} recorded successfully! Remaining balance: UGX {remaining:,.0f}'
                    
                    from django.urls import reverse
                    # If AJAX, return JSON; otherwise redirect
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': True,
                            'message': message,
                            'payment_id': payment.payment_id,
                            'amount': float(payment.amount),
                            'balance_due': float(payment.invoice.get_balance_due()),
                            'invoice_status': payment.invoice.status,
                            'redirect_url': reverse('billing:invoice_detail', kwargs={'pk': payment.invoice.pk})
                        })
                    else:
                        return redirect('billing:invoice_detail', pk=payment.invoice.pk)
                else:
                    from django.urls import reverse
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': True,
                            'message': f'Payment {payment.payment_id} recorded successfully!',
                            'payment_id': payment.payment_id,
                            'amount': float(payment.amount),
                            'redirect_url': reverse('billing:payment_list')
                        })
                    else:
                        return redirect('billing:payment_list')
        except Exception as e:
            return JsonResponse({
                'success': False,
                'errors': {'__all__': [str(e)]},
                'message': f'Error creating payment: {str(e)}'
            }, status=400)
    else:
        # Return form errors for client-side display
        errors = {}
        for field, error_list in form.errors.items():
            errors[field] = error_list
        
        return JsonResponse({
            'success': False,
            'errors': errors,
            'message': 'Please correct the errors below and try again.'
        }, status=400)

@login_required
def invoices_for_patient_ajax(request):
    """Return a list of open (unpaid) invoices for a given patient (by pk)."""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    patient_id = request.GET.get('patient')
    if not patient_id:
        return JsonResponse({'error': 'Missing patient parameter'}, status=400)
    try:
        patient = Patient.objects.get(pk=patient_id)
    except Patient.DoesNotExist:
        return JsonResponse({'error': 'Patient not found'}, status=404)

    # Consider invoices not fully paid
    invoices_qs = Invoice.objects.filter(patient=patient).exclude(status='paid').order_by('-created_at')

    data = []
    for inv in invoices_qs[:50]:
        try:
            balance = float(inv.get_balance_due()) if hasattr(inv, 'get_balance_due') else float(inv.total_amount)
        except Exception:
            balance = float(inv.total_amount)
        data.append({
            'id': inv.pk,
            'invoice_number': getattr(inv, 'invoice_number', f'INV-{inv.pk}'),
            'status': inv.status,
            'total_amount': float(inv.total_amount),
            'balance_due': balance,
            'created_at': inv.created_at.isoformat() if hasattr(inv, 'created_at') else None,
        })

    return JsonResponse({'success': True, 'invoices': data})


# ==================== EXPORT VIEWS ====================

@login_required
def export_invoices_csv(request):
    """Export filtered invoices as CSV."""
    invoices = Invoice.objects.select_related('patient').all()

    status_filter = request.GET.get('status')
    if status_filter:
        invoices = invoices.filter(status=status_filter)

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        invoices = invoices.filter(issue_date__gte=date_from)
    if date_to:
        invoices = invoices.filter(issue_date__lte=date_to)

    search_query = request.GET.get('search')
    if search_query:
        invoices = invoices.filter(
            Q(invoice_number__icontains=search_query) |
            Q(patient__first_name__icontains=search_query) |
            Q(patient__last_name__icontains=search_query)
        )

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="invoices_export.csv"'

    writer = csv.writer(response)
    writer.writerow(['Invoice #', 'Patient', 'Issue Date', 'Due Date', 'Status', 'Subtotal', 'Tax', 'Discount', 'Total', 'Balance Due', 'Notes'])

    for inv in invoices.order_by('-created_at'):
        try:
            balance = float(inv.get_balance_due()) if hasattr(inv, 'get_balance_due') else float(inv.total_amount)
        except Exception:
            balance = float(inv.total_amount)
        writer.writerow([
            inv.invoice_number,
            inv.patient.get_full_name(),
            inv.issue_date.strftime('%Y-%m-%d') if inv.issue_date else '',
            inv.due_date.strftime('%Y-%m-%d') if inv.due_date else '',
            inv.get_status_display(),
            float(inv.subtotal),
            float(inv.tax_amount),
            float(inv.discount_amount),
            float(inv.total_amount),
            balance,
            inv.notes or '',
        ])

    return response


@login_required
def export_payments_csv(request):
    """Export filtered payments as CSV."""
    payments = Payment.objects.select_related('patient', 'invoice').all()

    status_filter = request.GET.get('status')
    if status_filter:
        payments = payments.filter(status=status_filter)

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        payments = payments.filter(payment_date__date__gte=date_from)
    if date_to:
        payments = payments.filter(payment_date__date__lte=date_to)

    patient_search = request.GET.get('patient_search')
    if patient_search:
        payments = payments.filter(
            Q(patient__first_name__icontains=patient_search) |
            Q(patient__last_name__icontains=patient_search)
        )

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="payments_export.csv"'

    writer = csv.writer(response)
    writer.writerow(['Payment ID', 'Patient', 'Invoice #', 'Date', 'Amount (UGX)', 'Method', 'Status', 'Reference', 'Processed By', 'Notes'])

    for pay in payments.order_by('-payment_date'):
        writer.writerow([
            pay.payment_id,
            pay.patient.get_full_name(),
            pay.invoice.invoice_number if pay.invoice else '',
            pay.payment_date.strftime('%Y-%m-%d') if pay.payment_date else '',
            float(pay.amount),
            pay.get_payment_method_display(),
            pay.get_status_display(),
            pay.reference_number or '',
            pay.processed_by.get_full_name() if pay.processed_by else '',
            pay.notes or '',
        ])

    return response


# ==================== REFUND VIEW ====================

@login_required
def payment_refund(request, pk):
    """Mark a payment as refunded and create an audit log entry."""
    payment = get_object_or_404(Payment, pk=pk)

    if request.method == 'POST':
        reason = request.POST.get('reason', '').strip()
        try:
            with transaction.atomic():
                original_status = payment.status
                payment.status = 'refunded'
                payment.save()

                # Re-open invoice if it was marked paid
                if payment.invoice and payment.invoice.status == 'paid':
                    payment.invoice.status = 'sent'
                    payment.invoice.save()

                BillingAuditLog.objects.create(
                    action='payment_refunded',
                    performed_by=request.user,
                    invoice=payment.invoice,
                    payment=payment,
                    details=f'Payment {payment.payment_id} (UGX {payment.amount:,}) refunded. Reason: {reason or "Not specified"}. Previous status: {original_status}.',
                )

            messages.success(request, f'Payment {payment.payment_id} has been marked as refunded.')
            if payment.invoice:
                return redirect('billing:invoice_detail', pk=payment.invoice.pk)
        except Exception as e:
            messages.error(request, f'Error processing refund: {str(e)}')

    return redirect('billing:payment_list')


# ==================== BILLING REPORTS VIEW ====================

@login_required
def billing_reports(request):
    """Comprehensive billing reports: revenue by service, aging summary, payment methods."""
    from django.db.models import Count, Avg
    from .models import GroupInvoice, GroupPayment
    import json

    today = date.today()
    current_month = today.replace(day=1)

    # Revenue by service (top services this month)
    revenue_by_service = (
        InvoiceLineItem.objects
        .filter(invoice__payments__status='completed', invoice__payments__payment_date__gte=current_month)
        .values('service__name', 'description')
        .annotate(total=Sum('total_amount'), count=Count('id'), paid_total=Sum('invoice__payments__amount'))
        .order_by('-total')[:15]
    )
    # Add balance_total to each service
    for svc in revenue_by_service:
        svc['balance_total'] = (svc['total'] or 0) - (svc['paid_total'] or 0)

    # Monthly totals for last 12 months - include both regular and group payments
    monthly_totals = []
    for i in range(11, -1, -1):
        m_start = (current_month - timedelta(days=32 * i)).replace(day=1)
        m_end = (m_start + timedelta(days=32)).replace(day=1)
        # Regular payments
        regular_rev = Payment.objects.filter(status='completed', payment_date__gte=m_start, payment_date__lt=m_end).aggregate(Sum('amount'))['amount__sum'] or 0
        # Group payments
        group_rev = GroupPayment.objects.filter(status='completed', payment_date__gte=m_start, payment_date__lt=m_end).aggregate(Sum('amount'))['amount__sum'] or 0
        monthly_totals.append({'month': m_start.strftime('%b %Y'), 'revenue': float(regular_rev + group_rev)})

    # Payment method breakdown (all-time) - combine regular and group payments
    method_breakdown = (
        Payment.objects.filter(status='completed')
        .values('payment_method')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )
    method_breakdown_display = []
    for m in method_breakdown:
        method_breakdown_display.append({
            'method': dict(Payment.PAYMENT_METHODS).get(m['payment_method'], m['payment_method']),
            'total': float(m['total']),
            'count': m['count'],
        })
    
    # Add group payment methods
    group_methods = GroupPayment.objects.filter(status='completed').values('payment_method').annotate(total=Sum('amount'), count=Count('id'))
    for gm in group_methods:
        method_name = dict(GroupPayment.PAYMENT_METHODS).get(gm['payment_method'], gm['payment_method'])
        # Check if method already exists
        existing = next((m for m in method_breakdown_display if m['method'] == method_name), None)
        if existing:
            existing['total'] += float(gm['total'])
            existing['count'] += gm['count']
        else:
            method_breakdown_display.append({
                'method': method_name,
                'total': float(gm['total']),
                'count': gm['count'],
            })
    # Re-sort by total
    method_breakdown_display.sort(key=lambda x: x['total'], reverse=True)

    # Aging summary - include both regular and group invoices
    aging_summary = {
        'current': {
            'invoices': Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__gte=today),
            'total': (Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__gte=today).aggregate(Sum('total_amount'))['total_amount__sum'] or 0) +
                     (GroupInvoice.objects.filter(status__in=['sent', 'overdue'], due_date__gte=today).aggregate(Sum('total_amount'))['total_amount__sum'] or 0),
        },
        'overdue_1_30': {
            'invoices': Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today, due_date__gte=today - timedelta(days=30)),
            'total': (Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today, due_date__gte=today - timedelta(days=30)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0) +
                     (GroupInvoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today, due_date__gte=today - timedelta(days=30)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0),
        },
        'overdue_31_60': {
            'invoices': Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=30), due_date__gte=today - timedelta(days=60)),
            'total': (Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=30), due_date__gte=today - timedelta(days=60)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0) +
                     (GroupInvoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=30), due_date__gte=today - timedelta(days=60)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0),
        },
        'overdue_61_90': {
            'invoices': Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=60), due_date__gte=today - timedelta(days=90)),
            'total': (Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=60), due_date__gte=today - timedelta(days=90)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0) +
                     (GroupInvoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=60), due_date__gte=today - timedelta(days=90)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0),
        },
        'overdue_90_plus': {
            'invoices': Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=90)),
            'total': (Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=90)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0) +
                     (GroupInvoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=90)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0),
        },
    }

    # Top patients by outstanding balance
    top_debtors = (
        Invoice.objects.filter(status__in=['sent', 'overdue'])
        .select_related('patient')
        .values('patient__id', 'patient__first_name', 'patient__last_name')
        .annotate(outstanding=Sum('total_amount'), invoice_count=Count('id'))
        .order_by('-outstanding')[:10]
    )

    # Overall stats - include both regular and group invoices/payments
    regular_invoiced = Invoice.objects.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    group_invoiced = GroupInvoice.objects.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    regular_collected = Payment.objects.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
    group_collected = GroupPayment.objects.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
    regular_outstanding = Invoice.objects.filter(status__in=['sent', 'overdue']).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    group_outstanding = GroupInvoice.objects.filter(status__in=['sent', 'overdue']).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    overall_stats = {
        'total_invoiced': regular_invoiced + group_invoiced,
        'total_collected': regular_collected + group_collected,
        'total_outstanding': regular_outstanding + group_outstanding,
        'total_invoices': Invoice.objects.count() + GroupInvoice.objects.count(),
        'paid_invoices': Invoice.objects.filter(status='paid').count() + GroupInvoice.objects.filter(status='paid').count(),
        'overdue_invoices': Invoice.objects.filter(status='overdue').count() + GroupInvoice.objects.filter(status='overdue').count(),
        'regular_invoiced': regular_invoiced,
        'group_invoiced': group_invoiced,
        'regular_collected': regular_collected,
        'group_collected': group_collected,
    }

    monthly_labels = json.dumps([m['month'] for m in monthly_totals])
    monthly_data = json.dumps([m['revenue'] for m in monthly_totals])

    # ---- PER-PATIENT BILLING SUMMARY ----
    patient_summary = (
        Invoice.objects
        .values('patient__id', 'patient__first_name', 'patient__last_name')
        .annotate(
            total_invoiced=Sum('total_amount'),
            invoice_count=Count('id'),
            paid_count=Count('id', filter=Q(status='paid')),
            overdue_count=Count('id', filter=Q(status='overdue')),
        )
        .order_by('-total_invoiced')[:30]
    )
    # Attach total paid per patient
    patient_ids = [p['patient__id'] for p in patient_summary]
    paid_by_patient = {
        entry['patient_id']: entry['total']
        for entry in Payment.objects.filter(status='completed', patient_id__in=patient_ids)
        .values('patient_id').annotate(total=Sum('amount'))
    }
    patient_summary_display = []
    for p in patient_summary:
        total_inv = float(p['total_invoiced'] or 0)
        total_paid = float(paid_by_patient.get(p['patient__id'], 0))
        patient_summary_display.append({
            'id': p['patient__id'],
            'name': f"{p['patient__first_name']} {p['patient__last_name']}",
            'total_invoiced': total_inv,
            'total_paid': total_paid,
            'balance': total_inv - total_paid,
            'invoice_count': p['invoice_count'],
            'paid_count': p['paid_count'],
            'overdue_count': p['overdue_count'],
        })

    # ---- APPOINTMENTS BILLED ----
    appt_billed = (
        InvoiceLineItem.objects
        .filter(appointment__isnull=False)
        .select_related(
            'appointment__patient', 'appointment__service',
            'appointment__provider', 'invoice'
        )
        .order_by('-invoice__created_at')[:50]
    )
    # Appointments revenue by service
    appt_by_service = (
        InvoiceLineItem.objects
        .filter(appointment__isnull=False)
        .values('appointment__service__name')
        .annotate(total=Sum('total_amount'), count=Count('id'))
        .order_by('-total')[:10]
    )
    appt_total = InvoiceLineItem.objects.filter(appointment__isnull=False).aggregate(Sum('total_amount'))['total_amount__sum'] or 0

    # ---- LAB TESTS BILLED ----
    lab_billed = (
        InvoiceLineItem.objects
        .filter(lab_test_request__isnull=False)
        .select_related(
            'lab_test_request__patient', 'lab_test_request__test',
            'invoice'
        )
        .prefetch_related('invoice__payments')
        .order_by('-invoice__created_at')[:50]
    )
    # Add paid_amount and balance_amount to each lab billed item
    for item in lab_billed:
        paid = float(sum(p.amount for p in item.invoice.payments.filter(status='completed')))
        item.paid_amount = paid
        item.balance_amount = float(item.total_amount) - paid
    # Lab tests revenue by test name
    lab_by_test = (
        InvoiceLineItem.objects
        .filter(lab_test_request__isnull=False)
        .values('lab_test_request__test__name')
        .annotate(total=Sum('total_amount'), count=Count('id'), paid_total=Sum('invoice__payments__amount'))
        .order_by('-total')[:10]
    )
    # Add balance_total to each lab test
    for t in lab_by_test:
        t['balance_total'] = (t['total'] or 0) - (t['paid_total'] or 0)
    lab_total = InvoiceLineItem.objects.filter(lab_test_request__isnull=False).aggregate(Sum('total_amount'))['total_amount__sum'] or 0

    # Revenue by service (all-time for reports)
    revenue_by_service_alltime = (
        InvoiceLineItem.objects
        .values('service__name', 'description')
        .annotate(total=Sum('total_amount'), count=Count('id'), avg=Avg('unit_price'), paid_total=Sum('invoice__payments__amount'))
        .order_by('-total')[:20]
    )
    # Add balance_total to each service
    for svc in revenue_by_service_alltime:
        svc['balance_total'] = (svc['total'] or 0) - (svc['paid_total'] or 0)

    # Chart: service revenue bar (top 10 all-time)
    svc_report_labels = json.dumps([
        (s['service__name'] or s['description'] or '–')[:25]
        for s in revenue_by_service_alltime[:10]
    ])
    svc_report_data = json.dumps([float(s['total'] or 0) for s in revenue_by_service_alltime[:10]])

    # Aging chart data (combined regular + group)
    aging_chart_labels = json.dumps(['Current', '1–30 Days', '31–60 Days', '61–90 Days', '90+ Days'])
    aging_chart_data = json.dumps([
        float(aging_summary['current']['total']),
        float(aging_summary['overdue_1_30']['total']),
        float(aging_summary['overdue_31_60']['total']),
        float(aging_summary['overdue_61_90']['total']),
        float(aging_summary['overdue_90_plus']['total']),
    ])

    # Method chart for reports
    method_chart_labels = json.dumps([m['method'] for m in method_breakdown_display])
    method_chart_data = json.dumps([m['total'] for m in method_breakdown_display])

    context = {
        'revenue_by_service': revenue_by_service,
        'revenue_by_service_alltime': revenue_by_service_alltime,
        'monthly_totals': monthly_totals,
        'monthly_labels': monthly_labels,
        'monthly_data': monthly_data,
        'method_breakdown': method_breakdown_display,
        'method_chart_labels': method_chart_labels,
        'method_chart_data': method_chart_data,
        'aging_summary': aging_summary,
        'aging_chart_labels': aging_chart_labels,
        'aging_chart_data': aging_chart_data,
        'top_debtors': top_debtors,
        'overall_stats': overall_stats,
        'today': today,
        'patient_summary': patient_summary_display,
        'appt_billed': appt_billed,
        'appt_by_service': appt_by_service,
        'appt_total': float(appt_total),
        'lab_billed': lab_billed,
        'lab_by_test': lab_by_test,
        'lab_total': float(lab_total),
        'svc_report_labels': svc_report_labels,
        'svc_report_data': svc_report_data,
    }
    return render(request, 'billing/billing_reports.html', context)


# ==================== EXCEL EXPORT ====================

@login_required
def export_billing_excel(request):
    """
    Export billing data as a multi-sheet Excel workbook.
    Sheet 1 (Overview) acts as a control/summary sheet linking to all others.
    Accepts filters: date_from, date_to, inv_status, pay_method, patient_q, section
    If section param is given, only that sheet + Overview are exported.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
    from django.db.models import Count, Avg, Max, Min

    # ── Filters ──────────────────────────────────────────────────────────────
    date_from    = request.GET.get('date_from')
    date_to      = request.GET.get('date_to')
    inv_status   = request.GET.get('inv_status')
    pay_method   = request.GET.get('pay_method')
    patient_q    = request.GET.get('patient_q', '').strip()
    section      = request.GET.get('section', 'all')  # all | invoices | payments | patients | lab | appt | services | aging

    today = date.today()
    current_month = today.replace(day=1)

    # ── Base querysets with filters ───────────────────────────────────────────
    inv_qs = Invoice.objects.select_related('patient', 'created_by').all()
    pay_qs = Payment.objects.select_related('patient', 'invoice', 'processed_by').all()

    if date_from:
        inv_qs = inv_qs.filter(issue_date__gte=date_from)
        pay_qs = pay_qs.filter(payment_date__date__gte=date_from)
    if date_to:
        inv_qs = inv_qs.filter(issue_date__lte=date_to)
        pay_qs = pay_qs.filter(payment_date__date__lte=date_to)
    if inv_status:
        inv_qs = inv_qs.filter(status=inv_status)
    if pay_method:
        pay_qs = pay_qs.filter(payment_method=pay_method)
    if patient_q:
        inv_qs = inv_qs.filter(Q(patient__first_name__icontains=patient_q) | Q(patient__last_name__icontains=patient_q))
        pay_qs = pay_qs.filter(Q(patient__first_name__icontains=patient_q) | Q(patient__last_name__icontains=patient_q))

    # ── Workbook helpers ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    HDR_FILL   = PatternFill('solid', fgColor='2C3E50')
    HDR_FONT   = Font(color='FFFFFF', bold=True, size=11)
    ALT_FILL   = PatternFill('solid', fgColor='F2F6FA')
    TITLE_FONT = Font(bold=True, size=13)
    LINK_FONT  = Font(color='1155CC', underline='single', bold=True)
    thin       = Side(style='thin', color='CCCCCC')
    BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)
    CURRENCY   = '#,##0'
    CENTER     = Alignment(horizontal='center', vertical='center')

    def make_header(ws, columns, row=1):
        for col_idx, title in enumerate(columns, 1):
            cell = ws.cell(row=row, column=col_idx, value=title)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER
            cell.border = BORDER
        ws.freeze_panes = ws.cell(row=row + 1, column=1)

    def auto_width(ws, min_w=10, max_w=50):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or '')))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 3, max_w))

    def style_row(ws, row_idx, n_cols, alternate=False):
        fill = ALT_FILL if alternate else None
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = BORDER
            if fill:
                cell.fill = fill

    def currency_cols(ws, col_indices, start_row, end_row):
        for c in col_indices:
            for r in range(start_row, end_row + 1):
                ws.cell(row=r, column=c).number_format = CURRENCY

    # ── Filter description (used in Overview header) ──────────────────────
    filter_desc = []
    if date_from:  filter_desc.append(f'From {date_from}')
    if date_to:    filter_desc.append(f'To {date_to}')
    if patient_q:  filter_desc.append(f'Patient: {patient_q}')
    if inv_status: filter_desc.append(f'Status: {inv_status}')

    # ── SHEET 1: Overview — Full Management Dashboard ──────────────────────
    ws_ov = wb.create_sheet('Overview')
    ws_ov.sheet_view.showGridLines = False
    ws_ov.sheet_properties.tabColor = '1A252F'

    # ── Extended style palette for Overview ──────────────────────────────
    NAVY_F   = PatternFill('solid', fgColor='1A252F')
    DARK_F   = PatternFill('solid', fgColor='2C3E50')
    BLUE_F   = PatternFill('solid', fgColor='2980B9')
    GREEN_F  = PatternFill('solid', fgColor='1E8449')
    RED_F    = PatternFill('solid', fgColor='C0392B')
    GOLD_F   = PatternFill('solid', fgColor='B7950B')
    TEAL_F   = PatternFill('solid', fgColor='148F77')
    PURP_F   = PatternFill('solid', fgColor='6C3483')
    ORNG_F   = PatternFill('solid', fgColor='BA4A00')
    SLATE_F  = PatternFill('solid', fgColor='1B4F72')
    LBL_F    = PatternFill('solid', fgColor='D6EAF8')
    ALT_F    = PatternFill('solid', fgColor='EBF5FB')
    YEL_F    = PatternFill('solid', fgColor='FDEBD0')
    INPUT_F  = PatternFill('solid', fgColor='FFEAA7')
    WHT_F    = PatternFill('solid', fgColor='FFFFFF')
    LGRY_F   = PatternFill('solid', fgColor='F2F3F4')
    DGRY_F   = PatternFill('solid', fgColor='CCD1D1')

    WF   = Font(color='FFFFFF', bold=True, size=11)
    WF_L = Font(color='FFFFFF', bold=True, size=14)
    WF_S = Font(color='FFFFFF', size=9, italic=True)
    DF   = Font(color='2C3E50', bold=True)
    DF_S = Font(color='2C3E50', size=9)
    LF   = Font(color='1155CC', underline='single', bold=True)
    GF   = Font(color='7F8C8D', italic=True, size=9)

    OV_BORDER = Border(
        left=Side(style='thin', color='AEB6BF'),
        right=Side(style='thin', color='AEB6BF'),
        top=Side(style='thin', color='AEB6BF'),
        bottom=Side(style='thin', color='AEB6BF'),
    )
    THICK_BORDER = Border(
        left=Side(style='medium', color='2C3E50'),
        right=Side(style='medium', color='2C3E50'),
        top=Side(style='medium', color='2C3E50'),
        bottom=Side(style='medium', color='2C3E50'),
    )
    C = Alignment(horizontal='center', vertical='center', wrap_text=True)
    L = Alignment(horizontal='left',   vertical='center', wrap_text=False)
    R = Alignment(horizontal='right',  vertical='center')
    CURR = '#,##0'
    PCT  = '0.0"%"'

    # ── Column widths ──────────────────────────────────────────────────────
    col_w = {'A': 2, 'B': 32, 'C': 22, 'D': 22, 'E': 22, 'F': 22, 'G': 22, 'H': 22, 'I': 2}
    for col_l, w in col_w.items():
        ws_ov.column_dimensions[col_l].width = w

    def ov_hdr(row, label, fill=DARK_F, fnt=None, ht=24):
        ws_ov.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        c = ws_ov.cell(row=row, column=2, value=label)
        c.fill = fill; c.font = fnt or Font(color='FFFFFF', bold=True, size=12)
        c.alignment = C
        for col in range(2, 9):
            ws_ov.cell(row=row, column=col).border = OV_BORDER
        ws_ov.row_dimensions[row].height = ht

    def ov_set(row, col, val, fill=None, fnt=None, align=None, fmt=None, merge_end=None, border=True):
        if merge_end:
            ws_ov.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end)
            for c2 in range(col, merge_end + 1):
                if border: ws_ov.cell(row=row, column=c2).border = OV_BORDER
        cell = ws_ov.cell(row=row, column=col, value=val)
        if fill:  cell.fill = fill
        if fnt:   cell.font = fnt
        if align: cell.alignment = align
        if fmt:   cell.number_format = fmt
        if border and not merge_end: cell.border = OV_BORDER
        return cell

    # ═══════════════════════════════════════════════════════════════════════
    # BANNER (rows 1-3)
    # ═══════════════════════════════════════════════════════════════════════
    ws_ov.row_dimensions[1].height = 42
    ws_ov.merge_cells('B1:H1')
    c = ws_ov.cell(row=1, column=2, value='PhysioNutrition Clinic  ·  Billing Management Dashboard')
    c.fill = NAVY_F; c.font = Font(color='FFFFFF', bold=True, size=20); c.alignment = C
    for col in range(2, 9): ws_ov.cell(row=1, column=col).border = OV_BORDER

    ws_ov.row_dimensions[2].height = 20
    ws_ov.merge_cells('B2:H2')
    c = ws_ov.cell(row=2, column=2,
        value='Comprehensive Billing Control Centre  |  All data sheets have AutoFilter enabled — use the ▾ dropdowns to filter/sort')
    c.fill = DARK_F; c.font = Font(color='AEB6BF', italic=True, size=9); c.alignment = C
    for col in range(2, 9): ws_ov.cell(row=2, column=col).border = OV_BORDER

    ws_ov.row_dimensions[3].height = 18
    ov_set(3, 2, 'Report Date:', LBL_F, DF, L)
    c_date = ws_ov.cell(row=3, column=3, value='=TODAY()')
    c_date.number_format = 'DD-MMM-YYYY'; c_date.font = Font(bold=True, color='2980B9'); c_date.alignment = L
    c_date.border = OV_BORDER
    filter_text = 'Filters: ' + (', '.join(filter_desc) if filter_desc else 'None — showing all data')
    ov_set(3, 4, filter_text, fill=LBL_F, fnt=Font(italic=True, color='5D6D7E', size=9), align=L, merge_end=8)

    ws_ov.row_dimensions[4].height = 6  # spacer

    # ═══════════════════════════════════════════════════════════════════════
    # SECTION 1 — KPI DASHBOARD  (rows 5-21)
    # Each KPI: 3 rows tall (label | big value | sub-note)
    # Two KPIs per row (columns B-D left, E-H right)
    # ═══════════════════════════════════════════════════════════════════════
    ov_hdr(5, '📊   KEY PERFORMANCE INDICATORS  —  Values are live Excel formulas; they recalculate every time you open this file', NAVY_F,
           Font(color='FFFFFF', bold=True, size=11), ht=26)

    kpi_items = [
        # (label, value_formula, sub_note, number_format, fill)
        ('Total Invoices',
         '=IFERROR(COUNTA(Invoices!A:A)-1,0)',
         'Draft + Sent + Paid + Overdue + Cancelled',
         '0', BLUE_F),
        ('Total Invoiced (UGX)',
         '=IFERROR(SUM(Invoices!I:I),0)',
         'Sum of all invoice totals (col I in Invoices sheet)',
         CURR, TEAL_F),
        ('Paid Invoices',
         '=IFERROR(COUNTIF(Invoices!E:E,"Paid"),0)',
         '=IFERROR(TEXT(COUNTIF(Invoices!E:E,"Paid")/(COUNTA(Invoices!A:A)-1)*100,"0.0")&"% of all invoices","")',
         '0', GREEN_F),
        ('Overdue Invoices',
         '=IFERROR(COUNTIF(Invoices!E:E,"Overdue"),0)',
         '=IFERROR(TEXT(COUNTIF(Invoices!E:E,"Overdue")/(COUNTA(Invoices!A:A)-1)*100,"0.0")&"% of all invoices","")',
         '0', RED_F),
        ('Total Payments',
         '=IFERROR(COUNTA(Payments!A:A)-1,0)',
         'All payment transactions recorded',
         '0', PURP_F),
        ('Total Collected (UGX)',
         '=IFERROR(SUMIF(Payments!G:G,"Completed",Payments!E:E),0)',
         'Completed payments only (col G = Status, col E = Amount)',
         CURR, GOLD_F),
        ('Outstanding Balance (UGX)',
         '=IFERROR(SUM(Invoices!J:J),0)',
         'Sum of Balance Due column (col J) in Invoices sheet',
         CURR, ORNG_F),
        ('Collection Rate',
         '=IFERROR(SUMIF(Payments!G:G,"Completed",Payments!E:E)/SUM(Invoices!I:I)*100,0)',
         'Total Collected ÷ Total Invoiced × 100',
         '0.0"%"', SLATE_F),
    ]

    kpi_row = 6
    for i, (label, val_f, sub_f, fmt, fill) in enumerate(kpi_items):
        if i % 2 == 0 and i > 0:
            kpi_row += 3
        cs = 2 if i % 2 == 0 else 5
        ce = 4 if i % 2 == 0 else 8
        # Label row
        ws_ov.merge_cells(start_row=kpi_row,   start_column=cs, end_row=kpi_row,   end_column=ce)
        lc = ws_ov.cell(row=kpi_row, column=cs, value=label)
        lc.fill = fill; lc.font = WF; lc.alignment = C
        for cc in range(cs, ce + 1): ws_ov.cell(row=kpi_row, column=cc).border = OV_BORDER
        ws_ov.row_dimensions[kpi_row].height = 18
        # Value row
        ws_ov.merge_cells(start_row=kpi_row+1, start_column=cs, end_row=kpi_row+1, end_column=ce)
        vc = ws_ov.cell(row=kpi_row+1, column=cs, value=val_f)
        vc.fill = fill; vc.font = Font(color='FFFFFF', bold=True, size=18); vc.alignment = C
        vc.number_format = fmt
        for cc in range(cs, ce + 1): ws_ov.cell(row=kpi_row+1, column=cc).border = OV_BORDER
        ws_ov.row_dimensions[kpi_row+1].height = 32
        # Sub-note row (formula or static text)
        ws_ov.merge_cells(start_row=kpi_row+2, start_column=cs, end_row=kpi_row+2, end_column=ce)
        sc = ws_ov.cell(row=kpi_row+2, column=cs, value=sub_f)
        sc.fill = fill; sc.font = Font(color='D6EAF8', italic=True, size=8); sc.alignment = C
        for cc in range(cs, ce + 1): ws_ov.cell(row=kpi_row+2, column=cc).border = OV_BORDER
        ws_ov.row_dimensions[kpi_row+2].height = 14
    kpi_row += 3

    r = kpi_row + 1  # spacer row
    ws_ov.row_dimensions[r].height = 8
    r += 1

    # ═══════════════════════════════════════════════════════════════════════
    # SECTION 2 — INVOICE STATUS BREAKDOWN
    # ═══════════════════════════════════════════════════════════════════════
    ov_hdr(r, '📋   INVOICE STATUS BREAKDOWN  —  COUNTIF & SUMIF formulas referencing the Invoices sheet', DARK_F, ht=22)
    r += 1
    for ci, h in enumerate(['Status', 'Count', 'Total Amount (UGX)', '% of Invoices', '% of Revenue'], 2):
        c = ws_ov.cell(row=r, column=ci, value=h)
        c.fill = HDR_FILL; c.font = WF; c.alignment = C; c.border = OV_BORDER
    ws_ov.row_dimensions[r].height = 20
    r += 1

    status_rows = [
        ('Draft',     'Draft',     PatternFill('solid', fgColor='D6EAF8')),
        ('Sent',      'Sent',      PatternFill('solid', fgColor='D1F2EB')),
        ('Paid',      'Paid',      PatternFill('solid', fgColor='EAFAF1')),
        ('Overdue',   'Overdue',   PatternFill('solid', fgColor='FDEDEC')),
        ('Cancelled', 'Cancelled', LGRY_F),
    ]
    total_inv_f = '=IFERROR(COUNTA(Invoices!A:A)-1,1)'
    total_rev_f = '=IFERROR(SUM(Invoices!I:I),1)'
    for label, display, sf in status_rows:
        count_f = f'=IFERROR(COUNTIF(Invoices!E:E,"{display}"),0)'
        total_f = f'=IFERROR(SUMIF(Invoices!E:E,"{display}",Invoices!I:I),0)'
        pct_c   = f'=IFERROR(C{r}/({total_inv_f})*100,0)'
        pct_r   = f'=IFERROR(D{r}/({total_rev_f})*100,0)'
        for ci, (val, fmt, aln) in enumerate([
            (label,   '@',  L), (count_f, '0',   R),
            (total_f, CURR, R), (pct_c,   PCT,   R), (pct_r, PCT, R),
        ], 2):
            c = ws_ov.cell(row=r, column=ci, value=val)
            c.fill = sf; c.font = DF if ci == 2 else DF_S
            c.alignment = aln; c.border = OV_BORDER; c.number_format = fmt
        ws_ov.row_dimensions[r].height = 17
        r += 1
    # Totals row
    for ci, (val, fmt) in enumerate([
        ('TOTAL', '@'), (f'=IFERROR(COUNTA(Invoices!A:A)-1,0)', '0'),
        (f'=IFERROR(SUM(Invoices!I:I),0)', CURR), ('100%', '@'), ('100%', '@'),
    ], 2):
        c = ws_ov.cell(row=r, column=ci, value=val)
        c.fill = DGRY_F; c.font = DF; c.alignment = R if ci > 2 else L
        c.border = OV_BORDER; c.number_format = fmt
    ws_ov.row_dimensions[r].height = 18
    r += 2  # spacer

    # ═══════════════════════════════════════════════════════════════════════
    # SECTION 3 — PAYMENT METHOD SUMMARY
    # ═══════════════════════════════════════════════════════════════════════
    ov_hdr(r, '💳   PAYMENT METHOD SUMMARY  —  SUMIF & COUNTIF formulas referencing the Payments sheet', DARK_F, ht=22)
    r += 1
    for ci, h in enumerate(['Method', 'Transactions', 'Collected (UGX)', '% of Collections', 'Avg Transaction (UGX)'], 2):
        c = ws_ov.cell(row=r, column=ci, value=h)
        c.fill = HDR_FILL; c.font = WF; c.alignment = C; c.border = OV_BORDER
    ws_ov.row_dimensions[r].height = 20
    r += 1

    method_rows = [
        ('Cash',          PatternFill('solid', fgColor='EAFAF1')),
        ('Mobile Money',  PatternFill('solid', fgColor='EBF5FB')),
        ('Bank Transfer', PatternFill('solid', fgColor='F4ECF7')),
        ('Card',          PatternFill('solid', fgColor='FEF9E7')),
        ('Insurance',     PatternFill('solid', fgColor='FDEDEC')),
        ('Other',         LGRY_F),
    ]
    total_pay_f = '=IFERROR(SUM(Payments!E:E),1)'
    for mname, mfill in method_rows:
        cnt_f = f'=IFERROR(COUNTIF(Payments!F:F,"{mname}"),0)'
        tot_f = f'=IFERROR(SUMIF(Payments!F:F,"{mname}",Payments!E:E),0)'
        pct_f = f'=IFERROR(D{r}/({total_pay_f})*100,0)'
        avg_f = f'=IFERROR(D{r}/C{r},0)'
        for ci, (val, fmt, aln) in enumerate([
            (mname, '@', L), (cnt_f, '0', R), (tot_f, CURR, R), (pct_f, PCT, R), (avg_f, CURR, R),
        ], 2):
            c = ws_ov.cell(row=r, column=ci, value=val)
            c.fill = mfill; c.font = DF if ci == 2 else DF_S
            c.alignment = aln; c.border = OV_BORDER; c.number_format = fmt
        ws_ov.row_dimensions[r].height = 17
        r += 1
    for ci, (val, fmt) in enumerate([
        ('TOTAL', '@'), (f'=IFERROR(COUNTA(Payments!A:A)-1,0)', '0'),
        (f'=IFERROR(SUM(Payments!E:E),0)', CURR), ('100%', '@'),
        (f'=IFERROR(SUM(Payments!E:E)/(COUNTA(Payments!A:A)-1),0)', CURR),
    ], 2):
        c = ws_ov.cell(row=r, column=ci, value=val)
        c.fill = DGRY_F; c.font = DF; c.alignment = R if ci > 2 else L
        c.border = OV_BORDER; c.number_format = fmt
    ws_ov.row_dimensions[r].height = 18
    r += 2

    # ═══════════════════════════════════════════════════════════════════════
    # SECTION 4 — PATIENT SEARCH TOOL
    # ═══════════════════════════════════════════════════════════════════════
    ov_hdr(r,
        '🔍   PATIENT SEARCH TOOL  —  Type a name in the yellow cell; all rows below update instantly via SUMIF/COUNTIF',
        NAVY_F, Font(color='FFFFFF', bold=True, size=11), ht=26)
    r += 1

    # Instructions strip
    ws_ov.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ic = ws_ov.cell(row=r, column=2,
        value='ℹ  Enter a full or partial patient name below. Partial matches work (e.g. "john" will match "Johnson Mary"). '
              'Results pull from the Patients sheet using SUMIF with wildcard (*name*).')
    ic.fill = ALT_F; ic.font = Font(italic=True, color='5D6D7E', size=9); ic.alignment = L
    for col in range(2, 9): ws_ov.cell(row=r, column=col).border = OV_BORDER
    ws_ov.row_dimensions[r].height = 16
    r += 1

    # Search input row
    ov_set(r, 2, '🔎  Enter Patient Name:', LBL_F, DF, L)
    ws_ov.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    si = ws_ov.cell(row=r, column=3, value='')
    si.fill = INPUT_F
    si.font = Font(color='2C3E50', bold=True, size=13)
    si.alignment = L
    si.border = Border(
        left=Side(style='medium', color='F39C12'), right=Side(style='medium', color='F39C12'),
        top=Side(style='medium', color='F39C12'),  bottom=Side(style='medium', color='F39C12'),
    )
    for col in range(3, 6): ws_ov.cell(row=r, column=col).border = OV_BORDER
    ov_set(r, 6, '← Type here. Partial names OK (e.g. "john", "mary")', LGRY_F,
           Font(italic=True, color='95A5A6', size=9), L, merge_end=8)
    ws_ov.row_dimensions[r].height = 26
    search_ref = f'C{r}'
    r += 1

    # Search results table header
    for ci, h in enumerate(['Billing Metric', 'Value (auto-calculated)', 'How it works (formula source)'], 2):
        c = ws_ov.cell(row=r, column=ci, value=h)
        c.fill = HDR_FILL; c.font = WF; c.alignment = C; c.border = OV_BORDER
    ws_ov.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
    for col in range(4, 9): ws_ov.cell(row=r, column=col).border = OV_BORDER
    ws_ov.row_dimensions[r].height = 20
    r += 1

    search_results = [
        ('Total Invoiced (UGX)',
         f'=IFERROR(SUMIF(Patients!A:A,"*"&{search_ref}&"*",Patients!E:E),0)',
         'SUMIF(Patients!A, "*"&name&"*", Patients!E)  — col E = Total Invoiced',
         CURR, PatternFill('solid', fgColor='EBF5FB')),
        ('Total Paid (UGX)',
         f'=IFERROR(SUMIF(Patients!A:A,"*"&{search_ref}&"*",Patients!F:F),0)',
         'SUMIF(Patients!A, "*"&name&"*", Patients!F)  — col F = Total Paid',
         CURR, PatternFill('solid', fgColor='EAFAF1')),
        ('Outstanding Balance (UGX)',
         f'=IFERROR(SUMIF(Patients!A:A,"*"&{search_ref}&"*",Patients!G:G),0)',
         'SUMIF(Patients!A, "*"&name&"*", Patients!G)  — col G = Balance',
         CURR, PatternFill('solid', fgColor='FDEDEC')),
        ('Collection Rate (%)',
         f'=IFERROR(SUMIF(Patients!A:A,"*"&{search_ref}&"*",Patients!F:F)'
         f'/SUMIF(Patients!A:A,"*"&{search_ref}&"*",Patients!E:E)*100,0)',
         'Paid ÷ Invoiced × 100  (cols F and E from Patients sheet)',
         PCT, PatternFill('solid', fgColor='FEF9E7')),
        ('Total Invoice Count',
         f'=IFERROR(SUMIF(Patients!A:A,"*"&{search_ref}&"*",Patients!B:B),0)',
         'SUMIF(Patients!A, "*"&name&"*", Patients!B)  — col B = Invoice Count',
         '0', PatternFill('solid', fgColor='EBF5FB')),
        ('Paid Invoice Count',
         f'=IFERROR(SUMIF(Patients!A:A,"*"&{search_ref}&"*",Patients!C:C),0)',
         'SUMIF(Patients!A, "*"&name&"*", Patients!C)  — col C = Paid Count',
         '0', PatternFill('solid', fgColor='EAFAF1')),
        ('Overdue Invoice Count',
         f'=IFERROR(SUMIF(Patients!A:A,"*"&{search_ref}&"*",Patients!D:D),0)',
         'SUMIF(Patients!A, "*"&name&"*", Patients!D)  — col D = Overdue Count',
         '0', PatternFill('solid', fgColor='FDEDEC')),
        ('Patients Matched',
         f'=IFERROR(COUNTIF(Patients!A:A,"*"&{search_ref}&"*"),0)',
         'COUNTIF(Patients!A, "*"&name&"*")  — number of rows matching the search',
         '0', PatternFill('solid', fgColor='F4ECF7')),
    ]
    for metric, val_f, detail, fmt, sf in search_results:
        ov_set(r, 2, metric,  sf,   DF,  L)
        ov_set(r, 3, val_f,   WHT_F, Font(bold=True, color='1A5276', size=12), R, fmt)
        ov_set(r, 4, detail,  LGRY_F, GF, L, merge_end=8)
        ws_ov.row_dimensions[r].height = 18
        r += 1
    r += 1

    # ═══════════════════════════════════════════════════════════════════════
    # SECTION 5 — SHEET NAVIGATOR
    # ═══════════════════════════════════════════════════════════════════════
    ov_hdr(r, '🗂️   SHEET NAVIGATOR  —  Click any sheet name or the "→ Go" button to jump directly to that data sheet', DARK_F, ht=22)
    r += 1
    for ci, h in enumerate(['Sheet', 'Contents', 'Quick Stats (live formula)', '', 'Jump'], 2):
        c = ws_ov.cell(row=r, column=ci, value=h)
        c.fill = HDR_FILL; c.font = WF; c.alignment = C; c.border = OV_BORDER
    ws_ov.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
    for col in range(4, 8): ws_ov.cell(row=r, column=col).border = OV_BORDER
    ws_ov.row_dimensions[r].height = 20
    r += 1

    nav_items = [
        ('Invoices',     'Invoices',     'All invoice records — status, amounts, dates, notes',
         '=IFERROR(COUNTA(Invoices!A:A)-1,0)&" invoice rows  |  Total: UGX "&TEXT(SUM(Invoices!I:I),"#,##0")'),
        ('Group_Invoices', 'Group Invoices', 'Group invoice records — consolidated billing for patient groups',
         '=IFERROR(COUNTA(Group_Invoices!A:A)-1,0)&" group invoice rows  |  Total: UGX "&TEXT(SUM(Group_Invoices!H:H),"#,##0")'),
        ('Payments',     'Payments',     'All payment transactions — method, status, reference',
         '=IFERROR(COUNTA(Payments!A:A)-1,0)&" payment rows  |  Collected: UGX "&TEXT(SUM(Payments!E:E),"#,##0")'),
        ('Patients',     'Patients',     'Per-patient billing summary (top 200 by invoiced amount)',
         '=IFERROR(COUNTA(Patients!A:A)-1,0)&" patient rows"'),
        ('Lab_Tests',    'Lab Tests',    'Lab test invoice line items with test name, date, status',
         '=IFERROR(COUNTA(Lab_Tests!A:A)-1,0)&" lab records  |  Revenue: UGX "&TEXT(SUM(Lab_Tests!F:F),"#,##0")'),
        ('Appointments', 'Appointments', 'Appointment invoice line items with service and provider',
         '=IFERROR(COUNTA(Appointments!A:A)-1,0)&" appt records  |  Revenue: UGX "&TEXT(SUM(Appointments!F:F),"#,##0")'),
        ('Services',     'Services',     'All-time revenue per service with avg / max / min pricing',
         '=IFERROR(COUNTA(Services!A:A)-1,0)&" service types  |  Total: UGX "&TEXT(SUM(Services!C:C),"#,##0")'),
        ('Aging',        'Aging',        'Accounts receivable aging buckets (combined regular + group)',
         '"5 aging buckets  |  Outstanding: UGX "&TEXT(SUM(Aging!C:C),"#,##0")'),
    ]
    nav_fill_alt = [ALT_F if i % 2 == 0 else WHT_F for i in range(len(nav_items))]
    for idx, (name, label, desc, stat_f) in enumerate(nav_items):
        nf = nav_fill_alt[idx]
        # Sheet name cell (hyperlink)
        nc = ws_ov.cell(row=r, column=2, value=label)
        nc.hyperlink = f'#{name}!A1'; nc.font = LF; nc.fill = nf; nc.border = OV_BORDER; nc.alignment = L
        # Description
        ov_set(r, 3, desc, nf, DF_S, L)
        # Stats formula
        ws_ov.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        sc = ws_ov.cell(row=r, column=4, value=stat_f)
        sc.fill = nf; sc.font = Font(bold=True, color='1A5276', size=9); sc.alignment = L
        for col in range(4, 8): ws_ov.cell(row=r, column=col).border = OV_BORDER
        # Jump button
        jc = ws_ov.cell(row=r, column=8, value='→ Go')
        jc.hyperlink = f'#{name}!A1'
        jc.fill = BLUE_F; jc.font = Font(color='FFFFFF', bold=True)
        jc.border = OV_BORDER; jc.alignment = C
        ws_ov.row_dimensions[r].height = 20
        r += 1
    r += 1

    # ═══════════════════════════════════════════════════════════════════════
    # SECTION 6 — TIPS & INSTRUCTIONS
    # ═══════════════════════════════════════════════════════════════════════
    ov_hdr(r, '💡   TIPS FOR USING THIS WORKBOOK', DARK_F, ht=20)
    r += 1
    tips = [
        '📌  Each data sheet (Invoices, Payments, etc.) has AutoFilter on the header row — click the ▾ arrows to filter/sort any column.',
        '🔍  Use the Patient Search Tool above to look up any patient by full or partial name. All 8 metrics update as you type.',
        '📊  All KPI values on this sheet are live Excel formulas — they recalculate every time the file is opened.',
        '💾  To export a filtered subset: apply column filters on any sheet, select visible rows (Ctrl+Shift+End), copy to a new sheet.',
        '📅  The "Report Date" cell uses =TODAY() so it always reflects the actual date the file was opened.',
        '➕  To add new records: append rows below the last data row in any sheet. Full-column ranges ensure they are included in formulas.',
        '⚠   Do NOT rename or delete any sheet — the Overview formulas reference sheet names directly (e.g. Invoices!E:E).',
        '🎨  Colour coding:  Blue = informational  |  Green = paid/collected  |  Red = overdue/outstanding  |  Gold = revenue totals.',
    ]
    for i, tip in enumerate(tips):
        ws_ov.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        tc = ws_ov.cell(row=r, column=2, value=tip)
        tc.fill = ALT_F if i % 2 == 0 else WHT_F
        tc.font = Font(color='2C3E50', size=9); tc.alignment = L
        for col in range(2, 9): ws_ov.cell(row=r, column=col).border = OV_BORDER
        ws_ov.row_dimensions[r].height = 16
        r += 1

    # ── SHEET 2: Invoices ──────────────────────────────────────────────────
    ws_inv = wb.create_sheet('Invoices')
    inv_cols = ['Invoice #', 'Patient', 'Issue Date', 'Due Date', 'Status',
                'Subtotal (UGX)', 'Tax (UGX)', 'Discount (UGX)', 'Total (UGX)', 'Balance Due (UGX)', 'Notes']
    make_header(ws_inv, inv_cols)
    for r, inv in enumerate(inv_qs.order_by('-issue_date'), 2):
        try: balance = float(inv.get_balance_due())
        except Exception: balance = float(inv.total_amount)
        row_data = [
            inv.invoice_number,
            inv.patient.get_full_name(),
            inv.issue_date,
            inv.due_date,
            inv.get_status_display(),
            float(inv.subtotal),
            float(inv.tax_amount),
            float(inv.discount_amount),
            float(inv.total_amount),
            balance,
            inv.notes or '',
        ]
        for c, val in enumerate(row_data, 1):
            ws_inv.cell(row=r, column=c, value=val).border = BORDER
        currency_cols(ws_inv, [6, 7, 8, 9, 10], r, r)
        style_row(ws_inv, r, len(inv_cols), alternate=(r % 2 == 0))
    auto_width(ws_inv)
    ws_inv.auto_filter.ref = f'A1:{get_column_letter(len(inv_cols))}1'

    # ── SHEET 3: Payments ──────────────────────────────────────────────────
    ws_pay = wb.create_sheet('Payments')
    pay_cols = ['Payment ID', 'Patient', 'Invoice #', 'Date', 'Amount (UGX)',
                'Method', 'Status', 'Reference', 'Processed By', 'Notes']
    make_header(ws_pay, pay_cols)
    for r, pay in enumerate(pay_qs.order_by('-payment_date'), 2):
        row_data = [
            pay.payment_id,
            pay.patient.get_full_name(),
            pay.invoice.invoice_number if pay.invoice else '',
            pay.payment_date.date() if pay.payment_date else '',
            float(pay.amount),
            pay.get_payment_method_display(),
            pay.get_status_display(),
            pay.reference_number or '',
            pay.processed_by.get_full_name() if pay.processed_by else '',
            pay.notes or '',
        ]
        for c, val in enumerate(row_data, 1):
            ws_pay.cell(row=r, column=c, value=val).border = BORDER
        currency_cols(ws_pay, [5], r, r)
        style_row(ws_pay, r, len(pay_cols), alternate=(r % 2 == 0))
    auto_width(ws_pay)
    ws_pay.auto_filter.ref = f'A1:{get_column_letter(len(pay_cols))}1'

    # ── SHEET 4: Patients ──────────────────────────────────────────────────
    ws_pat = wb.create_sheet('Patients')
    pat_cols = ['Patient', 'Total Invoices', 'Paid', 'Overdue',
                'Total Invoiced (UGX)', 'Total Paid (UGX)', 'Balance (UGX)']
    make_header(ws_pat, pat_cols)
    pat_summary = (
        Invoice.objects.values('patient__id', 'patient__first_name', 'patient__last_name')
        .annotate(
            total_invoiced=Sum('total_amount'),
            invoice_count=Count('id'),
            paid_count=Count('id', filter=Q(status='paid')),
            overdue_count=Count('id', filter=Q(status='overdue')),
        ).order_by('-total_invoiced')[:200]
    )
    pat_ids = [p['patient__id'] for p in pat_summary]
    paid_map = {
        e['patient_id']: float(e['total'])
        for e in Payment.objects.filter(status='completed', patient_id__in=pat_ids)
        .values('patient_id').annotate(total=Sum('amount'))
    }
    for r, p in enumerate(pat_summary, 2):
        total_inv = float(p['total_invoiced'] or 0)
        total_paid = paid_map.get(p['patient__id'], 0)
        row_data = [
            f"{p['patient__first_name']} {p['patient__last_name']}",
            p['invoice_count'], p['paid_count'], p['overdue_count'],
            total_inv, total_paid, total_inv - total_paid,
        ]
        for c, val in enumerate(row_data, 1):
            ws_pat.cell(row=r, column=c, value=val).border = BORDER
        currency_cols(ws_pat, [5, 6, 7], r, r)
        style_row(ws_pat, r, len(pat_cols), alternate=(r % 2 == 0))
    auto_width(ws_pat)
    ws_pat.auto_filter.ref = f'A1:{get_column_letter(len(pat_cols))}1'

    # ── SHEET 5: Lab Tests ─────────────────────────────────────────────────
    ws_lab = wb.create_sheet('Lab_Tests')
    lab_cols = ['Invoice #', 'Patient', 'Test', 'Date Requested', 'Status', 'Amount (UGX)']
    make_header(ws_lab, lab_cols)
    lab_items = (
        InvoiceLineItem.objects.filter(lab_test_request__isnull=False)
        .select_related('lab_test_request__patient', 'lab_test_request__test', 'invoice')
        .order_by('-invoice__created_at')
    )
    if date_from:
        lab_items = lab_items.filter(invoice__issue_date__gte=date_from)
    if date_to:
        lab_items = lab_items.filter(invoice__issue_date__lte=date_to)
    if patient_q:
        lab_items = lab_items.filter(
            Q(lab_test_request__patient__first_name__icontains=patient_q) |
            Q(lab_test_request__patient__last_name__icontains=patient_q)
        )
    for r, item in enumerate(lab_items, 2):
        ltr = item.lab_test_request
        row_data = [
            item.invoice.invoice_number,
            ltr.patient.get_full_name(),
            ltr.test.name,
            ltr.date_requested.date() if ltr.date_requested else '',
            ltr.get_status_display(),
            float(item.total_amount),
        ]
        for c, val in enumerate(row_data, 1):
            ws_lab.cell(row=r, column=c, value=val).border = BORDER
        currency_cols(ws_lab, [6], r, r)
        style_row(ws_lab, r, len(lab_cols), alternate=(r % 2 == 0))
    auto_width(ws_lab)
    ws_lab.auto_filter.ref = f'A1:{get_column_letter(len(lab_cols))}1'

    # ── SHEET 6: Appointments ──────────────────────────────────────────────
    ws_appt = wb.create_sheet('Appointments')
    appt_cols = ['Invoice #', 'Patient', 'Service', 'Appt Date', 'Provider', 'Amount (UGX)']
    make_header(ws_appt, appt_cols)
    appt_items = (
        InvoiceLineItem.objects.filter(appointment__isnull=False)
        .select_related('appointment__patient', 'appointment__service', 'appointment__provider', 'invoice')
        .order_by('-invoice__created_at')
    )
    if date_from:
        appt_items = appt_items.filter(invoice__issue_date__gte=date_from)
    if date_to:
        appt_items = appt_items.filter(invoice__issue_date__lte=date_to)
    if patient_q:
        appt_items = appt_items.filter(
            Q(appointment__patient__first_name__icontains=patient_q) |
            Q(appointment__patient__last_name__icontains=patient_q)
        )
    for r, item in enumerate(appt_items, 2):
        appt = item.appointment
        row_data = [
            item.invoice.invoice_number,
            appt.patient.get_full_name(),
            appt.service.name,
            appt.appointment_date,
            appt.provider.get_full_name() if appt.provider else '',
            float(item.total_amount),
        ]
        for c, val in enumerate(row_data, 1):
            ws_appt.cell(row=r, column=c, value=val).border = BORDER
        currency_cols(ws_appt, [6], r, r)
        style_row(ws_appt, r, len(appt_cols), alternate=(r % 2 == 0))
    auto_width(ws_appt)
    ws_appt.auto_filter.ref = f'A1:{get_column_letter(len(appt_cols))}1'

    # ── SHEET 7: Services ──────────────────────────────────────────────────
    ws_svc = wb.create_sheet('Services')
    svc_cols = ['Service / Description', 'Times Billed', 'Total Revenue (UGX)', 'Avg Unit Price (UGX)', 'Max Price (UGX)', 'Min Price (UGX)']
    make_header(ws_svc, svc_cols)
    svc_qs = (
        InvoiceLineItem.objects.values('service__name', 'description')
        .annotate(
            total=Sum('total_amount'), count=Count('id'),
            avg=Avg('unit_price'),
            max_p=Max('unit_price'), min_p=Min('unit_price'),
        ).order_by('-total')
    )
    if date_from:
        svc_qs = svc_qs.filter(invoice__issue_date__gte=date_from)
    if date_to:
        svc_qs = svc_qs.filter(invoice__issue_date__lte=date_to)
    for r, s in enumerate(svc_qs, 2):
        row_data = [
            s['service__name'] or s['description'] or '–',
            s['count'],
            float(s['total'] or 0),
            float(s['avg'] or 0),
            float(s['max_p'] or 0),
            float(s['min_p'] or 0),
        ]
        for c, val in enumerate(row_data, 1):
            ws_svc.cell(row=r, column=c, value=val).border = BORDER
        currency_cols(ws_svc, [3, 4, 5, 6], r, r)
        style_row(ws_svc, r, len(svc_cols), alternate=(r % 2 == 0))
    auto_width(ws_svc)
    ws_svc.auto_filter.ref = f'A1:{get_column_letter(len(svc_cols))}1'

    # ── SHEET 8: Aging ─────────────────────────────────────────────────────
    ws_age = wb.create_sheet('Aging')
    age_cols = ['Bucket', 'Invoice Count', 'Total Outstanding (UGX)']
    make_header(ws_age, age_cols)
    
    # Regular invoice aging
    regular_buckets = [
        ('Current (not yet due)',  Invoice.objects.filter(status__in=['sent','overdue'], due_date__gte=today)),
        ('1–30 Days Overdue',      Invoice.objects.filter(status__in=['sent','overdue'], due_date__lt=today, due_date__gte=today-timedelta(days=30))),
        ('31–60 Days Overdue',     Invoice.objects.filter(status__in=['sent','overdue'], due_date__lt=today-timedelta(days=30), due_date__gte=today-timedelta(days=60))),
        ('61–90 Days Overdue',     Invoice.objects.filter(status__in=['sent','overdue'], due_date__lt=today-timedelta(days=60), due_date__gte=today-timedelta(days=90))),
        ('90+ Days Overdue',       Invoice.objects.filter(status__in=['sent','overdue'], due_date__lt=today-timedelta(days=90))),
    ]
    
    # Group invoice aging
    group_buckets = [
        ('Current (Group)',  GroupInvoice.objects.filter(status__in=['sent','overdue'], due_date__gte=today)),
        ('1–30 Days Overdue (Group)',      GroupInvoice.objects.filter(status__in=['sent','overdue'], due_date__lt=today, due_date__gte=today-timedelta(days=30))),
        ('31–60 Days Overdue (Group)',     GroupInvoice.objects.filter(status__in=['sent','overdue'], due_date__lt=today-timedelta(days=30), due_date__gte=today-timedelta(days=60))),
        ('61–90 Days Overdue (Group)',     GroupInvoice.objects.filter(status__in=['sent','overdue'], due_date__lt=today-timedelta(days=60), due_date__gte=today-timedelta(days=90))),
        ('90+ Days Overdue (Group)',       GroupInvoice.objects.filter(status__in=['sent','overdue'], due_date__lt=today-timedelta(days=90))),
    ]
    
    row_num = 2
    for (label, regular_qs), (group_label, group_qs) in zip(regular_buckets, group_buckets):
        regular_total = float(regular_qs.aggregate(Sum('total_amount'))['total_amount__sum'] or 0)
        group_total = float(group_qs.aggregate(Sum('total_amount'))['total_amount__sum'] or 0)
        total_count = regular_qs.count() + group_qs.count()
        total_amount = regular_total + group_total
        row_data = [label, total_count, total_amount]
        for c, val in enumerate(row_data, 1):
            ws_age.cell(row=row_num, column=c, value=val).border = BORDER
        currency_cols(ws_age, [3], row_num, row_num)
        style_row(ws_age, row_num, 3, alternate=(row_num % 2 == 0))
        ws_age.row_dimensions[row_num].height = 20
        row_num += 1
    
    # Add combined totals row
    row_num += 1
    total_outstanding = 0
    total_count = 0
    for _, regular_qs in regular_buckets:
        total_outstanding += float(regular_qs.aggregate(Sum('total_amount'))['total_amount__sum'] or 0)
        total_count += regular_qs.count()
    for _, group_qs in group_buckets:
        total_outstanding += float(group_qs.aggregate(Sum('total_amount'))['total_amount__sum'] or 0)
        total_count += group_qs.count()
    
    for c, val in enumerate(['TOTAL OUTSTANDING', total_count, total_outstanding], 1):
        cell = ws_age.cell(row=row_num, column=c, value=val)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill('solid', fgColor='D4E6F1')
        cell.border = BORDER
    currency_cols(ws_age, [3], row_num, row_num)
    ws_age.row_dimensions[row_num].height = 24
    
    auto_width(ws_age)

    # ── SHEET 9: Group Invoices ────────────────────────────────────────────
    ws_grp = wb.create_sheet('Group_Invoices')
    grp_cols = ['Invoice #', 'Group Name', 'Type', 'Period', 'Status',
                'Subtotal (UGX)', 'Tax (UGX)', 'Total (UGX)', 'Balance Due (UGX)', 
                'Patients', 'Services', 'Due Date']
    make_header(ws_grp, grp_cols)
    group_inv_qs = GroupInvoice.objects.select_related('patient_group').all()
    if date_from:
        group_inv_qs = group_inv_qs.filter(issue_date__gte=date_from)
    if date_to:
        group_inv_qs = group_inv_qs.filter(issue_date__lte=date_to)
    for r, inv in enumerate(group_inv_qs.order_by('-issue_date'), 2):
        try: balance = float(inv.get_balance_due())
        except Exception: balance = float(inv.total_amount)
        period_str = f"{inv.period_start.strftime('%d/%m/%Y')} - {inv.period_end.strftime('%d/%m/%Y')}" if inv.period_start and inv.period_end else '—'
        row_data = [
            inv.invoice_number,
            inv.patient_group.name if inv.patient_group else '—',
            inv.get_invoice_type_display(),
            period_str,
            inv.get_status_display(),
            float(inv.subtotal),
            float(inv.tax_amount),
            float(inv.total_amount),
            balance,
            inv.get_patient_count(),
            inv.get_service_count(),
            inv.due_date,
        ]
        for c, val in enumerate(row_data, 1):
            ws_grp.cell(row=r, column=c, value=val).border = BORDER
        currency_cols(ws_grp, [6, 7, 8, 9], r, r)
        style_row(ws_grp, r, len(grp_cols), alternate=(r % 2 == 0))
    auto_width(ws_grp)
    ws_grp.auto_filter.ref = f'A1:{get_column_letter(len(grp_cols))}1'

    # ── Build and return response ──────────────────────────────────────────
    filename = f"billing_report_{today.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def publish_invoice(request, pk):
    """Generate pixel-perfect invoice PDF via Playwright/Edge, upload to Cloudinary."""
    import re, io as _io, os, tempfile, base64
    import qrcode as qrcode_lib
    import cloudinary, cloudinary.uploader
    from pathlib import Path
    from django.template.loader import render_to_string
    from clinic_settings.models import ClinicSettings

    invoice = get_object_or_404(Invoice, pk=pk)
    line_items = invoice.line_items.all()
    payments = invoice.payments.all()

    try:
        clinic_settings = ClinicSettings.objects.first()
    except Exception:
        clinic_settings = None

    # Build safe filename
    patient_name_raw = invoice.patient.get_full_name() or str(invoice.patient)
    safe_name = re.sub(r'[^\w]', '_', patient_name_raw).strip('_') or f'Patient_{pk}'

    folder = 'invoices'
    public_id_base = f'{safe_name}_invoice_{invoice.invoice_number}'
    cloud_name = cloudinary.config().cloud_name
    predicted_url = f'https://res.cloudinary.com/{cloud_name}/raw/upload/{folder}/{public_id_base}.pdf?dl=1'

    # Generate QR code pointing to Cloudflare R2 public download URL
    from django.conf import settings as _settings
    r2_public = getattr(_settings, 'R2_PUBLIC_URL', '').rstrip('/')
    qr_url = f'{r2_public}/{folder}/{public_id_base}.pdf' if r2_public else predicted_url
    qr = qrcode_lib.QRCode(version=1, box_size=10, border=2)
    qr.add_data(qr_url)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color='black', back_color='white')
    buf = _io.BytesIO()
    qr_img.save(buf, format='PNG')
    qr_base64 = base64.b64encode(buf.getvalue()).decode()

    context = {
        'invoice': invoice,
        'line_items': line_items,
        'payments': payments,
        'clinic_settings': clinic_settings,
        'qr_code': qr_base64,
        'cloud_url': predicted_url,
    }

    base_url = request.build_absolute_uri('/')
    html_string = render_to_string('billing/invoice_pdf.html', context, request=request)
    html_string = html_string.replace('<head>', f'<head><base href="{base_url}">', 1)

    tmp_html = None
    try:
        with tempfile.NamedTemporaryFile(suffix='.html', delete=False, mode='w', encoding='utf-8') as f:
            f.write(html_string)
            tmp_html = f.name
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(channel='msedge')
            page = browser.new_page()
            page.goto(Path(tmp_html).as_uri())
            page.wait_for_load_state('networkidle')
            pdf_bytes = page.pdf(
                print_background=True,
                format='A4',
                margin={'top': '0mm', 'right': '0mm', 'bottom': '0mm', 'left': '0mm'},
            )
            browser.close()
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'PDF generation failed: {e}'}, status=500)
    finally:
        if tmp_html and os.path.exists(tmp_html):
            os.unlink(tmp_html)

    target = request.GET.get('target', 'all')  # 'cloudinary', 'firebase', or 'all'

    cloud_url = invoice.invoice_pdf_url or ''
    gdrive_url = invoice.gdrive_pdf_url or ''

    # Upload to Cloudinary
    if target in ('cloudinary', 'all'):
        try:
            upload_result = cloudinary.uploader.upload(
                _io.BytesIO(pdf_bytes),
                resource_type='raw',
                folder=folder,
                public_id=public_id_base,
                overwrite=True,
                format='pdf',
            )
            secure_url = upload_result.get('secure_url', '')
            cloud_url = (secure_url + '?dl=1') if secure_url else predicted_url
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Cloudinary upload failed: {e}'}, status=500)

    # Upload to Firebase
    if target in ('firebase', 'all'):
        try:
            from clinic_system.gdrive_utils import upload_pdf_to_drive
            gdrive_url = upload_pdf_to_drive(
                pdf_bytes, f'{public_id_base}.pdf', subfolder='invoices',
            )
        except Exception as e:
            if target == 'firebase':
                return JsonResponse({'success': False, 'error': f'Firebase upload failed: {e}'}, status=500)
            import logging
            logging.getLogger(__name__).warning(f'Firebase upload failed for invoice {pk}: {e}')

    invoice.invoice_pdf_url = cloud_url
    invoice.gdrive_pdf_url = gdrive_url
    invoice.save(update_fields=['invoice_pdf_url', 'gdrive_pdf_url'])

    return JsonResponse({'success': True, 'url': cloud_url, 'gdrive_url': gdrive_url})


@login_required
def publish_receipt(request, pk):
    """Generate pixel-perfect payment receipt PDF via Playwright/Edge, upload to Cloudinary."""
    import re, io as _io, os, tempfile, base64
    import qrcode as qrcode_lib
    import cloudinary, cloudinary.uploader
    from pathlib import Path
    from django.template.loader import render_to_string
    from clinic_settings.models import ClinicSettings

    payment = get_object_or_404(Payment, pk=pk)

    try:
        clinic_settings = ClinicSettings.objects.first()
    except Exception:
        clinic_settings = None

    # Build safe filename
    patient_name_raw = payment.patient.get_full_name() or str(payment.patient)
    safe_name = re.sub(r'[^\w]', '_', patient_name_raw).strip('_') or f'Patient_{pk}'

    folder = 'receipts'
    public_id_base = f'{safe_name}_receipt_{payment.payment_id}'
    cloud_name = cloudinary.config().cloud_name
    predicted_url = f'https://res.cloudinary.com/{cloud_name}/raw/upload/{folder}/{public_id_base}.pdf?dl=1'

    # Generate QR code pointing to Cloudflare R2 public download URL
    from django.conf import settings as _settings
    r2_public = getattr(_settings, 'R2_PUBLIC_URL', '').rstrip('/')
    qr_url = f'{r2_public}/{folder}/{public_id_base}.pdf' if r2_public else predicted_url
    qr = qrcode_lib.QRCode(version=1, box_size=10, border=2)
    qr.add_data(qr_url)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color='black', back_color='white')
    buf = _io.BytesIO()
    qr_img.save(buf, format='PNG')
    qr_base64 = base64.b64encode(buf.getvalue()).decode()

    context = {
        'payment': payment,
        'clinic_settings': clinic_settings,
        'qr_code': qr_base64,
        'cloud_url': predicted_url,
    }

    base_url = request.build_absolute_uri('/')
    html_string = render_to_string('billing/payment_receipt.html', context, request=request)
    html_string = html_string.replace('<head>', f'<head><base href="{base_url}">', 1)

    tmp_html = None
    try:
        with tempfile.NamedTemporaryFile(suffix='.html', delete=False, mode='w', encoding='utf-8') as f:
            f.write(html_string)
            tmp_html = f.name
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(channel='msedge')
            page = browser.new_page()
            page.goto(Path(tmp_html).as_uri())
            page.wait_for_load_state('networkidle')
            pdf_bytes = page.pdf(
                print_background=True,
                format='A4',
                margin={'top': '0mm', 'right': '0mm', 'bottom': '0mm', 'left': '0mm'},
            )
            browser.close()
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'PDF generation failed: {e}'}, status=500)
    finally:
        if tmp_html and os.path.exists(tmp_html):
            os.unlink(tmp_html)

    target = request.GET.get('target', 'all')  # 'cloudinary', 'firebase', or 'all'

    cloud_url = payment.receipt_pdf_url or ''
    gdrive_url = payment.gdrive_pdf_url or ''

    # Upload to Cloudinary
    if target in ('cloudinary', 'all'):
        try:
            upload_result = cloudinary.uploader.upload(
                _io.BytesIO(pdf_bytes),
                resource_type='raw',
                folder=folder,
                public_id=public_id_base,
                overwrite=True,
                format='pdf',
            )
            secure_url = upload_result.get('secure_url', '')
            cloud_url = (secure_url + '?dl=1') if secure_url else predicted_url
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Cloudinary upload failed: {e}'}, status=500)

    # Upload to Firebase
    if target in ('firebase', 'all'):
        try:
            from clinic_system.gdrive_utils import upload_pdf_to_drive
            gdrive_url = upload_pdf_to_drive(
                pdf_bytes, f'{public_id_base}.pdf', subfolder='receipts',
            )
        except Exception as e:
            if target == 'firebase':
                return JsonResponse({'success': False, 'error': f'Firebase upload failed: {e}'}, status=500)
            import logging
            logging.getLogger(__name__).warning(f'Firebase upload failed for receipt {pk}: {e}')

    payment.receipt_pdf_url = cloud_url
    payment.gdrive_pdf_url = gdrive_url
    payment.save(update_fields=['receipt_pdf_url', 'gdrive_pdf_url'])

    return JsonResponse({'success': True, 'url': cloud_url, 'gdrive_url': gdrive_url})
