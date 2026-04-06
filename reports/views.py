from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum, Avg, Q
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.core.cache import cache
from datetime import date, timedelta, datetime
from decimal import Decimal
import json
import time
from patients.models import Patient
from appointments.models import Appointment, Service
from billing.models import Invoice, Payment, InvoiceLineItem
from accounts.models import User
from .models import ReportAuditLog, ReportExport, ReportConfiguration
from .utils import ReportAuditMixin, create_report_export, get_report_performance_metrics

class ReportsView(ReportAuditMixin):
    """Base class for report views with audit logging"""
    pass

@login_required
def reports_dashboard(request):
    """Enhanced reports dashboard with audit logging"""
    start_time = time.time()
    audit_mixin = ReportAuditMixin()
    # Date range for reports (default to current month)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date:
        start_date = date.today().replace(day=1)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    if not end_date:
        end_date = date.today()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Patient statistics
    total_patients = Patient.objects.filter(is_active=True).count()
    new_patients = Patient.objects.filter(
        registration_date__range=[start_date, end_date]
    ).count()
    
    # Appointment statistics
    total_appointments = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date]
    ).count()
    
    completed_appointments = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date],
        status='completed'
    ).count()
    
    # Revenue statistics
    total_revenue = Payment.objects.filter(
        payment_date__date__range=[start_date, end_date],
        status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Outstanding balance - calculate from unpaid/partially paid invoices
    unpaid_invoices = Invoice.objects.filter(
        status__in=['sent', 'overdue']
    )
    outstanding_balance = 0
    for invoice in unpaid_invoices:
        outstanding_balance += invoice.get_balance_due()
    
    # Service popularity
    popular_services = Service.objects.annotate(
        appointment_count=Count('appointment')
    ).order_by('-appointment_count')[:5]
    
    service_labels = []
    service_data = []
    for service in popular_services:
        service_labels.append(service.name)
        service_data.append(service.appointment_count)
    
    # Chart data for dashboard - 7 day revenue trend
    revenue_labels = []
    revenue_data = []
    for i in range(7):
        day = end_date - timedelta(days=i)
        revenue = Payment.objects.filter(
            payment_date__date=day,
            status='completed'
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        revenue_labels.insert(0, day.strftime('%m/%d'))
        revenue_data.insert(0, float(revenue))
    
    # Patient demographics
    gender_data = Patient.objects.values('gender').annotate(count=Count('id'))
    gender_labels = []
    gender_counts = []
    for item in gender_data:
        gender_labels.append(item['gender'])
        gender_counts.append(item['count'])
    
    # Department statistics
    from patients.models import Assessment
    physio_count = Assessment.objects.filter(
        department='physiotherapy',
        assessment_date__range=[start_date, end_date]
    ).count()
    
    nutrition_count = Assessment.objects.filter(
        department='nutrition',
        assessment_date__range=[start_date, end_date]
    ).count()
    
    general_count = Assessment.objects.filter(
        department='general',
        assessment_date__range=[start_date, end_date]
    ).count()
    
    # Appointment status breakdown
    scheduled_appointments = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date],
        status='scheduled'
    ).count()
    
    cancelled_appointments = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date],
        status='cancelled'
    ).count()
    
    no_show_appointments = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date],
        status='no_show'
    ).count()
    
    # Calculate execution time and log activity
    execution_time = time.time() - start_time
    audit_mixin.log_report_activity(
        request=request,
        report_type='dashboard',
        report_name='Reports Dashboard',
        action='viewed',
        execution_time=execution_time,
        record_count=total_patients + total_appointments,
        parameters={
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat()
        }
    )
    
    # Get recent audit activities for dashboard
    recent_activities = ReportAuditLog.objects.filter(
        user=request.user
    ).order_by('-timestamp')[:10]
    
    # Get performance metrics
    performance_metrics = get_report_performance_metrics()
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_patients': total_patients,
        'new_patients': new_patients,
        'total_appointments': total_appointments,
        'completed_appointments': completed_appointments,
        'scheduled_appointments': scheduled_appointments,
        'cancelled_appointments': cancelled_appointments,
        'no_show_appointments': no_show_appointments,
        'total_revenue': total_revenue,
        'outstanding_balance': outstanding_balance,
        'popular_services': popular_services,
        'service_labels': json.dumps(service_labels),
        'service_data': json.dumps(service_data),
        'revenue_labels': json.dumps(revenue_labels),
        'revenue_data': json.dumps(revenue_data),
        'gender_labels': json.dumps(gender_labels),
        'gender_data': json.dumps(gender_counts),
        'physio_count': physio_count,
        'nutrition_count': nutrition_count,
        'general_count': general_count,
        'satisfaction_score': 0,  # Calculate from real data if needed
        'recent_activities': recent_activities,
        'performance_metrics': performance_metrics,
        'execution_time': round(execution_time, 2),
    }
    return render(request, 'reports/dashboard.html', context)

def _parse_patient_dates(request):
    """Parse date range from request GET params for patient reports."""
    period = request.GET.get('period', 'this_month')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    end_date = date.today()

    if period == 'custom' and date_from and date_to:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    elif period == 'last_month':
        start_date = (end_date.replace(day=1) - timedelta(days=1)).replace(day=1)
        end_date = end_date.replace(day=1) - timedelta(days=1)
    elif period == 'this_quarter':
        quarter = (end_date.month - 1) // 3 + 1
        start_date = date(end_date.year, (quarter - 1) * 3 + 1, 1)
    elif period == 'last_quarter':
        quarter = (end_date.month - 1) // 3 + 1
        if quarter == 1:
            start_date = date(end_date.year - 1, 10, 1)
            end_date = date(end_date.year - 1, 12, 31)
        else:
            start_date = date(end_date.year, (quarter - 2) * 3 + 1, 1)
            end_date = date(end_date.year, (quarter - 1) * 3, 1) - timedelta(days=1)
    elif period == 'this_year':
        start_date = date(end_date.year, 1, 1)
    elif period == 'last_year':
        start_date = date(end_date.year - 1, 1, 1)
        end_date = date(end_date.year - 1, 12, 31)
    else:  # this_month
        start_date = end_date.replace(day=1)
    return start_date, end_date, period, date_from, date_to


def _get_patient_demographics(patients, start_date, end_date):
    """Gather patient demographic statistics."""
    from patients.models import Assessment

    total_patients = patients.count()
    new_patients = patients.filter(registration_date__range=[start_date, end_date]).count()

    # Average age
    ages = []
    age_buckets = [0, 0, 0, 0, 0]  # 0-18, 19-35, 36-55, 56-75, 76+
    for p in patients:
        try:
            a = p.get_age()
            if a and isinstance(a, (int, float)):
                a = int(a)
                ages.append(a)
                if a <= 18: age_buckets[0] += 1
                elif a <= 35: age_buckets[1] += 1
                elif a <= 55: age_buckets[2] += 1
                elif a <= 75: age_buckets[3] += 1
                else: age_buckets[4] += 1
        except (ValueError, TypeError, AttributeError):
            continue
    average_age = round(sum(ages) / len(ages), 1) if ages else 0
    age_labels = ['0-18', '19-35', '36-55', '56-75', '76+']

    # Gender distribution
    gender_qs = patients.values('gender').annotate(count=Count('id'))
    gender_labels = [dict(Patient.GENDER_CHOICES).get(i['gender'], i['gender'] or 'Unknown') for i in gender_qs]
    gender_data = [i['count'] for i in gender_qs]

    # Blood type distribution
    blood_qs = patients.exclude(blood_type='').values('blood_type').annotate(count=Count('id')).order_by('-count')
    blood_labels = [i['blood_type'] for i in blood_qs]
    blood_data = [i['count'] for i in blood_qs]

    # Registration trend (6 months)
    trend_labels, trend_data = [], []
    for i in range(6):
        ms = (end_date.replace(day=1) - timedelta(days=i * 30)).replace(day=1)
        me = (ms + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        cnt = Patient.objects.filter(registration_date__range=[ms, me]).count()
        trend_labels.insert(0, ms.strftime('%b %Y'))
        trend_data.insert(0, cnt)

    # Insurance breakdown
    ins_qs = patients.exclude(insurance_provider='').values('insurance_provider').annotate(count=Count('id')).order_by('-count')[:10]
    insurance_labels = [i['insurance_provider'] for i in ins_qs]
    insurance_data = [i['count'] for i in ins_qs]
    insured_count = patients.exclude(insurance_provider='').count()

    # Patient groups
    group_qs = patients.exclude(patient_group=None).values('patient_group__name').annotate(count=Count('id')).order_by('-count')

    # Visit reasons
    visit_qs = patients.exclude(reason_for_visit='').values('reason_for_visit').annotate(count=Count('id')).order_by('-count')
    visit_reasons = [{'reason': dict(Patient.VISIT_REASON_CHOICES).get(i['reason_for_visit'], i['reason_for_visit']), 'count': i['count']} for i in visit_qs]

    # Department stats
    physio_patients = patients.filter(assessments__department='physiotherapy').distinct().count()
    nutrition_patients = patients.filter(assessments__department='nutrition').distinct().count()
    general_patients = patients.filter(assessments__department='general').distinct().count()
    if physio_patients == 0 and nutrition_patients == 0 and general_patients == 0:
        physio_patients = patients.filter(appointments__service__category__icontains='physiotherapy').distinct().count()
        nutrition_patients = patients.filter(appointments__service__category__icontains='nutrition').distinct().count()
        general_patients = patients.filter(appointments__service__category__icontains='general').distinct().count()

    # Provider stats
    provider_stats = Assessment.objects.filter(
        assessment_date__range=[start_date, end_date]
    ).values('assessed_by__first_name', 'assessed_by__last_name', 'department').annotate(
        patient_count=Count('patient', distinct=True), assessment_count=Count('id')
    ).order_by('-patient_count')
    provider_data = []
    for s in provider_stats:
        if s['assessed_by__first_name']:
            provider_data.append({
                'name': f"{s['assessed_by__first_name']} {s['assessed_by__last_name']}",
                'department': (s['department'] or '').title() or 'N/A',
                'patient_count': s['patient_count'], 'assessment_count': s['assessment_count'],
            })

    return {
        'total_patients': total_patients, 'new_patients': new_patients,
        'average_age': average_age,
        'age_labels': age_labels, 'age_data': age_buckets,
        'gender_labels': gender_labels, 'gender_data': gender_data,
        'blood_labels': blood_labels, 'blood_data': blood_data,
        'trend_labels': trend_labels, 'trend_data': trend_data,
        'insurance_labels': insurance_labels, 'insurance_data': insurance_data,
        'insured_count': insured_count, 'uninsured_count': total_patients - insured_count,
        'patient_groups': list(group_qs), 'visit_reasons': visit_reasons,
        'physio_patients': physio_patients, 'nutrition_patients': nutrition_patients,
        'general_patients': general_patients,
        'provider_data': provider_data,
    }


def _get_patient_labtest_stats(patients, start_date, end_date):
    """Gather patient-by-lab-test statistics."""
    from laboratory.models import LabTestRequest, LabTestResult

    req_q = LabTestRequest.objects.filter(patient__in=patients, date_requested__date__range=[start_date, end_date])
    total_requests = req_q.count()
    completed = req_q.filter(status='completed').count()
    pending = req_q.filter(status__in=['requested', 'sample_collected', 'in_progress']).count()
    cancelled = req_q.filter(status='cancelled').count()
    completion_rate = round(completed / total_requests * 100, 1) if total_requests else 0

    # Patients with tests
    patients_with_tests = req_q.values('patient').distinct().count()
    patients_without_tests = patients.count() - patients_with_tests

    # Top tests
    top_tests = list(req_q.values('test__name', 'test__category').annotate(cnt=Count('id')).order_by('-cnt')[:10])

    # Tests by category
    by_category = list(req_q.values('test__category').annotate(cnt=Count('id')).order_by('-cnt'))

    # By priority
    by_priority = {i['priority']: i['cnt'] for i in req_q.values('priority').annotate(cnt=Count('id'))}

    # Results stats
    results_q = LabTestResult.objects.filter(request__patient__in=patients, date_reported__date__range=[start_date, end_date])
    total_results = results_q.count()
    abnormal_results = results_q.filter(is_abnormal=True).count()
    abnormal_rate = round(abnormal_results / total_results * 100, 1) if total_results else 0

    # Monthly trend
    monthly_labels, monthly_data = [], []
    for i in range(6):
        ms = (end_date.replace(day=1) - timedelta(days=i * 30)).replace(day=1)
        me = (ms + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        cnt = LabTestRequest.objects.filter(patient__in=patients, date_requested__date__range=[ms, me]).count()
        monthly_labels.insert(0, ms.strftime('%b %Y'))
        monthly_data.insert(0, cnt)

    # Top patients by test count
    top_patients_lab = list(req_q.values(
        'patient__patient_id', 'patient__first_name', 'patient__last_name'
    ).annotate(cnt=Count('id')).order_by('-cnt')[:10])

    return {
        'total_requests': total_requests, 'completed': completed,
        'pending': pending, 'cancelled': cancelled,
        'completion_rate': completion_rate,
        'patients_with_tests': patients_with_tests,
        'patients_without_tests': patients_without_tests,
        'top_tests': top_tests, 'by_category': by_category,
        'by_priority': by_priority,
        'total_results': total_results, 'abnormal_results': abnormal_results,
        'abnormal_rate': abnormal_rate,
        'monthly_labels': monthly_labels, 'monthly_data': monthly_data,
        'top_patients': top_patients_lab,
    }


def _get_patient_medication_stats(patients, start_date, end_date):
    """Gather patient-by-medication/prescription statistics."""
    from pharmacy.models import Prescription, PrescriptionItem, StockMovement, Medication
    from django.db.models import F

    presc_q = Prescription.objects.filter(patient__in=patients, prescribed_date__date__range=[start_date, end_date])
    total_prescriptions = presc_q.count()
    dispensed = presc_q.filter(status='dispensed').count()
    pending = presc_q.filter(status='pending').count()
    cancelled = presc_q.filter(status='cancelled').count()
    dispensed_rate = round(dispensed / total_prescriptions * 100, 1) if total_prescriptions else 0

    patients_with_rx = presc_q.values('patient').distinct().count()
    patients_without_rx = patients.count() - patients_with_rx

    # Top prescribed medications (from PrescriptionItem)
    top_meds = list(PrescriptionItem.objects.filter(
        prescription__patient__in=patients,
        prescription__prescribed_date__date__range=[start_date, end_date]
    ).values('medication__name', 'medication__form').annotate(
        cnt=Count('id'), total_qty=Sum('quantity')
    ).order_by('-cnt')[:10])

    # Sales to patients (StockMovement out referencing SALE with patient prescriptions)
    sales_q = StockMovement.objects.filter(
        created_at__date__range=[start_date, end_date],
        movement_type='out', reference__icontains='SALE'
    )
    total_sales_qty = sales_q.aggregate(Sum('quantity'))['quantity__sum'] or 0
    total_sales_revenue = sales_q.annotate(rev=F('quantity') * F('batch__selling_price')).aggregate(total=Sum('rev'))['total'] or 0

    # Prescriptions by prescriber
    by_prescriber = list(presc_q.values(
        'prescribed_by__first_name', 'prescribed_by__last_name'
    ).annotate(cnt=Count('id')).order_by('-cnt')[:10])

    # Monthly prescription trend
    monthly_labels, monthly_data = [], []
    for i in range(6):
        ms = (end_date.replace(day=1) - timedelta(days=i * 30)).replace(day=1)
        me = (ms + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        cnt = Prescription.objects.filter(patient__in=patients, prescribed_date__date__range=[ms, me]).count()
        monthly_labels.insert(0, ms.strftime('%b %Y'))
        monthly_data.insert(0, cnt)

    # Top patients by prescription count
    top_patients_rx = list(presc_q.values(
        'patient__patient_id', 'patient__first_name', 'patient__last_name'
    ).annotate(cnt=Count('id')).order_by('-cnt')[:10])

    return {
        'total_prescriptions': total_prescriptions,
        'dispensed': dispensed, 'pending': pending, 'cancelled': cancelled,
        'dispensed_rate': dispensed_rate,
        'patients_with_rx': patients_with_rx, 'patients_without_rx': patients_without_rx,
        'top_meds': top_meds, 'by_prescriber': by_prescriber,
        'total_sales_qty': total_sales_qty, 'total_sales_revenue': float(total_sales_revenue),
        'monthly_labels': monthly_labels, 'monthly_data': monthly_data,
        'top_patients': top_patients_rx,
    }


def _get_patient_billing_stats(patients, start_date, end_date):
    """Gather patient-by-billing statistics."""
    inv_q = Invoice.objects.filter(patient__in=patients, created_at__date__range=[start_date, end_date])
    total_invoices = inv_q.count()
    total_billed = inv_q.aggregate(Sum('total_amount'))['total_amount__sum'] or 0

    pay_q = Payment.objects.filter(invoice__patient__in=patients, payment_date__date__range=[start_date, end_date], status='completed')
    total_paid = pay_q.aggregate(Sum('amount'))['amount__sum'] or 0
    outstanding = float(total_billed) - float(total_paid)

    patients_with_inv = inv_q.values('patient').distinct().count()
    avg_invoice = inv_q.aggregate(Avg('total_amount'))['total_amount__avg'] or 0

    # By status
    by_status = list(inv_q.values('status').annotate(cnt=Count('id'), total=Sum('total_amount')).order_by('-cnt'))

    # By payment method
    by_pm = list(pay_q.values('payment_method').annotate(total=Sum('amount'), cnt=Count('id')).order_by('-total'))

    # Top patients by billing
    top_patients_bill = list(inv_q.values(
        'patient__patient_id', 'patient__first_name', 'patient__last_name'
    ).annotate(total=Sum('total_amount'), cnt=Count('id')).order_by('-total')[:10])

    # Monthly billing trend
    monthly_labels, monthly_billed, monthly_paid = [], [], []
    for i in range(6):
        ms = (end_date.replace(day=1) - timedelta(days=i * 30)).replace(day=1)
        me = (ms + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        b = Invoice.objects.filter(patient__in=patients, created_at__date__range=[ms, me]).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        p = Payment.objects.filter(invoice__patient__in=patients, payment_date__date__range=[ms, me], status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
        monthly_labels.insert(0, ms.strftime('%b %Y'))
        monthly_billed.insert(0, float(b))
        monthly_paid.insert(0, float(p))

    return {
        'total_invoices': total_invoices, 'total_billed': float(total_billed),
        'total_paid': float(total_paid), 'outstanding': outstanding,
        'patients_with_inv': patients_with_inv,
        'avg_invoice': float(avg_invoice),
        'by_status': by_status, 'by_pm': by_pm,
        'top_patients': top_patients_bill,
        'monthly_labels': monthly_labels, 'monthly_billed': monthly_billed, 'monthly_paid': monthly_paid,
    }


def _get_patient_appointment_stats(patients, start_date, end_date):
    """Gather patient-by-appointment statistics."""
    apt_q = Appointment.objects.filter(patient__in=patients, appointment_date__range=[start_date, end_date])
    total = apt_q.count()
    completed = apt_q.filter(status='completed').count()
    scheduled = apt_q.filter(status='scheduled').count()
    cancelled = apt_q.filter(status='cancelled').count()
    no_show = apt_q.filter(status='no_show').count()
    completion_rate = round(completed / total * 100, 1) if total else 0
    cancellation_rate = round(cancelled / total * 100, 1) if total else 0

    patients_with_appts = apt_q.values('patient').distinct().count()

    # By service
    by_service = list(apt_q.values('service__name').annotate(cnt=Count('id')).order_by('-cnt')[:10])

    # By provider
    by_provider = list(apt_q.values('provider__first_name', 'provider__last_name').annotate(
        total=Count('id'), comp=Count('id', filter=Q(status='completed'))
    ).order_by('-total')[:10])

    # Top patients by appointment count
    top_patients_appt = list(apt_q.values(
        'patient__patient_id', 'patient__first_name', 'patient__last_name'
    ).annotate(cnt=Count('id')).order_by('-cnt')[:10])

    # Monthly trend
    monthly_labels, monthly_data = [], []
    for i in range(6):
        ms = (end_date.replace(day=1) - timedelta(days=i * 30)).replace(day=1)
        me = (ms + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        cnt = Appointment.objects.filter(patient__in=patients, appointment_date__range=[ms, me]).count()
        monthly_labels.insert(0, ms.strftime('%b %Y'))
        monthly_data.insert(0, cnt)

    return {
        'total': total, 'completed': completed, 'scheduled': scheduled,
        'cancelled': cancelled, 'no_show': no_show,
        'completion_rate': completion_rate, 'cancellation_rate': cancellation_rate,
        'patients_with_appts': patients_with_appts,
        'by_service': by_service, 'by_provider': by_provider,
        'top_patients': top_patients_appt,
        'monthly_labels': monthly_labels, 'monthly_data': monthly_data,
    }


@login_required
def patient_reports(request):
    from clinic_settings.models import EnabledModule

    start_date, end_date, period, date_from, date_to = _parse_patient_dates(request)
    enabled = EnabledModule.get_enabled_modules()
    gender_filter = request.GET.get('gender', '')

    # Base patient queryset
    patients = Patient.objects.filter(is_active=True)
    if gender_filter:
        patients = patients.filter(gender=gender_filter)

    # Build tab list based on enabled modules
    app_tabs = [{'key': 'demographics', 'label': 'Demographics', 'icon': 'bi-people-fill'}]
    if 'laboratory' in enabled:
        app_tabs.append({'key': 'labtests', 'label': 'Lab Tests', 'icon': 'bi-droplet-half'})
    if 'pharmacy' in enabled:
        app_tabs.append({'key': 'medications', 'label': 'Medications', 'icon': 'bi-capsule'})
    if 'billing' in enabled:
        app_tabs.append({'key': 'billing', 'label': 'Billing', 'icon': 'bi-credit-card'})
    if 'appointments' in enabled:
        app_tabs.append({'key': 'appointments', 'label': 'Appointments', 'icon': 'bi-calendar-check'})

    # Gather stats per section
    demographics = _get_patient_demographics(patients, start_date, end_date)
    labtests = _get_patient_labtest_stats(patients, start_date, end_date) if 'laboratory' in enabled else {}
    medications = _get_patient_medication_stats(patients, start_date, end_date) if 'pharmacy' in enabled else {}
    billing = _get_patient_billing_stats(patients, start_date, end_date) if 'billing' in enabled else {}
    appointments = _get_patient_appointment_stats(patients, start_date, end_date) if 'appointments' in enabled else {}

    # Form choices for the custom report generator modal
    from patients.models import PatientGroup
    form_choices = {
        'genders': Patient.GENDER_CHOICES,
        'blood_types': Patient.BLOOD_TYPE_CHOICES,
        'visit_reasons': Patient.VISIT_REASON_CHOICES,
        'patient_groups': PatientGroup.objects.filter(is_active=True),
        'providers': User.objects.filter(is_active=True).order_by('first_name'),
    }
    if 'laboratory' in enabled:
        from laboratory.models import LabTest as LabTestType, TestCategory
        form_choices['lab_tests'] = LabTestType.objects.filter(is_active=True).order_by('name')
        form_choices['lab_categories'] = TestCategory.objects.filter(is_active=True)
    if 'pharmacy' in enabled:
        from pharmacy.models import Medication
        form_choices['medications_list'] = Medication.objects.filter(is_active=True).order_by('name')
        form_choices['med_forms'] = Medication.FORM_CHOICES
    if 'billing' in enabled:
        form_choices['invoice_statuses'] = Invoice.STATUS_CHOICES
        form_choices['payment_methods'] = Payment.PAYMENT_METHODS
    if 'appointments' in enabled:
        form_choices['services'] = Service.objects.filter(is_active=True).order_by('name')
        form_choices['service_categories'] = Service.SERVICE_CATEGORIES
        form_choices['appt_statuses'] = Appointment.STATUS_CHOICES

    context = {
        'period': period, 'date_from': date_from, 'date_to': date_to,
        'start_date': start_date, 'end_date': end_date,
        'gender_filter': gender_filter,
        'app_tabs': app_tabs, 'enabled_modules': enabled,
        'demographics': demographics,
        'labtests': labtests, 'medications': medications,
        'billing': billing, 'appointments': appointments,
        # JSON for charts
        'demographics_json': json.dumps(demographics, default=str),
        'labtests_json': json.dumps(labtests, default=str) if labtests else '{}',
        'medications_json': json.dumps(medications, default=str) if medications else '{}',
        'billing_json': json.dumps(billing, default=str) if billing else '{}',
        'appointments_json': json.dumps(appointments, default=str) if appointments else '{}',
        # Form choices for custom report modal
        **form_choices,
    }
    return render(request, 'reports/patient_reports.html', context)


def _add_patient_report_charts(wb, demo, lab, meds, bill, appt):
    """Add well-designed charts to every sheet of the patient report workbook."""
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList

    def _write_data(ws, labels, value_lists, col_start=11):
        """Write chart source data in far-right columns and return (base_row, end_row, col_start)."""
        base = ws.max_row + 3
        ws.cell(row=base, column=col_start, value='Category')
        for j, name in enumerate(value_lists.keys()):
            ws.cell(row=base, column=col_start + 1 + j, value=name)
        for i, lbl in enumerate(labels):
            ws.cell(row=base + 1 + i, column=col_start, value=str(lbl))
            for j, vals in enumerate(value_lists.values()):
                ws.cell(row=base + 1 + i, column=col_start + 1 + j, value=vals[i] if i < len(vals) else 0)
        return base, base + len(labels), col_start

    def _pie(ws, title, labels, values, anchor, w=13, h=9):
        if not labels or not values or not any(values):
            return
        base, end, cs = _write_data(ws, labels, {'Value': values})
        chart = PieChart()
        chart.title = title
        chart.style = 10
        chart.width, chart.height = w, h
        chart.add_data(Reference(ws, min_col=cs + 1, min_row=base + 1, max_row=end))
        chart.set_categories(Reference(ws, min_col=cs, min_row=base + 1, max_row=end))
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True
        chart.dataLabels.showVal = False
        ws.add_chart(chart, anchor)

    def _bar(ws, title, labels, values, anchor, w=14, h=9):
        if not labels or not values:
            return
        base, end, cs = _write_data(ws, labels, {'Value': values})
        chart = BarChart()
        chart.type = 'col'
        chart.style = 10
        chart.title = title
        chart.width, chart.height = w, h
        chart.add_data(Reference(ws, min_col=cs + 1, min_row=base, max_row=end), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=cs, min_row=base + 1, max_row=end))
        chart.shape = 4
        ws.add_chart(chart, anchor)

    def _line(ws, title, labels, series_dict, anchor, w=15, h=10):
        if not labels:
            return
        base, end, cs = _write_data(ws, labels, series_dict)
        chart = LineChart()
        chart.style = 10
        chart.title = title
        chart.width, chart.height = w, h
        for j in range(len(series_dict)):
            chart.add_data(Reference(ws, min_col=cs + 1 + j, min_row=base, max_row=end), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=cs, min_row=base + 1, max_row=end))
        ws.add_chart(chart, anchor)

    # ── Overview ──
    ws = wb['Overview']
    _pie(ws, 'Patients by Department',
         ['Physiotherapy', 'Nutrition', 'General'],
         [demo['physio_patients'], demo['nutrition_patients'], demo['general_patients']], 'E4')

    # ── Demographics ──
    if 'Demographics' in wb.sheetnames:
        ws = wb['Demographics']
        _bar(ws, 'Age Distribution', demo['age_labels'], demo['age_data'], 'E4')
        _pie(ws, 'Gender Distribution', demo['gender_labels'], demo['gender_data'], 'E20')
        if demo.get('blood_labels'):
            _pie(ws, 'Blood Type', demo['blood_labels'], demo['blood_data'], 'L20')
        if demo.get('insurance_labels'):
            _pie(ws, 'Insurance Breakdown', demo['insurance_labels'], demo['insurance_data'], 'L4')
        _line(ws, 'Registration Trend', demo['trend_labels'], {'New Registrations': demo['trend_data']}, 'E36')

    # ── Lab Tests ──
    if lab and 'Lab Tests' in wb.sheetnames:
        ws = wb['Lab Tests']
        _pie(ws, 'Test Status', ['Completed', 'Pending', 'Cancelled'],
             [lab['completed'], lab['pending'], lab['cancelled']], 'E4')
        _line(ws, 'Monthly Lab Requests', lab['monthly_labels'], {'Requests': lab['monthly_data']}, 'E20')
        if lab.get('by_category'):
            _pie(ws, 'Tests by Category',
                 [c.get('test__category', 'N/A') for c in lab['by_category']],
                 [c.get('cnt', 0) for c in lab['by_category']], 'L4')
        if lab.get('top_tests'):
            _bar(ws, 'Top Requested Tests',
                 [t.get('test__name', '')[:20] for t in lab['top_tests'][:7]],
                 [t.get('cnt', 0) for t in lab['top_tests'][:7]], 'L20')

    # ── Medications ──
    if meds and 'Medications' in wb.sheetnames:
        ws = wb['Medications']
        _pie(ws, 'Prescription Status', ['Dispensed', 'Pending', 'Cancelled'],
             [meds['dispensed'], meds['pending'], meds['cancelled']], 'E4')
        _line(ws, 'Monthly Prescriptions', meds['monthly_labels'], {'Prescriptions': meds['monthly_data']}, 'E20')
        if meds.get('top_meds'):
            _bar(ws, 'Top Medications',
                 [m.get('medication__name', '')[:20] for m in meds['top_meds'][:7]],
                 [m.get('cnt', 0) for m in meds['top_meds'][:7]], 'L4')

    # ── Billing ──
    if bill and 'Billing' in wb.sheetnames:
        ws = wb['Billing']
        _line(ws, 'Billing vs Payments', bill['monthly_labels'],
              {'Billed': bill['monthly_billed'], 'Paid': bill['monthly_paid']}, 'E4')
        if bill.get('by_pm'):
            _pie(ws, 'Payment Methods',
                 [p.get('payment_method', 'N/A') for p in bill['by_pm']],
                 [float(p.get('total', 0) or 0) for p in bill['by_pm']], 'E20')
        if bill.get('by_status'):
            _pie(ws, 'Invoices by Status',
                 [s.get('status', '') for s in bill['by_status']],
                 [s.get('cnt', 0) for s in bill['by_status']], 'L4')

    # ── Appointments ──
    if appt and 'Appointments' in wb.sheetnames:
        ws = wb['Appointments']
        _pie(ws, 'Appointment Status', ['Completed', 'Scheduled', 'Cancelled', 'No Show'],
             [appt['completed'], appt['scheduled'], appt['cancelled'], appt['no_show']], 'E4')
        _line(ws, 'Monthly Appointments', appt['monthly_labels'], {'Appointments': appt['monthly_data']}, 'E20')
        if appt.get('by_service'):
            _bar(ws, 'By Service',
                 [s.get('service__name', 'N/A')[:20] for s in appt['by_service'][:7]],
                 [s.get('cnt', 0) for s in appt['by_service'][:7]], 'L4')


@login_required
def download_patient_report(request):
    """Generate multi-sheet Excel patient report with charts."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from clinic_settings.models import EnabledModule

    start_date, end_date, period, date_from, date_to = _parse_patient_dates(request)
    enabled = EnabledModule.get_enabled_modules()
    gender_filter = request.GET.get('gender', '')

    patients = Patient.objects.filter(is_active=True)
    if gender_filter:
        patients = patients.filter(gender=gender_filter)

    demo = _get_patient_demographics(patients, start_date, end_date)
    lab = _get_patient_labtest_stats(patients, start_date, end_date) if 'laboratory' in enabled else {}
    meds = _get_patient_medication_stats(patients, start_date, end_date) if 'pharmacy' in enabled else {}
    bill = _get_patient_billing_stats(patients, start_date, end_date) if 'billing' in enabled else {}
    appt = _get_patient_appointment_stats(patients, start_date, end_date) if 'appointments' in enabled else {}

    wb = Workbook()
    title_font = Font(name='Arial', size=14, bold=True, color='1B4F72')
    sub_font = Font(name='Arial', size=11, bold=True, color='2E86C1')
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')
    money_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def write_header(ws, row, cols):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        return row + 1

    def write_row(ws, row, vals, fmt=None):
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = thin_border
            if fmt and c in fmt:
                cell.number_format = fmt[c]
        return row + 1

    def auto_width(ws, cols):
        for c in range(1, cols + 1):
            max_len = 12
            for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)) + 2)
            ws.column_dimensions[get_column_letter(c)].width = min(max_len, 40)

    # ── OVERVIEW SHEET ──
    ws_o = wb.active
    ws_o.title = 'Overview'
    ws_o.cell(row=1, column=1, value='Patient Report - Overview').font = title_font
    ws_o.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}').font = Font(name='Arial', size=10, italic=True)
    r = 4
    ws_o.cell(row=r, column=1, value='Summary Statistics').font = sub_font; r += 1
    r = write_header(ws_o, r, ['Metric', 'Value'])
    for k, v in [
        ('Total Patients', demo['total_patients']),
        ('New Patients (period)', demo['new_patients']),
        ('Average Age', demo['average_age']),
        ('Insured Patients', demo['insured_count']),
        ('Uninsured Patients', demo['uninsured_count']),
        ('Physiotherapy Patients', demo['physio_patients']),
        ('Nutrition Patients', demo['nutrition_patients']),
        ('General Patients', demo['general_patients']),
    ]:
        r = write_row(ws_o, r, [k, v])

    r += 1
    ws_o.cell(row=r, column=1, value='Sheets in this Report').font = sub_font; r += 1
    sheets = ['Demographics']
    if lab: sheets.append('Lab Tests')
    if meds: sheets.append('Medications')
    if bill: sheets.append('Billing')
    if appt: sheets.append('Appointments')
    for s in sheets:
        ws_o.cell(row=r, column=1, value=f'→ {s}'); r += 1
    auto_width(ws_o, 2)

    # ── DEMOGRAPHICS SHEET ──
    ws_d = wb.create_sheet('Demographics')
    ws_d.cell(row=1, column=1, value='Patient Demographics').font = title_font
    ws_d.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}').font = Font(name='Arial', size=10, italic=True)
    r = 4
    ws_d.cell(row=r, column=1, value='Age Distribution').font = sub_font; r += 1
    r = write_header(ws_d, r, ['Age Group', 'Count'])
    for lbl, cnt in zip(demo['age_labels'], demo['age_data']):
        r = write_row(ws_d, r, [lbl, cnt])
    r += 1
    ws_d.cell(row=r, column=1, value='Gender Distribution').font = sub_font; r += 1
    r = write_header(ws_d, r, ['Gender', 'Count'])
    for lbl, cnt in zip(demo['gender_labels'], demo['gender_data']):
        r = write_row(ws_d, r, [lbl, cnt])
    r += 1
    if demo['blood_labels']:
        ws_d.cell(row=r, column=1, value='Blood Type Distribution').font = sub_font; r += 1
        r = write_header(ws_d, r, ['Blood Type', 'Count'])
        for lbl, cnt in zip(demo['blood_labels'], demo['blood_data']):
            r = write_row(ws_d, r, [lbl, cnt])
        r += 1
    if demo['insurance_labels']:
        ws_d.cell(row=r, column=1, value='Insurance Providers').font = sub_font; r += 1
        r = write_header(ws_d, r, ['Provider', 'Patients'])
        for lbl, cnt in zip(demo['insurance_labels'], demo['insurance_data']):
            r = write_row(ws_d, r, [lbl, cnt])
        r += 1
    if demo['visit_reasons']:
        ws_d.cell(row=r, column=1, value='Visit Reasons').font = sub_font; r += 1
        r = write_header(ws_d, r, ['Reason', 'Count'])
        for vr in demo['visit_reasons']:
            r = write_row(ws_d, r, [vr['reason'], vr['count']])
        r += 1
    ws_d.cell(row=r, column=1, value='Registration Trend').font = sub_font; r += 1
    r = write_header(ws_d, r, ['Month', 'New Registrations'])
    for lbl, cnt in zip(demo['trend_labels'], demo['trend_data']):
        r = write_row(ws_d, r, [lbl, cnt])
    r += 1
    if demo['provider_data']:
        ws_d.cell(row=r, column=1, value='Provider Statistics').font = sub_font; r += 1
        r = write_header(ws_d, r, ['Provider', 'Department', 'Patients', 'Assessments'])
        for p in demo['provider_data']:
            r = write_row(ws_d, r, [p['name'], p['department'], p['patient_count'], p['assessment_count']])
    auto_width(ws_d, 4)

    # ── LAB TESTS SHEET ──
    if lab:
        ws_l = wb.create_sheet('Lab Tests')
        ws_l.cell(row=1, column=1, value='Patients by Lab Tests').font = title_font
        ws_l.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}').font = Font(name='Arial', size=10, italic=True)
        r = 4
        ws_l.cell(row=r, column=1, value='Lab Test Statistics').font = sub_font; r += 1
        r = write_header(ws_l, r, ['Metric', 'Value'])
        for k, v in [
            ('Total Test Requests', lab['total_requests']),
            ('Completed', lab['completed']),
            ('Pending', lab['pending']),
            ('Cancelled', lab['cancelled']),
            ('Completion Rate %', lab['completion_rate']),
            ('Patients with Tests', lab['patients_with_tests']),
            ('Patients without Tests', lab['patients_without_tests']),
            ('Total Results', lab['total_results']),
            ('Abnormal Results', lab['abnormal_results']),
            ('Abnormal Rate %', lab['abnormal_rate']),
        ]:
            r = write_row(ws_l, r, [k, v])
        r += 1
        if lab['top_tests']:
            ws_l.cell(row=r, column=1, value='Top Requested Tests').font = sub_font; r += 1
            r = write_header(ws_l, r, ['Test Name', 'Category', 'Count'])
            for t in lab['top_tests']:
                r = write_row(ws_l, r, [t.get('test__name', ''), t.get('test__category', ''), t.get('cnt', 0)])
            r += 1
        if lab['by_category']:
            ws_l.cell(row=r, column=1, value='Tests by Category').font = sub_font; r += 1
            r = write_header(ws_l, r, ['Category', 'Count'])
            for c_ in lab['by_category']:
                r = write_row(ws_l, r, [c_.get('test__category', 'N/A'), c_.get('cnt', 0)])
            r += 1
        if lab['top_patients']:
            ws_l.cell(row=r, column=1, value='Top Patients by Test Count').font = sub_font; r += 1
            r = write_header(ws_l, r, ['Patient ID', 'Name', 'Tests'])
            for p in lab['top_patients']:
                name = f"{p.get('patient__first_name', '')} {p.get('patient__last_name', '')}".strip()
                r = write_row(ws_l, r, [p.get('patient__patient_id', ''), name, p.get('cnt', 0)])
            r += 1
        ws_l.cell(row=r, column=1, value='Monthly Trend').font = sub_font; r += 1
        r = write_header(ws_l, r, ['Month', 'Requests'])
        for lbl, cnt in zip(lab['monthly_labels'], lab['monthly_data']):
            r = write_row(ws_l, r, [lbl, cnt])
        auto_width(ws_l, 3)

    # ── MEDICATIONS SHEET ──
    if meds:
        ws_m = wb.create_sheet('Medications')
        ws_m.cell(row=1, column=1, value='Patients by Medications').font = title_font
        ws_m.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}').font = Font(name='Arial', size=10, italic=True)
        r = 4
        ws_m.cell(row=r, column=1, value='Prescription Statistics').font = sub_font; r += 1
        r = write_header(ws_m, r, ['Metric', 'Value'])
        for k, v in [
            ('Total Prescriptions', meds['total_prescriptions']),
            ('Dispensed', meds['dispensed']),
            ('Pending', meds['pending']),
            ('Cancelled', meds['cancelled']),
            ('Dispensed Rate %', meds['dispensed_rate']),
            ('Patients with Rx', meds['patients_with_rx']),
            ('Patients without Rx', meds['patients_without_rx']),
            ('Total Sales Qty', meds['total_sales_qty']),
            ('Total Sales Revenue', meds['total_sales_revenue']),
        ]:
            r = write_row(ws_m, r, [k, v])
        r += 1
        if meds['top_meds']:
            ws_m.cell(row=r, column=1, value='Top Prescribed Medications').font = sub_font; r += 1
            r = write_header(ws_m, r, ['Medication', 'Form', 'Prescriptions', 'Total Qty'])
            for m in meds['top_meds']:
                r = write_row(ws_m, r, [m.get('medication__name', ''), m.get('medication__form', ''), m.get('cnt', 0), m.get('total_qty', 0)])
            r += 1
        if meds['by_prescriber']:
            ws_m.cell(row=r, column=1, value='Prescriptions by Prescriber').font = sub_font; r += 1
            r = write_header(ws_m, r, ['Prescriber', 'Count'])
            for p in meds['by_prescriber']:
                name = f"{p.get('prescribed_by__first_name', '')} {p.get('prescribed_by__last_name', '')}".strip()
                r = write_row(ws_m, r, [name, p.get('cnt', 0)])
            r += 1
        if meds['top_patients']:
            ws_m.cell(row=r, column=1, value='Top Patients by Prescription Count').font = sub_font; r += 1
            r = write_header(ws_m, r, ['Patient ID', 'Name', 'Prescriptions'])
            for p in meds['top_patients']:
                name = f"{p.get('patient__first_name', '')} {p.get('patient__last_name', '')}".strip()
                r = write_row(ws_m, r, [p.get('patient__patient_id', ''), name, p.get('cnt', 0)])
            r += 1
        ws_m.cell(row=r, column=1, value='Monthly Trend').font = sub_font; r += 1
        r = write_header(ws_m, r, ['Month', 'Prescriptions'])
        for lbl, cnt in zip(meds['monthly_labels'], meds['monthly_data']):
            r = write_row(ws_m, r, [lbl, cnt])
        auto_width(ws_m, 4)

    # ── BILLING SHEET ──
    if bill:
        ws_b = wb.create_sheet('Billing')
        ws_b.cell(row=1, column=1, value='Patients by Billing').font = title_font
        ws_b.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}').font = Font(name='Arial', size=10, italic=True)
        r = 4
        ws_b.cell(row=r, column=1, value='Billing Statistics').font = sub_font; r += 1
        r = write_header(ws_b, r, ['Metric', 'Value'])
        for k, v in [
            ('Total Invoices', bill['total_invoices']),
            ('Total Billed', bill['total_billed']),
            ('Total Paid', bill['total_paid']),
            ('Outstanding', bill['outstanding']),
            ('Patients with Invoices', bill['patients_with_inv']),
            ('Average Invoice', bill['avg_invoice']),
        ]:
            r = write_row(ws_b, r, [k, v])
        r += 1
        if bill['by_status']:
            ws_b.cell(row=r, column=1, value='Invoices by Status').font = sub_font; r += 1
            r = write_header(ws_b, r, ['Status', 'Count', 'Total Amount'])
            for s in bill['by_status']:
                r = write_row(ws_b, r, [s.get('status', ''), s.get('cnt', 0), float(s.get('total', 0) or 0)], fmt={3: money_fmt})
            r += 1
        if bill['by_pm']:
            ws_b.cell(row=r, column=1, value='Payments by Method').font = sub_font; r += 1
            r = write_header(ws_b, r, ['Method', 'Count', 'Total'])
            for pm in bill['by_pm']:
                r = write_row(ws_b, r, [pm.get('payment_method', ''), pm.get('cnt', 0), float(pm.get('total', 0) or 0)], fmt={3: money_fmt})
            r += 1
        if bill['top_patients']:
            ws_b.cell(row=r, column=1, value='Top Patients by Billing').font = sub_font; r += 1
            r = write_header(ws_b, r, ['Patient ID', 'Name', 'Total Billed', 'Invoices'])
            for p in bill['top_patients']:
                name = f"{p.get('patient__first_name', '')} {p.get('patient__last_name', '')}".strip()
                r = write_row(ws_b, r, [p.get('patient__patient_id', ''), name, float(p.get('total', 0) or 0), p.get('cnt', 0)], fmt={3: money_fmt})
            r += 1
        ws_b.cell(row=r, column=1, value='Monthly Trend').font = sub_font; r += 1
        r = write_header(ws_b, r, ['Month', 'Billed', 'Paid'])
        for lbl, b_, p_ in zip(bill['monthly_labels'], bill['monthly_billed'], bill['monthly_paid']):
            r = write_row(ws_b, r, [lbl, b_, p_], fmt={2: money_fmt, 3: money_fmt})
        auto_width(ws_b, 4)

    # ── APPOINTMENTS SHEET ──
    if appt:
        ws_a = wb.create_sheet('Appointments')
        ws_a.cell(row=1, column=1, value='Patients by Appointments').font = title_font
        ws_a.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}').font = Font(name='Arial', size=10, italic=True)
        r = 4
        ws_a.cell(row=r, column=1, value='Appointment Statistics').font = sub_font; r += 1
        r = write_header(ws_a, r, ['Metric', 'Value'])
        for k, v in [
            ('Total Appointments', appt['total']),
            ('Completed', appt['completed']),
            ('Scheduled', appt['scheduled']),
            ('Cancelled', appt['cancelled']),
            ('No Show', appt['no_show']),
            ('Completion Rate %', appt['completion_rate']),
            ('Cancellation Rate %', appt['cancellation_rate']),
            ('Patients with Appointments', appt['patients_with_appts']),
        ]:
            r = write_row(ws_a, r, [k, v])
        r += 1
        if appt['by_service']:
            ws_a.cell(row=r, column=1, value='By Service').font = sub_font; r += 1
            r = write_header(ws_a, r, ['Service', 'Count'])
            for s in appt['by_service']:
                r = write_row(ws_a, r, [s.get('service__name', 'N/A'), s.get('cnt', 0)])
            r += 1
        if appt['by_provider']:
            ws_a.cell(row=r, column=1, value='By Provider').font = sub_font; r += 1
            r = write_header(ws_a, r, ['Provider', 'Total', 'Completed'])
            for p in appt['by_provider']:
                name = f"{p.get('provider__first_name', '')} {p.get('provider__last_name', '')}".strip() or 'N/A'
                r = write_row(ws_a, r, [name, p.get('total', 0), p.get('comp', 0)])
            r += 1
        if appt['top_patients']:
            ws_a.cell(row=r, column=1, value='Top Patients by Appointment Count').font = sub_font; r += 1
            r = write_header(ws_a, r, ['Patient ID', 'Name', 'Appointments'])
            for p in appt['top_patients']:
                name = f"{p.get('patient__first_name', '')} {p.get('patient__last_name', '')}".strip()
                r = write_row(ws_a, r, [p.get('patient__patient_id', ''), name, p.get('cnt', 0)])
            r += 1
        ws_a.cell(row=r, column=1, value='Monthly Trend').font = sub_font; r += 1
        r = write_header(ws_a, r, ['Month', 'Appointments'])
        for lbl, cnt in zip(appt['monthly_labels'], appt['monthly_data']):
            r = write_row(ws_a, r, [lbl, cnt])
        auto_width(ws_a, 3)

    # Add charts to all sheets
    _add_patient_report_charts(wb, demo, lab, meds, bill, appt)

    # Write response
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Patient_Report_{start_date}_{end_date}.xlsx"'
    return response


def _parse_financial_dates(request):
    """Parse date range from request GET params for financial reports."""
    period = request.GET.get('period', 'this_month')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    end_date = date.today()

    if period == 'custom' and date_from and date_to:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    elif period == 'last_month':
        start_date = (end_date.replace(day=1) - timedelta(days=1)).replace(day=1)
        end_date = end_date.replace(day=1) - timedelta(days=1)
    elif period == 'this_quarter':
        quarter = (end_date.month - 1) // 3 + 1
        start_date = date(end_date.year, (quarter - 1) * 3 + 1, 1)
    elif period == 'last_quarter':
        quarter = (end_date.month - 1) // 3 + 1
        if quarter == 1:
            start_date = date(end_date.year - 1, 10, 1)
            end_date = date(end_date.year - 1, 12, 31)
        else:
            start_date = date(end_date.year, (quarter - 2) * 3 + 1, 1)
            end_date = date(end_date.year, (quarter - 1) * 3, 1) - timedelta(days=1)
    elif period == 'this_year':
        start_date = date(end_date.year, 1, 1)
    elif period == 'last_year':
        start_date = date(end_date.year - 1, 1, 1)
        end_date = date(end_date.year - 1, 12, 31)
    else:
        start_date = end_date.replace(day=1)

    # Previous period for comparison
    period_days = (end_date - start_date).days
    prev_end = start_date - timedelta(days=1)
    prev_start = prev_end - timedelta(days=period_days)

    return start_date, end_date, prev_start, prev_end, period, date_from, date_to


def _get_billing_stats(start_date, end_date, prev_start, prev_end):
    """Gather billing/invoicing statistics."""
    from billing.models import InsuranceClaim
    from django.db.models import F

    payments_query = Payment.objects.filter(
        payment_date__date__range=[start_date, end_date], status='completed'
    )
    services_revenue = payments_query.aggregate(Sum('amount'))['amount__sum'] or 0
    payments_received = services_revenue

    invoices_query = Invoice.objects.filter(created_at__date__range=[start_date, end_date])
    total_invoices = invoices_query.count()
    avg_invoice_value = invoices_query.aggregate(Avg('total_amount'))['total_amount__avg'] or 0

    outstanding_query = Invoice.objects.filter(status__in=['sent', 'overdue'])
    outstanding_amount = outstanding_query.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    collection_rate = (float(payments_received) / float(services_revenue) * 100) if services_revenue > 0 else 0

    # Previous period revenue
    prev_revenue = Payment.objects.filter(
        payment_date__date__range=[prev_start, prev_end], status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    revenue_change = services_revenue - prev_revenue
    revenue_change_pct = (revenue_change / prev_revenue * 100) if prev_revenue > 0 else 0

    # Payment methods breakdown
    pm = payments_query.values('payment_method').annotate(total=Sum('amount'))
    pm_labels = [i['payment_method'] for i in pm]
    pm_data = [float(i['total']) for i in pm]

    # Service revenue by category
    srv = InvoiceLineItem.objects.filter(
        invoice__created_at__date__range=[start_date, end_date],
        invoice__status__in=['sent', 'paid']
    ).values('service__category').annotate(total_revenue=Sum('total_amount')).order_by('-total_revenue')
    srv_labels = []
    srv_data = []
    for i in srv:
        srv_labels.append(dict(Service.SERVICE_CATEGORIES).get(i['service__category'], i['service__category'] or 'Other'))
        srv_data.append(float(i['total_revenue']))

    # Top services
    top_srv = InvoiceLineItem.objects.filter(
        invoice__created_at__date__range=[start_date, end_date],
        invoice__status__in=['sent', 'paid']
    ).values('service__name').annotate(
        total_revenue=Sum('total_amount'), service_count=Count('id')
    ).order_by('-total_revenue')[:5]
    top_services = [{'name': i['service__name'] or 'Unknown', 'total_revenue': float(i['total_revenue'] or 0), 'count': i['service_count']} for i in top_srv]

    # Outstanding aging
    today = date.today()
    o30 = Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__gte=today - timedelta(days=30)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    o60 = Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__range=[today - timedelta(days=60), today - timedelta(days=31)]).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    o90 = Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__range=[today - timedelta(days=90), today - timedelta(days=61)]).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    o90p = Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=90)).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    tot_out = o30 + o60 + o90 + o90p

    # Monthly revenue (current & prev year)
    monthly_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    cy = date.today().year
    current_year_data = []
    previous_year_data = []
    for m in range(1, 13):
        current_year_data.append(float(Payment.objects.filter(payment_date__year=cy, payment_date__month=m, status='completed').aggregate(Sum('amount'))['amount__sum'] or 0))
        previous_year_data.append(float(Payment.objects.filter(payment_date__year=cy - 1, payment_date__month=m, status='completed').aggregate(Sum('amount'))['amount__sum'] or 0))

    # Revenue trend (daily)
    period_days = (end_date - start_date).days + 1
    trend_labels = []
    trend_data = []
    for i in range(min(period_days, 90)):
        day = start_date + timedelta(days=i)
        rev = Payment.objects.filter(payment_date__date=day, status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
        trend_labels.append(day.strftime('%m/%d'))
        trend_data.append(float(rev))

    # Insurance claims
    claims_q = InsuranceClaim.objects.filter(submission_date__range=[start_date, end_date])
    claims_submitted = claims_q.count()
    claims_approved = claims_q.filter(status='approved').count()
    claims_pending = claims_q.filter(status__in=['submitted', 'pending']).count()
    claims_denied = claims_q.filter(status='denied').count()
    claims_total = float(claims_q.aggregate(Sum('claim_amount'))['claim_amount__sum'] or 0)
    claims_reimbursed = float(claims_q.filter(status__in=['approved', 'paid']).aggregate(Sum('approved_amount'))['approved_amount__sum'] or 0)

    # Invoices by status
    inv_status = Invoice.objects.filter(created_at__date__range=[start_date, end_date]).values('status').annotate(cnt=Count('id'))
    inv_by_status = {i['status']: i['cnt'] for i in inv_status}

    return {
        'services_revenue': services_revenue,
        'payments_received': payments_received,
        'total_invoices': total_invoices,
        'avg_invoice_value': avg_invoice_value,
        'outstanding_amount': outstanding_amount,
        'collection_rate': round(collection_rate, 1),
        'prev_revenue': prev_revenue,
        'revenue_change': revenue_change,
        'revenue_change_pct': round(revenue_change_pct, 1),
        'pm_labels': pm_labels, 'pm_data': pm_data,
        'srv_labels': srv_labels, 'srv_data': srv_data,
        'top_services': top_services,
        'o30': o30, 'o60': o60, 'o90': o90, 'o90p': o90p, 'tot_out': tot_out,
        'o30_pct': round((o30 / tot_out * 100) if tot_out else 0, 1),
        'o60_pct': round((o60 / tot_out * 100) if tot_out else 0, 1),
        'o90_pct': round((o90 / tot_out * 100) if tot_out else 0, 1),
        'o90p_pct': round((o90p / tot_out * 100) if tot_out else 0, 1),
        'monthly_labels': monthly_labels,
        'current_year_data': current_year_data,
        'previous_year_data': previous_year_data,
        'trend_labels': trend_labels, 'trend_data': trend_data,
        'claims_submitted': claims_submitted, 'claims_approved': claims_approved,
        'claims_pending': claims_pending, 'claims_denied': claims_denied,
        'claims_total': claims_total, 'claims_reimbursed': claims_reimbursed,
        'inv_by_status': inv_by_status,
    }


def _get_pharmacy_stats(start_date, end_date):
    """Gather pharmacy statistics."""
    from pharmacy.models import StockMovement, Batch, Medication, PurchaseOrder
    from django.db.models import F

    # Sales
    sales_q = StockMovement.objects.filter(
        created_at__date__range=[start_date, end_date], movement_type='out', reference__icontains='SALE'
    )
    sales_revenue = sales_q.annotate(rev=F('quantity') * F('batch__selling_price')).aggregate(total=Sum('rev'))['total'] or 0
    sales_cogs = sales_q.annotate(cost=F('quantity') * F('batch__cost_price')).aggregate(total=Sum('cost'))['total'] or 0
    sales_profit = sales_revenue - sales_cogs
    sales_margin = (sales_profit / sales_revenue * 100) if sales_revenue > 0 else 0
    total_items_sold = sales_q.aggregate(Sum('quantity'))['quantity__sum'] or 0
    total_transactions = sales_q.count()

    # Stock overview
    total_medications = Medication.objects.filter(is_active=True).count()
    low_stock = Medication.objects.filter(is_active=True).count()  # will refine
    low_stock_items = []
    for med in Medication.objects.filter(is_active=True):
        if med.current_stock <= med.reorder_level:
            low_stock_items.append({'name': str(med), 'stock': med.current_stock, 'reorder': med.reorder_level})
    low_stock_count = len(low_stock_items)

    # Expiring batches (within 90 days)
    from django.utils import timezone as tz
    expiring = Batch.objects.filter(is_active=True, expiry_date__lte=tz.now().date() + timedelta(days=90), expiry_date__gt=tz.now().date())
    expired = Batch.objects.filter(is_active=True, expiry_date__lte=tz.now().date())

    # Stock value
    active_batches = Batch.objects.filter(is_active=True, expiry_date__gt=tz.now().date())
    stock_cost_val = sum(b.quantity_remaining * b.cost_price for b in active_batches)
    stock_sell_val = sum(b.quantity_remaining * b.selling_price for b in active_batches)

    # Purchase orders
    po_q = PurchaseOrder.objects.filter(order_date__range=[start_date, end_date])
    po_count = po_q.count()
    po_total = po_q.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    po_by_status = {i['status']: i['cnt'] for i in po_q.values('status').annotate(cnt=Count('id'))}

    # Top selling medications
    top_meds = sales_q.values('batch__medication__name').annotate(
        qty=Sum('quantity'),
        rev=Sum(F('quantity') * F('batch__selling_price'))
    ).order_by('-qty')[:10]

    # Monthly sales trend
    monthly_labels = []
    monthly_data = []
    for i in range(6):
        ms = (end_date.replace(day=1) - timedelta(days=i * 30)).replace(day=1)
        me = (ms + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        rev = StockMovement.objects.filter(
            created_at__date__range=[ms, me], movement_type='out', reference__icontains='SALE'
        ).annotate(r=F('quantity') * F('batch__selling_price')).aggregate(t=Sum('r'))['t'] or 0
        monthly_labels.insert(0, ms.strftime('%b %Y'))
        monthly_data.insert(0, float(rev))

    return {
        'sales_revenue': float(sales_revenue), 'sales_cogs': float(sales_cogs),
        'sales_profit': float(sales_profit), 'sales_margin': round(sales_margin, 1),
        'total_items_sold': total_items_sold, 'total_transactions': total_transactions,
        'total_medications': total_medications, 'low_stock_count': low_stock_count,
        'low_stock_items': low_stock_items[:10],
        'expiring_count': expiring.count(), 'expired_count': expired.count(),
        'stock_cost_val': float(stock_cost_val), 'stock_sell_val': float(stock_sell_val),
        'po_count': po_count, 'po_total': float(po_total), 'po_by_status': po_by_status,
        'top_meds': list(top_meds),
        'monthly_labels': monthly_labels, 'monthly_data': monthly_data,
    }


def _get_laboratory_stats(start_date, end_date):
    """Gather laboratory statistics."""
    from laboratory.models import LabTestRequest, LabTestResult, LabTest

    req_q = LabTestRequest.objects.filter(date_requested__date__range=[start_date, end_date])
    total_requests = req_q.count()
    by_status = {i['status']: i['cnt'] for i in req_q.values('status').annotate(cnt=Count('id'))}
    by_priority = {i['priority']: i['cnt'] for i in req_q.values('priority').annotate(cnt=Count('id'))}

    completed = req_q.filter(status='completed').count()
    pending = req_q.filter(status__in=['requested', 'sample_collected', 'in_progress']).count()
    cancelled = req_q.filter(status='cancelled').count()
    completion_rate = (completed / total_requests * 100) if total_requests > 0 else 0

    # Results stats
    results_q = LabTestResult.objects.filter(date_reported__date__range=[start_date, end_date])
    total_results = results_q.count()
    abnormal_results = results_q.filter(is_abnormal=True).count()
    verified_results = results_q.filter(verified=True).count()

    # Top requested tests
    top_tests = req_q.values('test__name', 'test__category').annotate(cnt=Count('id')).order_by('-cnt')[:10]

    # Tests by category
    by_category = req_q.values('test__category').annotate(cnt=Count('id')).order_by('-cnt')

    # Unique patients tested
    unique_patients = req_q.values('patient').distinct().count()

    # Monthly trend
    monthly_labels = []
    monthly_data = []
    for i in range(6):
        ms = (end_date.replace(day=1) - timedelta(days=i * 30)).replace(day=1)
        me = (ms + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        cnt = LabTestRequest.objects.filter(date_requested__date__range=[ms, me]).count()
        monthly_labels.insert(0, ms.strftime('%b %Y'))
        monthly_data.insert(0, cnt)

    # TAT (turnaround time) - avg days from request to result
    from django.db.models import Avg as DjAvg
    avg_tat = 0
    completed_with_results = req_q.filter(status='completed', result__isnull=False)
    if completed_with_results.exists():
        tat_list = []
        for r in completed_with_results.select_related('result')[:100]:
            if r.result and r.result.date_reported and r.date_requested:
                diff = (r.result.date_reported - r.date_requested).total_seconds() / 3600
                tat_list.append(diff)
        avg_tat = round(sum(tat_list) / len(tat_list), 1) if tat_list else 0

    return {
        'total_requests': total_requests, 'completed': completed,
        'pending': pending, 'cancelled': cancelled,
        'completion_rate': round(completion_rate, 1),
        'by_status': by_status, 'by_priority': by_priority,
        'total_results': total_results, 'abnormal_results': abnormal_results,
        'verified_results': verified_results,
        'top_tests': list(top_tests), 'by_category': list(by_category),
        'unique_patients': unique_patients,
        'monthly_labels': monthly_labels, 'monthly_data': monthly_data,
        'avg_tat_hours': avg_tat,
    }


def _get_budget_stats(start_date, end_date):
    """Gather budget & expense statistics."""
    from budget.models import Expense, ExpenseCategory, Budget

    exp_q = Expense.objects.filter(expense_date__range=[start_date, end_date], status__in=['approved', 'paid'])
    total_expenses = exp_q.aggregate(Sum('amount'))['amount__sum'] or 0
    total_expense_count = exp_q.count()
    pending_expenses = Expense.objects.filter(expense_date__range=[start_date, end_date], status='pending').count()
    rejected_expenses = Expense.objects.filter(expense_date__range=[start_date, end_date], status='rejected').count()

    # By category
    by_cat = exp_q.values('category__name', 'category__color', 'category__icon').annotate(total=Sum('amount'), cnt=Count('id')).order_by('-total')

    # By payment method
    by_pm = exp_q.values('payment_method').annotate(total=Sum('amount'), cnt=Count('id')).order_by('-total')

    # Top vendors
    top_vendors = exp_q.exclude(vendor_name='').values('vendor_name').annotate(total=Sum('amount'), cnt=Count('id')).order_by('-total')[:10]

    # Monthly trend
    monthly_labels = []
    monthly_data = []
    for i in range(6):
        ms = (end_date.replace(day=1) - timedelta(days=i * 30)).replace(day=1)
        me = (ms + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        amt = Expense.objects.filter(expense_date__range=[ms, me], status__in=['approved', 'paid']).aggregate(Sum('amount'))['amount__sum'] or 0
        monthly_labels.insert(0, ms.strftime('%b %Y'))
        monthly_data.insert(0, float(amt))

    # Active budgets
    active_budgets = Budget.objects.filter(status='active', start_date__lte=end_date, end_date__gte=start_date)
    budget_list = []
    for b in active_budgets:
        budget_list.append({
            'name': b.name, 'total': float(b.total_amount),
            'spent': float(b.get_spent_amount()),
            'remaining': float(b.get_remaining_amount()),
            'utilization': round(float(b.get_utilization_percentage()), 1),
        })

    return {
        'total_expenses': float(total_expenses),
        'total_expense_count': total_expense_count,
        'pending_expenses': pending_expenses,
        'rejected_expenses': rejected_expenses,
        'by_category': list(by_cat), 'by_payment_method': list(by_pm),
        'top_vendors': list(top_vendors),
        'monthly_labels': monthly_labels, 'monthly_data': monthly_data,
        'active_budgets': budget_list,
    }


def _get_appointments_stats(start_date, end_date):
    """Gather appointment statistics."""
    apt_q = Appointment.objects.filter(appointment_date__range=[start_date, end_date])
    total = apt_q.count()
    completed = apt_q.filter(status='completed').count()
    scheduled = apt_q.filter(status='scheduled').count()
    cancelled = apt_q.filter(status='cancelled').count()
    no_show = apt_q.filter(status='no_show').count()

    completion_rate = (completed / total * 100) if total > 0 else 0
    cancellation_rate = (cancelled / total * 100) if total > 0 else 0

    # Revenue from completed appointments
    total_revenue = 0
    for a in apt_q.filter(status='completed').select_related('service'):
        if a.service:
            total_revenue += float(a.service.price)

    # By service
    by_service = apt_q.values('service__name').annotate(cnt=Count('id')).order_by('-cnt')[:10]

    # By provider
    by_provider = apt_q.values('provider__first_name', 'provider__last_name').annotate(
        total=Count('id'), comp=Count('id', filter=Q(status='completed'))
    ).order_by('-total')[:10]

    unique_patients = apt_q.values('patient').distinct().count()

    # Monthly trend
    monthly_labels = []
    monthly_data = []
    for i in range(6):
        ms = (end_date.replace(day=1) - timedelta(days=i * 30)).replace(day=1)
        me = (ms + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        cnt = Appointment.objects.filter(appointment_date__range=[ms, me]).count()
        monthly_labels.insert(0, ms.strftime('%b %Y'))
        monthly_data.insert(0, cnt)

    return {
        'total': total, 'completed': completed, 'scheduled': scheduled,
        'cancelled': cancelled, 'no_show': no_show,
        'completion_rate': round(completion_rate, 1),
        'cancellation_rate': round(cancellation_rate, 1),
        'total_revenue': total_revenue, 'unique_patients': unique_patients,
        'by_service': list(by_service), 'by_provider': list(by_provider),
        'monthly_labels': monthly_labels, 'monthly_data': monthly_data,
    }


@login_required
def financial_reports(request):
    from clinic_settings.models import EnabledModule

    start_date, end_date, prev_start, prev_end, period, date_from, date_to = _parse_financial_dates(request)
    enabled = EnabledModule.get_enabled_modules()

    # Build list of active financial tabs
    app_tabs = []
    if 'billing' in enabled:
        app_tabs.append({'key': 'billing', 'label': 'Billing & Invoicing', 'icon': 'bi-credit-card'})
    if 'pharmacy' in enabled:
        app_tabs.append({'key': 'pharmacy', 'label': 'Pharmacy', 'icon': 'bi-capsule'})
    if 'laboratory' in enabled:
        app_tabs.append({'key': 'laboratory', 'label': 'Laboratory', 'icon': 'bi-droplet-half'})
    if 'budget' in enabled:
        app_tabs.append({'key': 'budget', 'label': 'Budget & Expenses', 'icon': 'bi-wallet2'})
    if 'appointments' in enabled:
        app_tabs.append({'key': 'appointments', 'label': 'Appointments', 'icon': 'bi-calendar-check'})

    # Gather stats only for enabled apps
    billing = _get_billing_stats(start_date, end_date, prev_start, prev_end) if 'billing' in enabled else {}
    pharmacy = _get_pharmacy_stats(start_date, end_date) if 'pharmacy' in enabled else {}
    lab = _get_laboratory_stats(start_date, end_date) if 'laboratory' in enabled else {}
    budget = _get_budget_stats(start_date, end_date) if 'budget' in enabled else {}
    appointments = _get_appointments_stats(start_date, end_date) if 'appointments' in enabled else {}

    # General overview totals
    total_revenue = float(billing.get('services_revenue', 0)) + float(pharmacy.get('sales_revenue', 0))
    total_expenses = float(budget.get('total_expenses', 0))
    net_balance = total_revenue - total_expenses

    # ── Choice data for custom report modal ──
    invoice_statuses = Invoice.STATUS_CHOICES if hasattr(Invoice, 'STATUS_CHOICES') else []
    payment_methods = Payment.PAYMENT_METHODS if hasattr(Payment, 'PAYMENT_METHODS') else []
    providers = User.objects.filter(role__in=['doctor', 'physiotherapist', 'nutritionist']).order_by('first_name')
    services = Service.objects.filter(is_active=True).order_by('name')
    service_categories = Service.SERVICE_CATEGORIES if hasattr(Service, 'SERVICE_CATEGORIES') else []

    context = {
        'period': period, 'date_from': date_from, 'date_to': date_to,
        'start_date': start_date, 'end_date': end_date,
        'enabled_modules': enabled, 'app_tabs': app_tabs,
        'total_revenue': total_revenue, 'total_expenses': total_expenses,
        'net_balance': net_balance,
        'billing': billing, 'pharmacy': pharmacy, 'lab': lab,
        'budget': budget, 'appointments': appointments,
        # JSON for charts
        'billing_json': json.dumps(billing, default=str) if billing else '{}',
        'pharmacy_json': json.dumps(pharmacy, default=str) if pharmacy else '{}',
        'lab_json': json.dumps(lab, default=str) if lab else '{}',
        'budget_json': json.dumps(budget, default=str) if budget else '{}',
        'appointments_json': json.dumps(appointments, default=str) if appointments else '{}',
        # Custom report modal choices
        'invoice_statuses': invoice_statuses,
        'payment_methods': payment_methods,
        'providers': providers,
        'services': services,
        'service_categories': service_categories,
    }
    return render(request, 'reports/financial_reports.html', context)

@login_required
def appointment_report(request):
    # Date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date:
        start_date = date.today().replace(day=1)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    if not end_date:
        end_date = date.today()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Appointment statistics
    appointments = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date]
    )
    
    total_appointments = appointments.count()
    
    # Status distribution
    status_stats = appointments.values('status').annotate(count=Count('id'))
    
    # Calculate individual status counts
    completed_count = appointments.filter(status='completed').count()
    scheduled_count = appointments.filter(status='scheduled').count()
    cancelled_count = appointments.filter(status='cancelled').count()
    no_show_count = appointments.filter(status='no_show').count()
    
    # Calculate rates
    completion_rate = (completed_count / total_appointments * 100) if total_appointments > 0 else 0
    cancellation_rate = (cancelled_count / total_appointments * 100) if total_appointments > 0 else 0
    no_show_rate = (no_show_count / total_appointments * 100) if total_appointments > 0 else 0
    
    # Service distribution with detailed stats
    service_stats = appointments.values('service__name').annotate(
        count=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        cancelled=Count('id', filter=Q(status='cancelled'))
    ).order_by('-count')
    
    # Provider workload with performance metrics
    provider_stats = appointments.values(
        'provider__first_name', 
        'provider__last_name'
    ).annotate(
        total_count=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        scheduled=Count('id', filter=Q(status='scheduled')),
        cancelled=Count('id', filter=Q(status='cancelled')),
        no_show=Count('id', filter=Q(status='no_show'))
    ).order_by('-total_count')
    
    # Calculate unique patients and providers
    unique_patients = appointments.values('patient').distinct().count()
    unique_providers = appointments.values('provider').distinct().count()
    
    # Average appointments per day
    date_diff = (end_date - start_date).days + 1
    avg_appointments_per_day = total_appointments / date_diff if date_diff > 0 else 0
    
    # Peak hours analysis (if time data available)
    hourly_distribution = []
    for hour in range(8, 18):  # 8 AM to 6 PM
        count = appointments.filter(
            appointment_time__hour=hour
        ).count() if hasattr(appointments.first(), 'appointment_time') else 0
        hourly_distribution.append({
            'hour': f"{hour}:00",
            'count': count
        })
    
    # Daily appointment trend (last 7 days within range)
    daily_labels = []
    daily_data = []
    daily_completed = []
    daily_cancelled = []
    
    current_date = max(start_date, end_date - timedelta(days=6))
    while current_date <= end_date:
        day_appointments = appointments.filter(appointment_date=current_date)
        daily_labels.append(current_date.strftime('%m/%d'))
        daily_data.append(day_appointments.count())
        daily_completed.append(day_appointments.filter(status='completed').count())
        daily_cancelled.append(day_appointments.filter(status='cancelled').count())
        current_date += timedelta(days=1)
    
    # Monthly trend (last 6 months)
    monthly_labels = []
    monthly_data = []
    for i in range(6):
        month_start = (end_date.replace(day=1) - timedelta(days=i*30)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        count = Appointment.objects.filter(
            appointment_date__range=[month_start, min(month_end, end_date)]
        ).count()
        monthly_labels.insert(0, month_start.strftime('%b %Y'))
        monthly_data.insert(0, count)
    
    # Department distribution (if applicable)
    dept_stats = appointments.values('service__name').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    # Revenue from completed appointments (if linked to invoices)
    completed_appointments = appointments.filter(status='completed')
    total_revenue = 0
    for apt in completed_appointments:
        if hasattr(apt, 'service') and apt.service:
            total_revenue += apt.service.price
    
    # ── Choice data for custom report modal ──
    all_services = Service.objects.filter(is_active=True).order_by('name')
    all_service_categories = Service.SERVICE_CATEGORIES if hasattr(Service, 'SERVICE_CATEGORIES') else []
    all_providers = User.objects.filter(role__in=['doctor', 'physiotherapist', 'nutritionist']).order_by('first_name')
    appt_statuses = Appointment.STATUS_CHOICES if hasattr(Appointment, 'STATUS_CHOICES') else []

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_appointments': total_appointments,
        'completed_count': completed_count,
        'scheduled_count': scheduled_count,
        'cancelled_count': cancelled_count,
        'no_show_count': no_show_count,
        'completion_rate': round(completion_rate, 1),
        'cancellation_rate': round(cancellation_rate, 1),
        'no_show_rate': round(no_show_rate, 1),
        'unique_patients': unique_patients,
        'unique_providers': unique_providers,
        'avg_appointments_per_day': round(avg_appointments_per_day, 1),
        'total_revenue': total_revenue,
        'status_stats': status_stats,
        'service_stats': service_stats,
        'provider_stats': provider_stats,
        'hourly_distribution': hourly_distribution,
        'daily_labels': json.dumps(daily_labels),
        'daily_data': json.dumps(daily_data),
        'daily_completed': json.dumps(daily_completed),
        'daily_cancelled': json.dumps(daily_cancelled),
        'monthly_labels': json.dumps(monthly_labels),
        'monthly_data': json.dumps(monthly_data),
        'dept_stats': dept_stats,
        # Custom report modal choices
        'all_services': all_services,
        'all_service_categories': all_service_categories,
        'all_providers': all_providers,
        'appt_statuses': appt_statuses,
    }
    return render(request, 'reports/appointment_report.html', context)

@login_required
def download_financial_report(request):
    """Generate a multi-sheet Excel financial report for all enabled apps."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from clinic_settings.models import EnabledModule

    start_date, end_date, prev_start, prev_end, period, date_from, date_to = _parse_financial_dates(request)
    enabled = EnabledModule.get_enabled_modules()

    wb = openpyxl.Workbook()
    # Styles
    title_font = Font(name='Arial', size=16, bold=True, color='1B5E96')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1B5E96', end_color='1B5E96', fill_type='solid')
    sub_font = Font(name='Arial', size=12, bold=True, color='2E8B57')
    data_font = Font(name='Arial', size=10)
    money_fmt = '#,##0'
    pct_fmt = '0.0"%"'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def write_header(ws, row, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        return row + 1

    def write_row(ws, row, values, fmt=None):
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = data_font
            cell.border = thin_border
            if fmt and c in fmt:
                cell.number_format = fmt[c]
        return row + 1

    def auto_width(ws, cols):
        for c in range(1, cols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 18

    # ── OVERVIEW SHEET ──
    ws = wb.active
    ws.title = 'Overview'
    ws.cell(row=1, column=1, value='Financial Reports & Analytics').font = title_font
    ws.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}').font = Font(name='Arial', size=11, italic=True)
    r = 4

    # Collect overview data
    billing = _get_billing_stats(start_date, end_date, prev_start, prev_end) if 'billing' in enabled else {}
    pharmacy = _get_pharmacy_stats(start_date, end_date) if 'pharmacy' in enabled else {}
    lab = _get_laboratory_stats(start_date, end_date) if 'laboratory' in enabled else {}
    budget_data = _get_budget_stats(start_date, end_date) if 'budget' in enabled else {}
    appt = _get_appointments_stats(start_date, end_date) if 'appointments' in enabled else {}

    total_revenue = float(billing.get('services_revenue', 0)) + float(pharmacy.get('sales_revenue', 0))
    total_expenses = float(budget_data.get('total_expenses', 0))

    ws.cell(row=r, column=1, value='Summary').font = sub_font
    r += 1
    overview_items = [
        ('Total Revenue', total_revenue),
        ('  Services Revenue', float(billing.get('services_revenue', 0))),
        ('  Pharmacy Sales', float(pharmacy.get('sales_revenue', 0))),
        ('Total Expenses', total_expenses),
        ('Net Balance', total_revenue - total_expenses),
    ]
    r = write_header(ws, r, ['Metric', 'Amount (UGX)'])
    for label, val in overview_items:
        r = write_row(ws, r, [label, val], fmt={2: money_fmt})
    r += 1

    # Links to other sheets
    ws.cell(row=r, column=1, value='Sections').font = sub_font
    r += 1
    sheet_names = []
    if 'billing' in enabled:
        sheet_names.append('Billing')
    if 'pharmacy' in enabled:
        sheet_names.append('Pharmacy')
    if 'laboratory' in enabled:
        sheet_names.append('Laboratory')
    if 'budget' in enabled:
        sheet_names.append('Budget & Expenses')
    if 'appointments' in enabled:
        sheet_names.append('Appointments')
    for sn in sheet_names:
        cell = ws.cell(row=r, column=1, value=sn)
        cell.font = Font(name='Arial', size=11, color='0563C1', underline='single')
        safe = sn.replace(' ', '_').replace('&', 'and')[:31]
        cell.hyperlink = f"#'{sn}'!A1"
        r += 1
    auto_width(ws, 2)

    # ── BILLING SHEET ──
    if 'billing' in enabled and billing:
        ws_b = wb.create_sheet('Billing')
        ws_b.cell(row=1, column=1, value='Billing & Invoicing Report').font = title_font
        ws_b.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}').font = Font(name='Arial', size=10, italic=True)
        r = 4
        ws_b.cell(row=r, column=1, value='Key Metrics').font = sub_font; r += 1
        r = write_header(ws_b, r, ['Metric', 'Value'])
        for k, v in [
            ('Revenue', billing['services_revenue']), ('Invoices', billing['total_invoices']),
            ('Avg Invoice', billing['avg_invoice_value']), ('Outstanding', billing['outstanding_amount']),
            ('Collection Rate %', billing['collection_rate']),
            ('Prev Period Revenue', billing['prev_revenue']),
            ('Revenue Change', billing['revenue_change']),
        ]:
            r = write_row(ws_b, r, [k, v])
        r += 1
        # Payment methods
        ws_b.cell(row=r, column=1, value='Payment Methods').font = sub_font; r += 1
        r = write_header(ws_b, r, ['Method', 'Amount'])
        for lbl, val in zip(billing['pm_labels'], billing['pm_data']):
            r = write_row(ws_b, r, [lbl, val], fmt={2: money_fmt})
        r += 1
        # Top services
        ws_b.cell(row=r, column=1, value='Top Services').font = sub_font; r += 1
        r = write_header(ws_b, r, ['Service', 'Revenue', 'Count'])
        for s in billing['top_services']:
            r = write_row(ws_b, r, [s['name'], s['total_revenue'], s['count']], fmt={2: money_fmt})
        r += 1
        # Outstanding aging
        ws_b.cell(row=r, column=1, value='Outstanding Aging').font = sub_font; r += 1
        r = write_header(ws_b, r, ['Age Bracket', 'Amount'])
        for lbl, val in [('0-30 days', billing['o30']), ('31-60 days', billing['o60']),
                         ('61-90 days', billing['o90']), ('90+ days', billing['o90p'])]:
            r = write_row(ws_b, r, [lbl, val], fmt={2: money_fmt})
        r += 1
        # Insurance claims
        ws_b.cell(row=r, column=1, value='Insurance Claims').font = sub_font; r += 1
        r = write_header(ws_b, r, ['Metric', 'Value'])
        for k, v in [('Submitted', billing['claims_submitted']), ('Approved', billing['claims_approved']),
                     ('Pending', billing['claims_pending']), ('Denied', billing['claims_denied']),
                     ('Total Claimed', billing['claims_total']), ('Reimbursed', billing['claims_reimbursed'])]:
            r = write_row(ws_b, r, [k, v])
        auto_width(ws_b, 3)

    # ── PHARMACY SHEET ──
    if 'pharmacy' in enabled and pharmacy:
        ws_p = wb.create_sheet('Pharmacy')
        ws_p.cell(row=1, column=1, value='Pharmacy Report').font = title_font
        ws_p.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}').font = Font(name='Arial', size=10, italic=True)
        r = 4
        ws_p.cell(row=r, column=1, value='Sales Summary').font = sub_font; r += 1
        r = write_header(ws_p, r, ['Metric', 'Value'])
        for k, v in [
            ('Sales Revenue', pharmacy['sales_revenue']), ('Cost of Goods', pharmacy['sales_cogs']),
            ('Gross Profit', pharmacy['sales_profit']), ('Profit Margin %', pharmacy['sales_margin']),
            ('Items Sold', pharmacy['total_items_sold']), ('Transactions', pharmacy['total_transactions']),
        ]:
            r = write_row(ws_p, r, [k, v])
        r += 1
        ws_p.cell(row=r, column=1, value='Stock Overview').font = sub_font; r += 1
        r = write_header(ws_p, r, ['Metric', 'Value'])
        for k, v in [
            ('Total Medications', pharmacy['total_medications']),
            ('Low Stock Items', pharmacy['low_stock_count']),
            ('Expiring (90 days)', pharmacy['expiring_count']),
            ('Expired', pharmacy['expired_count']),
            ('Stock Value (Cost)', pharmacy['stock_cost_val']),
            ('Stock Value (Sell)', pharmacy['stock_sell_val']),
            ('Purchase Orders', pharmacy['po_count']),
            ('PO Total', pharmacy['po_total']),
        ]:
            r = write_row(ws_p, r, [k, v])
        r += 1
        # Top medications
        ws_p.cell(row=r, column=1, value='Top Selling Medications').font = sub_font; r += 1
        r = write_header(ws_p, r, ['Medication', 'Qty Sold', 'Revenue'])
        for m in pharmacy['top_meds']:
            r = write_row(ws_p, r, [m.get('batch__medication__name', ''), m.get('qty', 0), float(m.get('rev', 0))], fmt={3: money_fmt})
        r += 1
        # Low stock items
        if pharmacy['low_stock_items']:
            ws_p.cell(row=r, column=1, value='Low Stock Alert').font = sub_font; r += 1
            r = write_header(ws_p, r, ['Medication', 'Current Stock', 'Reorder Level'])
            for item in pharmacy['low_stock_items']:
                r = write_row(ws_p, r, [item['name'], item['stock'], item['reorder']])
        auto_width(ws_p, 3)

    # ── LABORATORY SHEET ──
    if 'laboratory' in enabled and lab:
        ws_l = wb.create_sheet('Laboratory')
        ws_l.cell(row=1, column=1, value='Laboratory Report').font = title_font
        ws_l.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}').font = Font(name='Arial', size=10, italic=True)
        r = 4
        ws_l.cell(row=r, column=1, value='Test Statistics').font = sub_font; r += 1
        r = write_header(ws_l, r, ['Metric', 'Value'])
        for k, v in [
            ('Total Requests', lab['total_requests']), ('Completed', lab['completed']),
            ('Pending', lab['pending']), ('Cancelled', lab['cancelled']),
            ('Completion Rate %', lab['completion_rate']),
            ('Total Results', lab['total_results']),
            ('Abnormal Results', lab['abnormal_results']),
            ('Verified Results', lab['verified_results']),
            ('Unique Patients', lab['unique_patients']),
            ('Avg TAT (hours)', lab['avg_tat_hours']),
        ]:
            r = write_row(ws_l, r, [k, v])
        r += 1
        # Top tests
        ws_l.cell(row=r, column=1, value='Top Requested Tests').font = sub_font; r += 1
        r = write_header(ws_l, r, ['Test', 'Category', 'Count'])
        for t in lab['top_tests']:
            r = write_row(ws_l, r, [t.get('test__name', ''), t.get('test__category', ''), t.get('cnt', 0)])
        r += 1
        # By priority
        ws_l.cell(row=r, column=1, value='By Priority').font = sub_font; r += 1
        r = write_header(ws_l, r, ['Priority', 'Count'])
        for k, v in lab['by_priority'].items():
            r = write_row(ws_l, r, [k, v])
        auto_width(ws_l, 3)

    # ── BUDGET SHEET ──
    if 'budget' in enabled and budget_data:
        ws_e = wb.create_sheet('Budget & Expenses')
        ws_e.cell(row=1, column=1, value='Budget & Expenses Report').font = title_font
        ws_e.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}').font = Font(name='Arial', size=10, italic=True)
        r = 4
        ws_e.cell(row=r, column=1, value='Expense Summary').font = sub_font; r += 1
        r = write_header(ws_e, r, ['Metric', 'Value'])
        for k, v in [
            ('Total Expenses', budget_data['total_expenses']),
            ('Expense Count', budget_data['total_expense_count']),
            ('Pending Approval', budget_data['pending_expenses']),
            ('Rejected', budget_data['rejected_expenses']),
        ]:
            r = write_row(ws_e, r, [k, v])
        r += 1
        # By category
        ws_e.cell(row=r, column=1, value='Expenses by Category').font = sub_font; r += 1
        r = write_header(ws_e, r, ['Category', 'Amount', 'Count'])
        for cat in budget_data['by_category']:
            r = write_row(ws_e, r, [cat['category__name'], float(cat['total']), cat['cnt']], fmt={2: money_fmt})
        r += 1
        # Top vendors
        if budget_data['top_vendors']:
            ws_e.cell(row=r, column=1, value='Top Vendors').font = sub_font; r += 1
            r = write_header(ws_e, r, ['Vendor', 'Amount', 'Count'])
            for v in budget_data['top_vendors']:
                r = write_row(ws_e, r, [v['vendor_name'], float(v['total']), v['cnt']], fmt={2: money_fmt})
            r += 1
        # Active budgets
        if budget_data['active_budgets']:
            ws_e.cell(row=r, column=1, value='Active Budgets').font = sub_font; r += 1
            r = write_header(ws_e, r, ['Budget', 'Total', 'Spent', 'Remaining', 'Utilization %'])
            for b in budget_data['active_budgets']:
                r = write_row(ws_e, r, [b['name'], b['total'], b['spent'], b['remaining'], b['utilization']], fmt={2: money_fmt, 3: money_fmt, 4: money_fmt})
        auto_width(ws_e, 5)

    # ── APPOINTMENTS SHEET ──
    if 'appointments' in enabled and appt:
        ws_a = wb.create_sheet('Appointments')
        ws_a.cell(row=1, column=1, value='Appointments Report').font = title_font
        ws_a.cell(row=2, column=1, value=f'Period: {start_date} to {end_date}').font = Font(name='Arial', size=10, italic=True)
        r = 4
        ws_a.cell(row=r, column=1, value='Appointment Statistics').font = sub_font; r += 1
        r = write_header(ws_a, r, ['Metric', 'Value'])
        for k, v in [
            ('Total Appointments', appt['total']), ('Completed', appt['completed']),
            ('Scheduled', appt['scheduled']), ('Cancelled', appt['cancelled']),
            ('No Show', appt['no_show']),
            ('Completion Rate %', appt['completion_rate']),
            ('Cancellation Rate %', appt['cancellation_rate']),
            ('Revenue (est.)', appt['total_revenue']),
            ('Unique Patients', appt['unique_patients']),
        ]:
            r = write_row(ws_a, r, [k, v])
        r += 1
        # By service
        ws_a.cell(row=r, column=1, value='By Service').font = sub_font; r += 1
        r = write_header(ws_a, r, ['Service', 'Count'])
        for s in appt['by_service']:
            r = write_row(ws_a, r, [s.get('service__name', 'N/A'), s.get('cnt', 0)])
        r += 1
        # By provider
        ws_a.cell(row=r, column=1, value='By Provider').font = sub_font; r += 1
        r = write_header(ws_a, r, ['Provider', 'Total', 'Completed'])
        for p in appt['by_provider']:
            name = f"{p.get('provider__first_name', '')} {p.get('provider__last_name', '')}".strip() or 'N/A'
            r = write_row(ws_a, r, [name, p.get('total', 0), p.get('comp', 0)])
        auto_width(ws_a, 3)

    # Write response
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Financial_Report_{start_date}_{end_date}.xlsx"'
    return response


@login_required
@require_http_methods(["POST"])
def export_report(request):
    """Export report to PDF, Excel, or CSV"""
    try:
        report_type = request.POST.get('report_type')
        export_format = request.POST.get('export_format')
        report_name = request.POST.get('report_name', f'{report_type.title()} Report')
        
        # Get report parameters from POST data
        parameters = {}
        for key, value in request.POST.items():
            if key not in ['report_type', 'export_format', 'report_name', 'csrfmiddlewaretoken']:
                parameters[key] = value
        
        # Generate report data based on type
        if report_type == 'dashboard':
            content_data = generate_dashboard_export_data(parameters)
        elif report_type == 'patient':
            content_data = generate_patient_export_data(parameters)
        elif report_type == 'financial':
            content_data = generate_financial_export_data(parameters)
        elif report_type == 'appointment':
            content_data = generate_appointment_export_data(parameters)
        else:
            return JsonResponse({'error': 'Invalid report type'}, status=400)
        
        # Create export
        response, export_record = create_report_export(
            user=request.user,
            report_type=report_type,
            report_name=report_name,
            export_format=export_format,
            content_data=content_data,
            parameters=parameters
        )
        
        # Log export activity
        audit_mixin = ReportAuditMixin()
        audit_mixin.log_report_activity(
            request=request,
            report_type=report_type,
            report_name=report_name,
            action=f'exported_{export_format}',
            record_count=len(content_data.get('tables', [{}])[0].get('data', [])),
            file_size=export_record.file_size,
            parameters=parameters
        )
        
        return response
        
    except Exception as e:
        # Log error
        audit_mixin = ReportAuditMixin()
        audit_mixin.log_report_activity(
            request=request,
            report_type=request.POST.get('report_type', 'unknown'),
            report_name=request.POST.get('report_name', 'Unknown Report'),
            action=f'exported_{request.POST.get("export_format", "unknown")}',
            success=False,
            error_message=str(e)
        )
        return JsonResponse({'error': str(e)}, status=500)

def generate_dashboard_export_data(parameters):
    """Generate export data for dashboard report"""
    # Parse date parameters
    start_date = datetime.strptime(parameters.get('start_date', date.today().replace(day=1).isoformat()), '%Y-%m-%d').date()
    end_date = datetime.strptime(parameters.get('end_date', date.today().isoformat()), '%Y-%m-%d').date()
    
    # Get dashboard data
    total_patients = Patient.objects.filter(is_active=True).count()
    new_patients = Patient.objects.filter(registration_date__range=[start_date, end_date]).count()
    total_appointments = Appointment.objects.filter(appointment_date__range=[start_date, end_date]).count()
    completed_appointments = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date], status='completed'
    ).count()
    total_revenue = Payment.objects.filter(
        payment_date__date__range=[start_date, end_date], status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Popular services data
    popular_services = Service.objects.annotate(
        appointment_count=Count('appointment')
    ).order_by('-appointment_count')[:10]
    
    services_data = []
    for service in popular_services:
        services_data.append([
            service.name,
            service.category,
            f"UGX {service.price:,.0f}",
            service.appointment_count
        ])
    
    return {
        'summary_stats': {
            'Total Patients': total_patients,
            'New Patients': new_patients,
            'Total Appointments': total_appointments,
            'Completed Appointments': completed_appointments,
            'Total Revenue': f"UGX {total_revenue:,.0f}",
            'Report Period': f"{start_date} to {end_date}"
        },
        'tables': [
            {
                'title': 'Popular Services',
                'headers': ['Service Name', 'Category', 'Price', 'Appointments'],
                'data': services_data
            }
        ]
    }

def generate_patient_export_data(parameters):
    """Generate export data for patient report"""
    # Get filter parameters
    date_range = parameters.get('date_range', 'last_30_days')
    gender = parameters.get('gender', '')
    
    # Calculate date range
    end_date = date.today()
    if date_range == 'last_30_days':
        start_date = end_date - timedelta(days=30)
    elif date_range == 'last_3_months':
        start_date = end_date - timedelta(days=90)
    elif date_range == 'last_6_months':
        start_date = end_date - timedelta(days=180)
    elif date_range == 'last_year':
        start_date = end_date - timedelta(days=365)
    else:
        start_date = end_date - timedelta(days=30)
    
    # Base patient queryset
    patients = Patient.objects.filter(is_active=True)
    if gender:
        patients = patients.filter(gender=gender)
    
    # Patient data for export
    patient_data = []
    for patient in patients[:100]:  # Limit for export
        try:
            age = patient.get_age() if hasattr(patient, 'get_age') else None
            age_display = str(int(age)) if age and isinstance(age, (int, float)) else 'N/A'
        except (ValueError, TypeError, AttributeError):
            age_display = 'N/A'
            
        patient_data.append([
            f"{patient.first_name} {patient.last_name}",
            patient.email,
            patient.phone,
            patient.gender,
            age_display,
            patient.registration_date.strftime('%Y-%m-%d'),
            patient.insurance_provider or 'None'
        ])
    
    # Calculate average age safely
    ages = []
    for p in patients:
        try:
            age = p.get_age() if hasattr(p, 'get_age') else None
            if age and isinstance(age, (int, float)):
                ages.append(int(age))
        except (ValueError, TypeError, AttributeError):
            continue
    
    avg_age = round(sum(ages) / len(ages), 1) if ages else 0
    
    return {
        'summary_stats': {
            'Total Patients': patients.count(),
            'New Patients (Period)': patients.filter(registration_date__range=[start_date, end_date]).count(),
            'Average Age': avg_age,
            'Report Period': f"{start_date} to {end_date}"
        },
        'tables': [
            {
                'title': 'Patient List',
                'headers': ['Name', 'Email', 'Phone', 'Gender', 'Age', 'Registration Date', 'Insurance'],
                'data': patient_data
            }
        ]
    }

def generate_financial_export_data(parameters):
    """Generate export data for financial report"""
    period = parameters.get('period', 'this_month')
    
    # Calculate date range based on period
    end_date = date.today()
    if period == 'this_month':
        start_date = end_date.replace(day=1)
    elif period == 'last_month':
        start_date = (end_date.replace(day=1) - timedelta(days=1)).replace(day=1)
        end_date = end_date.replace(day=1) - timedelta(days=1)
    elif period == 'this_quarter':
        quarter = (end_date.month - 1) // 3 + 1
        start_date = date(end_date.year, (quarter - 1) * 3 + 1, 1)
    elif period == 'this_year':
        start_date = date(end_date.year, 1, 1)
    else:
        start_date = end_date.replace(day=1)
    
    # Financial metrics
    total_revenue = Payment.objects.filter(
        payment_date__date__range=[start_date, end_date], status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    outstanding_amount = Invoice.objects.filter(
        status__in=['sent', 'overdue']
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Recent invoices data
    recent_invoices = Invoice.objects.filter(
        created_at__date__range=[start_date, end_date]
    ).order_by('-created_at')[:50]
    
    invoice_data = []
    for invoice in recent_invoices:
        invoice_data.append([
            invoice.invoice_number,
            f"{invoice.patient.first_name} {invoice.patient.last_name}" if invoice.patient else 'N/A',
            f"UGX {invoice.total_amount:,.0f}",
            invoice.status.title(),
            invoice.created_at.strftime('%Y-%m-%d'),
            invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else 'N/A'
        ])
    
    return {
        'summary_stats': {
            'Total Revenue': f"UGX {total_revenue:,.0f}",
            'Outstanding Amount': f"UGX {outstanding_amount:,.0f}",
            'Total Invoices': recent_invoices.count(),
            'Report Period': f"{start_date} to {end_date}"
        },
        'tables': [
            {
                'title': 'Recent Invoices',
                'headers': ['Invoice #', 'Patient', 'Amount', 'Status', 'Created', 'Due Date'],
                'data': invoice_data
            }
        ]
    }

def generate_appointment_export_data(parameters):
    """Generate export data for appointment report"""
    start_date = datetime.strptime(parameters.get('start_date', date.today().replace(day=1).isoformat()), '%Y-%m-%d').date()
    end_date = datetime.strptime(parameters.get('end_date', date.today().isoformat()), '%Y-%m-%d').date()
    
    # Appointment data
    appointments = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date]
    ).order_by('-appointment_date')[:100]
    
    appointment_data = []
    for appointment in appointments:
        appointment_data.append([
            appointment.appointment_date.strftime('%Y-%m-%d %H:%M'),
            f"{appointment.patient.first_name} {appointment.patient.last_name}" if appointment.patient else 'N/A',
            f"{appointment.provider.first_name} {appointment.provider.last_name}" if appointment.provider else 'N/A',
            appointment.service.name if appointment.service else 'N/A',
            appointment.status.title(),
            appointment.notes[:100] + '...' if appointment.notes and len(appointment.notes) > 100 else appointment.notes or ''
        ])
    
    return {
        'summary_stats': {
            'Total Appointments': appointments.count(),
            'Completed': appointments.filter(status='completed').count(),
            'Cancelled': appointments.filter(status='cancelled').count(),
            'No Show': appointments.filter(status='no_show').count(),
            'Report Period': f"{start_date} to {end_date}"
        },
        'tables': [
            {
                'title': 'Appointments',
                'headers': ['Date/Time', 'Patient', 'Provider', 'Service', 'Status', 'Notes'],
                'data': appointment_data
            }
        ]
    }

@login_required
def audit_log(request):
    """View audit log for reports"""
    # Get filter parameters
    user_filter = request.GET.get('user')
    report_type_filter = request.GET.get('report_type')
    action_filter = request.GET.get('action')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Base queryset
    audit_logs = ReportAuditLog.objects.all()
    
    # Apply filters
    if user_filter:
        audit_logs = audit_logs.filter(user__username__icontains=user_filter)
    if report_type_filter:
        audit_logs = audit_logs.filter(report_type=report_type_filter)
    if action_filter:
        audit_logs = audit_logs.filter(action=action_filter)
    if date_from:
        audit_logs = audit_logs.filter(timestamp__date__gte=date_from)
    if date_to:
        audit_logs = audit_logs.filter(timestamp__date__lte=date_to)
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(audit_logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get filter options
    users = User.objects.filter(report_activities__isnull=False).distinct()
    report_types = ReportAuditLog.objects.values_list('report_type', flat=True).distinct()
    actions = ReportAuditLog.objects.values_list('action', flat=True).distinct()
    
    context = {
        'page_obj': page_obj,
        'users': users,
        'report_types': report_types,
        'actions': actions,
        'filters': {
            'user': user_filter,
            'report_type': report_type_filter,
            'action': action_filter,
            'date_from': date_from,
            'date_to': date_to,
        }
    }
    return render(request, 'reports/audit_log.html', context)

@login_required
def report_performance(request):
    """View report performance metrics with comprehensive analytics"""
    metrics = get_report_performance_metrics()
    
    # Get detailed performance data
    from django.db.models import Avg, Max, Min, Sum
    
    # Date range for analysis
    end_date = date.today()
    start_date = end_date - timedelta(days=30)
    
    # Total reports generated
    total_reports = ReportAuditLog.objects.filter(
        timestamp__date__range=[start_date, end_date]
    ).count()
    
    # Success rate
    successful_reports = ReportAuditLog.objects.filter(
        timestamp__date__range=[start_date, end_date],
        success=True
    ).count()
    success_rate = (successful_reports / total_reports * 100) if total_reports > 0 else 0
    
    # Execution time statistics by report type
    execution_stats = ReportAuditLog.objects.filter(
        timestamp__date__range=[start_date, end_date],
        success=True
    ).values('report_type').annotate(
        avg_time=Avg('execution_time'),
        max_time=Max('execution_time'),
        min_time=Min('execution_time'),
        total_count=Count('id'),
        total_records=Sum('record_count')
    ).order_by('-avg_time')
    
    # Most used reports
    popular_reports = ReportAuditLog.objects.filter(
        timestamp__date__range=[start_date, end_date]
    ).values('report_type', 'report_name').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # User activity
    user_activity = ReportAuditLog.objects.filter(
        timestamp__date__range=[start_date, end_date]
    ).values('user__first_name', 'user__last_name').annotate(
        report_count=Count('id'),
        avg_execution_time=Avg('execution_time')
    ).order_by('-report_count')[:10]
    
    # Action distribution
    action_stats = ReportAuditLog.objects.filter(
        timestamp__date__range=[start_date, end_date]
    ).values('action').annotate(
        count=Count('id')
    ).order_by('-count')
    
    action_labels = []
    action_data = []
    for stat in action_stats:
        action_labels.append(stat['action'].title())
        action_data.append(stat['count'])
    
    # Daily report generation trend (last 30 days)
    daily_stats = []
    daily_success = []
    daily_failures = []
    
    for i in range(30):
        day = end_date - timedelta(days=i)
        total = ReportAuditLog.objects.filter(
            timestamp__date=day
        ).count()
        success = ReportAuditLog.objects.filter(
            timestamp__date=day,
            success=True
        ).count()
        failed = total - success
        
        daily_stats.insert(0, {
            'date': day.strftime('%m/%d'),
            'count': total,
            'success': success,
            'failed': failed
        })
        daily_success.insert(0, success)
        daily_failures.insert(0, failed)
    
    # Hourly distribution (peak usage times)
    hourly_stats = []
    for hour in range(24):
        count = ReportAuditLog.objects.filter(
            timestamp__date__range=[start_date, end_date],
            timestamp__hour=hour
        ).count()
        hourly_stats.append({
            'hour': f"{hour}:00",
            'count': count
        })
    
    # Performance recommendations
    recommendations = []
    
    # Check for slow reports
    slow_reports = execution_stats.filter(avg_time__gt=5)
    if slow_reports.exists():
        recommendations.append({
            'type': 'warning',
            'message': f"{slow_reports.count()} report type(s) have average execution time > 5 seconds. Consider optimization."
        })
    
    # Check failure rate
    if success_rate < 95:
        recommendations.append({
            'type': 'danger',
            'message': f"Success rate is {success_rate:.1f}%. Investigate failed reports."
        })
    
    # Check for high usage
    if total_reports > 1000:
        recommendations.append({
            'type': 'info',
            'message': f"{total_reports} reports generated in last 30 days. Consider implementing caching."
        })
    
    # Average response time
    avg_response_time = ReportAuditLog.objects.filter(
        timestamp__date__range=[start_date, end_date],
        success=True
    ).aggregate(Avg('execution_time'))['execution_time__avg'] or 0
    
    # Fastest and slowest reports
    fastest_report = execution_stats.order_by('avg_time').first()
    slowest_report = execution_stats.order_by('-avg_time').first()
    
    context = {
        'metrics': metrics,
        'total_reports': total_reports,
        'successful_reports': successful_reports,
        'success_rate': round(success_rate, 1),
        'avg_response_time': round(avg_response_time, 3),
        'fastest_report': fastest_report,
        'slowest_report': slowest_report,
        'execution_stats': execution_stats,
        'popular_reports': popular_reports,
        'user_activity': user_activity,
        'action_stats': action_stats,
        'action_labels': json.dumps(action_labels),
        'action_data': json.dumps(action_data),
        'daily_stats': daily_stats,
        'daily_labels': json.dumps([stat['date'] for stat in daily_stats]),
        'daily_data': json.dumps([stat['count'] for stat in daily_stats]),
        'daily_success': json.dumps(daily_success),
        'daily_failures': json.dumps(daily_failures),
        'hourly_stats': hourly_stats,
        'recommendations': recommendations,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'reports/performance.html', context)

@login_required
def physiotherapy_reports(request):
    """Physiotherapy department specific reports"""
    from patients.models import Assessment
    
    # Get filter parameters
    start_date = request.GET.get('start_date', date.today().replace(day=1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Physiotherapy assessments
    physio_assessments = Assessment.objects.filter(
        department='physiotherapy',
        assessment_date__range=[start_date, end_date]
    ).select_related('patient', 'assessed_by')
    
    # Statistics
    total_assessments = physio_assessments.count()
    first_visits = physio_assessments.filter(assessment_type='first_visit').count()
    follow_ups = physio_assessments.filter(assessment_type='follow_up').count()
    
    # Common diagnoses
    diagnoses = {}
    for assessment in physio_assessments:
        if assessment.diagnosis:
            diagnosis = assessment.diagnosis.strip()
            diagnoses[diagnosis] = diagnoses.get(diagnosis, 0) + 1
    
    common_diagnoses = sorted(diagnoses.items(), key=lambda x: x[1], reverse=True)[:10]
    
    # Treatment outcomes
    follow_up_required = physio_assessments.filter(follow_up_required=True).count()
    
    # Therapist performance
    therapist_stats = physio_assessments.values('assessed_by__first_name', 'assessed_by__last_name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Monthly trend
    monthly_labels = []
    monthly_data = []
    for i in range(6):
        month_start = (end_date.replace(day=1) - timedelta(days=i*30)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        count = Assessment.objects.filter(
            department='physiotherapy',
            assessment_date__range=[month_start, month_end]
        ).count()
        
        monthly_labels.insert(0, month_start.strftime('%b %Y'))
        monthly_data.insert(0, count)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_assessments': total_assessments,
        'first_visits': first_visits,
        'follow_ups': follow_ups,
        'common_diagnoses': common_diagnoses,
        'follow_up_required': follow_up_required,
        'therapist_stats': therapist_stats,
        'monthly_labels': json.dumps(monthly_labels),
        'monthly_data': json.dumps(monthly_data),
        'recent_assessments': physio_assessments.order_by('-assessment_date')[:20],
    }
    return render(request, 'reports/physiotherapy_reports.html', context)

@login_required
def nutrition_reports(request):
    """Nutrition department specific reports"""
    from patients.models import Assessment
    
    # Get filter parameters
    start_date = request.GET.get('start_date', date.today().replace(day=1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Nutrition assessments
    nutrition_assessments = Assessment.objects.filter(
        department='nutrition',
        assessment_date__range=[start_date, end_date]
    ).select_related('patient', 'assessed_by')
    
    # Statistics
    total_assessments = nutrition_assessments.count()
    first_visits = nutrition_assessments.filter(assessment_type='first_visit').count()
    follow_ups = nutrition_assessments.filter(assessment_type='follow_up').count()
    
    # Common diagnoses/conditions
    diagnoses = {}
    for assessment in nutrition_assessments:
        if assessment.diagnosis:
            diagnosis = assessment.diagnosis.strip()
            diagnoses[diagnosis] = diagnoses.get(diagnosis, 0) + 1
    
    common_conditions = sorted(diagnoses.items(), key=lambda x: x[1], reverse=True)[:10]
    
    # Follow-up tracking
    follow_up_required = nutrition_assessments.filter(follow_up_required=True).count()
    
    # Nutritionist performance
    nutritionist_stats = nutrition_assessments.values('assessed_by__first_name', 'assessed_by__last_name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Monthly trend
    monthly_labels = []
    monthly_data = []
    for i in range(6):
        month_start = (end_date.replace(day=1) - timedelta(days=i*30)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        count = Assessment.objects.filter(
            department='nutrition',
            assessment_date__range=[month_start, month_end]
        ).count()
        
        monthly_labels.insert(0, month_start.strftime('%b %Y'))
        monthly_data.insert(0, count)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_assessments': total_assessments,
        'first_visits': first_visits,
        'follow_ups': follow_ups,
        'common_conditions': common_conditions,
        'follow_up_required': follow_up_required,
        'nutritionist_stats': nutritionist_stats,
        'monthly_labels': json.dumps(monthly_labels),
        'monthly_data': json.dumps(monthly_data),
        'recent_assessments': nutrition_assessments.order_by('-assessment_date')[:20],
    }
    return render(request, 'reports/nutrition_reports.html', context)

@login_required
def clinical_summary_report(request):
    """Comprehensive clinical summary report across all departments"""
    from patients.models import Assessment, VitalSigns
    
    # Get filter parameters
    start_date = request.GET.get('start_date', date.today().replace(day=1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Assessment statistics by department
    physio_count = Assessment.objects.filter(
        department='physiotherapy',
        assessment_date__range=[start_date, end_date]
    ).count()
    
    nutrition_count = Assessment.objects.filter(
        department='nutrition',
        assessment_date__range=[start_date, end_date]
    ).count()
    
    general_count = Assessment.objects.filter(
        department='general',
        assessment_date__range=[start_date, end_date]
    ).count()
    
    # Vital signs monitoring
    vital_signs_count = VitalSigns.objects.filter(
        recorded_date__range=[start_date, end_date]
    ).count()
    
    # Patient engagement
    unique_patients = Assessment.objects.filter(
        assessment_date__range=[start_date, end_date]
    ).values('patient').distinct().count()
    
    # Department distribution
    dept_labels = ['Physiotherapy', 'Nutrition', 'General']
    dept_data = [physio_count, nutrition_count, general_count]
    
    # Assessment type distribution
    first_visits = Assessment.objects.filter(
        assessment_date__range=[start_date, end_date],
        assessment_type='first_visit'
    ).count()
    
    follow_ups = Assessment.objects.filter(
        assessment_date__range=[start_date, end_date],
        assessment_type='follow_up'
    ).count()
    
    # Recent assessments across all departments
    recent_assessments = Assessment.objects.filter(
        assessment_date__range=[start_date, end_date]
    ).select_related('patient', 'assessed_by').order_by('-assessment_date')[:30]
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'physio_count': physio_count,
        'nutrition_count': nutrition_count,
        'general_count': general_count,
        'vital_signs_count': vital_signs_count,
        'unique_patients': unique_patients,
        'first_visits': first_visits,
        'follow_ups': follow_ups,
        'dept_labels': json.dumps(dept_labels),
        'dept_data': json.dumps(dept_data),
        'recent_assessments': recent_assessments,
    }
    return render(request, 'reports/clinical_summary.html', context)


# ─────────────────────────────────────────────────────────
#  Custom Patient Report Generator
# ─────────────────────────────────────────────────────────

def _apply_custom_filters(request):
    """Parse all custom report filters and return a filtered Patient queryset + applied_filters dict."""
    from clinic_settings.models import EnabledModule
    from patients.models import PatientGroup

    enabled = EnabledModule.get_enabled_modules()
    patients = Patient.objects.filter(is_active=True)
    applied = {}

    # ── Demographics ──
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        patients = patients.filter(registration_date__date__gte=date_from)
        applied['Registration From'] = date_from
    if date_to:
        patients = patients.filter(registration_date__date__lte=date_to)
        applied['Registration To'] = date_to
    gender = request.GET.get('gender')
    if gender:
        patients = patients.filter(gender=gender)
        applied['Gender'] = dict(Patient.GENDER_CHOICES).get(gender, gender)
    blood_type = request.GET.get('blood_type')
    if blood_type:
        patients = patients.filter(blood_type=blood_type)
        applied['Blood Type'] = blood_type
    insurance = request.GET.get('insurance_provider')
    if insurance:
        patients = patients.filter(insurance_provider__icontains=insurance)
        applied['Insurance Provider'] = insurance
    group_id = request.GET.get('patient_group')
    if group_id:
        patients = patients.filter(patient_group_id=group_id)
        try:
            applied['Patient Group'] = PatientGroup.objects.get(pk=group_id).name
        except PatientGroup.DoesNotExist:
            pass
    city = request.GET.get('city')
    if city:
        patients = patients.filter(city__icontains=city)
        applied['City'] = city
    visit_reason = request.GET.get('visit_reason')
    if visit_reason:
        patients = patients.filter(reason_for_visit=visit_reason)
        applied['Visit Reason'] = dict(Patient.VISIT_REASON_CHOICES).get(visit_reason, visit_reason)

    # ── Laboratory Filters ──
    if 'laboratory' in enabled:
        lab_test_id = request.GET.get('lab_test')
        lab_category = request.GET.get('lab_category')
        lab_status = request.GET.get('lab_status')
        lab_priority = request.GET.get('lab_priority')
        lab_abnormal = request.GET.get('lab_abnormal')
        lab_date_from = request.GET.get('lab_date_from')
        lab_date_to = request.GET.get('lab_date_to')
        lq = Q()
        if lab_test_id:
            lq &= Q(lab_requests__test_id=lab_test_id)
            from laboratory.models import LabTest as LT
            try:
                applied['Lab Test'] = LT.objects.get(pk=lab_test_id).name
            except LT.DoesNotExist:
                pass
        if lab_category:
            lq &= Q(lab_requests__test__category=lab_category)
            applied['Lab Category'] = lab_category.title()
        if lab_status:
            lq &= Q(lab_requests__status=lab_status)
            applied['Lab Status'] = lab_status.replace('_', ' ').title()
        if lab_priority:
            lq &= Q(lab_requests__priority=lab_priority)
            applied['Lab Priority'] = lab_priority.title()
        if lab_abnormal == 'yes':
            lq &= Q(lab_requests__result__is_abnormal=True)
            applied['Abnormal Results'] = 'Yes'
        elif lab_abnormal == 'no':
            lq &= Q(lab_requests__result__is_abnormal=False)
            applied['Abnormal Results'] = 'No'
        if lab_date_from:
            lq &= Q(lab_requests__date_requested__date__gte=lab_date_from)
            applied['Lab Date From'] = lab_date_from
        if lab_date_to:
            lq &= Q(lab_requests__date_requested__date__lte=lab_date_to)
            applied['Lab Date To'] = lab_date_to
        if lq != Q():
            patients = patients.filter(lq).distinct()

    # ── Pharmacy Filters ──
    if 'pharmacy' in enabled:
        med_id = request.GET.get('medication')
        med_form = request.GET.get('med_form')
        rx_status = request.GET.get('rx_status')
        prescriber_id = request.GET.get('prescriber')
        rx_date_from = request.GET.get('rx_date_from')
        rx_date_to = request.GET.get('rx_date_to')
        rq = Q()
        if med_id:
            rq &= (Q(prescription__items__medication_id=med_id) | Q(prescription__medication_id=med_id))
            from pharmacy.models import Medication as Med
            try:
                applied['Medication'] = Med.objects.get(pk=med_id).name
            except Med.DoesNotExist:
                pass
        if med_form:
            rq &= Q(prescription__items__medication__form=med_form)
            applied['Medication Form'] = med_form.title()
        if rx_status:
            rq &= Q(prescription__status=rx_status)
            applied['Prescription Status'] = rx_status.title()
        if prescriber_id:
            rq &= Q(prescription__prescribed_by_id=prescriber_id)
            try:
                u = User.objects.get(pk=prescriber_id)
                applied['Prescriber'] = u.get_full_name()
            except User.DoesNotExist:
                pass
        if rx_date_from:
            rq &= Q(prescription__prescribed_date__date__gte=rx_date_from)
            applied['Rx Date From'] = rx_date_from
        if rx_date_to:
            rq &= Q(prescription__prescribed_date__date__lte=rx_date_to)
            applied['Rx Date To'] = rx_date_to
        if rq != Q():
            patients = patients.filter(rq).distinct()

    # ── Billing Filters ──
    if 'billing' in enabled:
        inv_status = request.GET.get('inv_status')
        payment_method = request.GET.get('payment_method')
        amount_min = request.GET.get('amount_min')
        amount_max = request.GET.get('amount_max')
        bill_date_from = request.GET.get('bill_date_from')
        bill_date_to = request.GET.get('bill_date_to')
        bq = Q()
        if inv_status:
            bq &= Q(invoices__status=inv_status)
            applied['Invoice Status'] = inv_status.title()
        if amount_min:
            bq &= Q(invoices__total_amount__gte=amount_min)
            applied['Min Amount'] = amount_min
        if amount_max:
            bq &= Q(invoices__total_amount__lte=amount_max)
            applied['Max Amount'] = amount_max
        if bill_date_from:
            bq &= Q(invoices__created_at__date__gte=bill_date_from)
            applied['Billing From'] = bill_date_from
        if bill_date_to:
            bq &= Q(invoices__created_at__date__lte=bill_date_to)
            applied['Billing To'] = bill_date_to
        if payment_method:
            bq &= Q(payments__payment_method=payment_method)
            applied['Payment Method'] = payment_method.replace('_', ' ').title()
        if bq != Q():
            patients = patients.filter(bq).distinct()

    # ── Appointment Filters ──
    if 'appointments' in enabled:
        service_id = request.GET.get('service')
        service_cat = request.GET.get('service_category')
        provider_id = request.GET.get('appt_provider')
        appt_status = request.GET.get('appt_status')
        appt_date_from = request.GET.get('appt_date_from')
        appt_date_to = request.GET.get('appt_date_to')
        aq = Q()
        if service_id:
            aq &= Q(appointments__service_id=service_id)
            try:
                applied['Service'] = Service.objects.get(pk=service_id).name
            except Service.DoesNotExist:
                pass
        if service_cat:
            aq &= Q(appointments__service__category=service_cat)
            applied['Service Category'] = service_cat.title()
        if provider_id:
            aq &= Q(appointments__provider_id=provider_id)
            try:
                u = User.objects.get(pk=provider_id)
                applied['Provider'] = u.get_full_name()
            except User.DoesNotExist:
                pass
        if appt_status:
            aq &= Q(appointments__status=appt_status)
            applied['Appointment Status'] = appt_status.replace('_', ' ').title()
        if appt_date_from:
            aq &= Q(appointments__appointment_date__gte=appt_date_from)
            applied['Appt Date From'] = appt_date_from
        if appt_date_to:
            aq &= Q(appointments__appointment_date__lte=appt_date_to)
            applied['Appt Date To'] = appt_date_to
        if aq != Q():
            patients = patients.filter(aq).distinct()

    # ── Age Filters (post-query) ──
    age_min = request.GET.get('age_min')
    age_max = request.GET.get('age_max')
    if age_min or age_max:
        filtered_ids = []
        for p in patients:
            try:
                age = p.get_age()
                if isinstance(age, (int, float)):
                    if age_min and int(age) < int(age_min):
                        continue
                    if age_max and int(age) > int(age_max):
                        continue
                    filtered_ids.append(p.pk)
            except (ValueError, TypeError, AttributeError):
                continue
        patients = Patient.objects.filter(pk__in=filtered_ids, is_active=True)
        if age_min:
            applied['Min Age'] = age_min
        if age_max:
            applied['Max Age'] = age_max

    return patients, applied, enabled


@login_required
def custom_patient_report(request):
    """Process custom report filters and render results on a new page."""
    from clinic_settings.models import EnabledModule
    from pharmacy.models import Prescription

    patients, applied, enabled = _apply_custom_filters(request)
    total_patients = patients.count()

    # Gather per-patient detail rows (cap at 500 for display)
    patient_data = []
    for p in patients.select_related('patient_group')[:500]:
        row = {
            'patient': p,
            'age': p.get_age(),
            'gender': p.get_gender_display() if p.gender else 'N/A',
            'blood_type': p.blood_type or 'N/A',
            'insurance': p.insurance_provider or 'None',
            'group': p.patient_group.name if p.patient_group else 'N/A',
            'registration_date': p.registration_date,
        }
        if 'laboratory' in enabled:
            row['lab_count'] = p.lab_requests.count()
        if 'pharmacy' in enabled:
            row['rx_count'] = Prescription.objects.filter(patient=p).count()
        if 'billing' in enabled:
            row['invoice_count'] = p.invoices.count()
            row['total_billed'] = p.invoices.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            row['total_paid'] = p.payments.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
        if 'appointments' in enabled:
            row['appt_count'] = p.appointments.count()
        patient_data.append(row)

    # Summary statistics
    summary = {'total': total_patients}
    gender_dist = patients.values('gender').annotate(cnt=Count('id'))
    summary['gender'] = {dict(Patient.GENDER_CHOICES).get(g['gender'], g['gender'] or 'Unknown'): g['cnt'] for g in gender_dist}
    if 'laboratory' in enabled:
        from laboratory.models import LabTestRequest
        summary['total_lab'] = LabTestRequest.objects.filter(patient__in=patients).count()
    if 'pharmacy' in enabled:
        summary['total_rx'] = Prescription.objects.filter(patient__in=patients).count()
    if 'billing' in enabled:
        inv_agg = Invoice.objects.filter(patient__in=patients).aggregate(total=Sum('total_amount'), cnt=Count('id'))
        summary['total_invoices'] = inv_agg['cnt']
        summary['total_billed'] = float(inv_agg['total'] or 0)
    if 'appointments' in enabled:
        summary['total_appts'] = Appointment.objects.filter(patient__in=patients).count()

    qs = request.GET.urlencode()
    context = {
        'patient_data': patient_data,
        'total_patients': total_patients,
        'summary': summary,
        'query_string': qs,
        'filters_applied': applied,
        'enabled_modules': enabled,
    }
    return render(request, 'reports/custom_report_results.html', context)


@login_required
def download_custom_patient_report(request):
    """Generate Excel workbook for custom filtered patient report with charts."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from pharmacy.models import Prescription

    patients, applied, enabled = _apply_custom_filters(request)
    total = patients.count()

    wb = Workbook()
    tf = Font(name='Calibri', size=14, bold=True, color='1B4F72')
    sf = Font(name='Calibri', size=11, bold=True, color='2E86C1')
    hf = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')
    alt = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
    bdr = Border(left=Side(style='thin'), right=Side(style='thin'),
                 top=Side(style='thin'), bottom=Side(style='thin'))
    money = '#,##0'

    def wh(ws, row, cols):
        for c, v in enumerate(cols, 1):
            cl = ws.cell(row=row, column=c, value=v)
            cl.font = hf; cl.fill = hfill; cl.alignment = Alignment(horizontal='center'); cl.border = bdr
        return row + 1

    def wr(ws, row, vals, fmt=None):
        for c, v in enumerate(vals, 1):
            cl = ws.cell(row=row, column=c, value=v)
            cl.border = bdr
            if fmt and c in fmt:
                cl.number_format = fmt[c]
            if row % 2 == 0:
                cl.fill = alt
        return row + 1

    def aw(ws, n):
        for c in range(1, n + 1):
            mx = 12
            for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
                for cell in row:
                    if cell.value:
                        mx = max(mx, len(str(cell.value)) + 2)
            ws.column_dimensions[get_column_letter(c)].width = min(mx, 40)

    def add_pie(ws, title, labels, values, anchor, w=13, h=9):
        if not labels or not values or not any(values):
            return
        base = ws.max_row + 3
        for i, (l, v) in enumerate(zip(labels, values)):
            ws.cell(row=base + i, column=10, value=str(l))
            ws.cell(row=base + i, column=11, value=v)
        chart = PieChart()
        chart.title = title; chart.style = 10; chart.width = w; chart.height = h
        chart.add_data(Reference(ws, min_col=11, min_row=base, max_row=base + len(labels) - 1))
        chart.set_categories(Reference(ws, min_col=10, min_row=base, max_row=base + len(labels) - 1))
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True; chart.dataLabels.showCatName = True
        ws.add_chart(chart, anchor)

    def add_bar(ws, title, labels, values, anchor, w=14, h=9):
        if not labels or not values:
            return
        base = ws.max_row + 3
        ws.cell(row=base, column=10, value='Category'); ws.cell(row=base, column=11, value='Count')
        for i, (l, v) in enumerate(zip(labels, values)):
            ws.cell(row=base + 1 + i, column=10, value=str(l))
            ws.cell(row=base + 1 + i, column=11, value=v)
        chart = BarChart()
        chart.type = 'col'; chart.style = 10; chart.title = title; chart.width = w; chart.height = h
        chart.add_data(Reference(ws, min_col=11, min_row=base, max_row=base + len(labels)), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=10, min_row=base + 1, max_row=base + len(labels)))
        ws.add_chart(chart, anchor)

    # ── Summary Sheet ──
    ws = wb.active
    ws.title = 'Report Summary'
    ws.cell(row=1, column=1, value='Custom Patient Report').font = tf
    ws.cell(row=2, column=1, value=f'Generated: {date.today()}').font = Font(name='Calibri', size=10, italic=True)
    r = 4
    if applied:
        ws.cell(row=r, column=1, value='Applied Filters').font = sf; r += 1
        r = wh(ws, r, ['Filter', 'Value'])
        for k, v in applied.items():
            r = wr(ws, r, [k, v])
        r += 1
    ws.cell(row=r, column=1, value='Summary').font = sf; r += 1
    r = wh(ws, r, ['Metric', 'Value'])
    r = wr(ws, r, ['Total Matching Patients', total])
    gender_dist = patients.values('gender').annotate(cnt=Count('id'))
    for g in gender_dist:
        label = dict(Patient.GENDER_CHOICES).get(g['gender'], g['gender'] or 'Unknown')
        r = wr(ws, r, [f'  {label}', g['cnt']])
    if 'laboratory' in enabled:
        from laboratory.models import LabTestRequest
        r = wr(ws, r, ['Total Lab Requests', LabTestRequest.objects.filter(patient__in=patients).count()])
    if 'pharmacy' in enabled:
        r = wr(ws, r, ['Total Prescriptions', Prescription.objects.filter(patient__in=patients).count()])
    if 'billing' in enabled:
        agg = Invoice.objects.filter(patient__in=patients).aggregate(t=Sum('total_amount'), c=Count('id'))
        r = wr(ws, r, ['Total Invoices', agg['c']])
        r = wr(ws, r, ['Total Billed', float(agg['t'] or 0)], fmt={2: money})
    if 'appointments' in enabled:
        r = wr(ws, r, ['Total Appointments', Appointment.objects.filter(patient__in=patients).count()])
    aw(ws, 2)

    # Gender pie on summary
    g_labels = [dict(Patient.GENDER_CHOICES).get(g['gender'], g['gender'] or 'Unknown') for g in gender_dist]
    g_values = [g['cnt'] for g in gender_dist]
    add_pie(ws, 'Gender Distribution', g_labels, g_values, 'E4')

    # ── Patient Details Sheet ──
    ws2 = wb.create_sheet('Patient Details')
    ws2.cell(row=1, column=1, value='Patient Details').font = tf
    r = 3
    headers = ['Patient ID', 'Name', 'Gender', 'Age', 'Blood Type', 'Insurance', 'Group', 'Registration Date']
    col_count = len(headers)
    if 'laboratory' in enabled:
        headers.append('Lab Tests')
        col_count += 1
    if 'pharmacy' in enabled:
        headers.append('Prescriptions')
        col_count += 1
    if 'billing' in enabled:
        headers.extend(['Invoices', 'Total Billed', 'Total Paid'])
        col_count += 3
    if 'appointments' in enabled:
        headers.append('Appointments')
        col_count += 1
    r = wh(ws2, r, headers)
    for p in patients.select_related('patient_group')[:500]:
        vals = [
            p.patient_id,
            p.get_full_name(),
            p.get_gender_display() if p.gender else 'N/A',
            p.get_age() if isinstance(p.get_age(), (int, float)) else 'N/A',
            p.blood_type or 'N/A',
            p.insurance_provider or 'None',
            p.patient_group.name if p.patient_group else 'N/A',
            p.registration_date.strftime('%Y-%m-%d') if p.registration_date else '',
        ]
        fmt_map = {}
        if 'laboratory' in enabled:
            vals.append(p.lab_requests.count())
        if 'pharmacy' in enabled:
            vals.append(Prescription.objects.filter(patient=p).count())
        if 'billing' in enabled:
            inv_cnt = p.invoices.count()
            billed = p.invoices.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            paid = p.payments.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
            vals.extend([inv_cnt, float(billed), float(paid)])
            idx = len(vals)
            fmt_map[idx - 1] = money
            fmt_map[idx] = money
        if 'appointments' in enabled:
            vals.append(p.appointments.count())
        r = wr(ws2, r, vals, fmt=fmt_map)
    aw(ws2, col_count)

    # ── Lab Details Sheet ──
    if 'laboratory' in enabled:
        from laboratory.models import LabTestRequest, LabTestResult
        ws3 = wb.create_sheet('Lab Test Details')
        ws3.cell(row=1, column=1, value='Lab Test Details').font = tf
        r = 3
        r = wh(ws3, r, ['Patient ID', 'Patient Name', 'Test', 'Category', 'Status', 'Priority', 'Date Requested', 'Abnormal'])
        reqs_base = LabTestRequest.objects.filter(patient__in=patients).select_related('patient', 'test').order_by('-date_requested')
        reqs = list(reqs_base[:1000])
        for req in reqs:
            abnormal = ''
            try:
                abnormal = 'Yes' if req.result.is_abnormal else 'No'
            except LabTestResult.DoesNotExist:
                abnormal = 'N/A'
            r = wr(ws3, r, [
                req.patient.patient_id, req.patient.get_full_name(),
                req.test.name, req.test.get_category_display(),
                req.get_status_display(), req.get_priority_display(),
                req.date_requested.strftime('%Y-%m-%d'), abnormal,
            ])
        aw(ws3, 8)
        # Test category chart
        cat_qs = reqs_base.values('test__category').annotate(cnt=Count('id')).order_by('-cnt')
        add_pie(ws3, 'Tests by Category', [c['test__category'] for c in cat_qs], [c['cnt'] for c in cat_qs], 'J3')
        # Status chart
        st_qs = reqs_base.values('status').annotate(cnt=Count('id')).order_by('-cnt')
        add_pie(ws3, 'Tests by Status', [s['status'].title() for s in st_qs], [s['cnt'] for s in st_qs], 'J20')

    # ── Prescription Details Sheet ──
    if 'pharmacy' in enabled:
        from pharmacy.models import PrescriptionItem
        ws4 = wb.create_sheet('Prescription Details')
        ws4.cell(row=1, column=1, value='Prescription Details').font = tf
        r = 3
        r = wh(ws4, r, ['Patient ID', 'Patient Name', 'Medication', 'Dosage', 'Frequency', 'Quantity', 'Status', 'Prescribed Date', 'Prescriber'])
        items_base = PrescriptionItem.objects.filter(
            prescription__patient__in=patients
        ).select_related('prescription__patient', 'prescription__prescribed_by', 'medication').order_by('-prescription__prescribed_date')
        items = list(items_base[:1000])
        for it in items:
            rx = it.prescription
            r = wr(ws4, r, [
                rx.patient.patient_id if rx.patient else '', rx.patient.get_full_name() if rx.patient else '',
                it.medication.name, it.dosage, it.frequency, it.quantity,
                rx.get_status_display(), rx.prescribed_date.strftime('%Y-%m-%d'),
                rx.prescribed_by.get_full_name() if rx.prescribed_by else 'N/A',
            ])
        aw(ws4, 9)
        # Top medications chart
        top = list(items_base.values('medication__name').annotate(cnt=Count('id')).order_by('-cnt')[:7])
        add_bar(ws4, 'Top Medications', [t['medication__name'] for t in top], [t['cnt'] for t in top], 'K3')

    # ── Billing Details Sheet ──
    if 'billing' in enabled:
        ws5 = wb.create_sheet('Billing Details')
        ws5.cell(row=1, column=1, value='Billing Details').font = tf
        r = 3
        r = wh(ws5, r, ['Patient ID', 'Patient Name', 'Invoice #', 'Status', 'Total Amount', 'Paid', 'Balance', 'Issue Date'])
        invs_base = Invoice.objects.filter(patient__in=patients).select_related('patient').order_by('-created_at')
        invs = list(invs_base[:1000])
        for inv in invs:
            paid = inv.get_total_paid()
            r = wr(ws5, r, [
                inv.patient.patient_id, inv.patient.get_full_name(),
                inv.invoice_number, inv.get_status_display(),
                float(inv.total_amount), float(paid), float(inv.total_amount - paid),
                inv.issue_date.strftime('%Y-%m-%d') if inv.issue_date else '',
            ], fmt={5: money, 6: money, 7: money})
        aw(ws5, 8)
        st_qs = invs_base.values('status').annotate(cnt=Count('id')).order_by('-cnt')
        add_pie(ws5, 'Invoices by Status', [s['status'].title() for s in st_qs], [s['cnt'] for s in st_qs], 'J3')

    # ── Appointment Details Sheet ──
    if 'appointments' in enabled:
        ws6 = wb.create_sheet('Appointment Details')
        ws6.cell(row=1, column=1, value='Appointment Details').font = tf
        r = 3
        r = wh(ws6, r, ['Patient ID', 'Patient Name', 'Service', 'Category', 'Provider', 'Date', 'Time', 'Status'])
        appts_base = Appointment.objects.filter(patient__in=patients).select_related('patient', 'service', 'provider').order_by('-appointment_date')
        appts = list(appts_base[:1000])
        for a in appts:
            r = wr(ws6, r, [
                a.patient.patient_id, a.patient.get_full_name(),
                a.service.name, a.service.get_category_display(),
                a.provider.get_full_name() if a.provider else 'N/A',
                a.appointment_date.strftime('%Y-%m-%d'), a.appointment_time.strftime('%H:%M'),
                a.get_status_display(),
            ])
        aw(ws6, 8)
        st_qs = appts_base.values('status').annotate(cnt=Count('id')).order_by('-cnt')
        add_pie(ws6, 'Appointment Status', [s['status'].replace('_', ' ').title() for s in st_qs], [s['cnt'] for s in st_qs], 'J3')
        sv_qs = list(appts_base.values('service__name').annotate(cnt=Count('id')).order_by('-cnt')[:7])
        add_bar(ws6, 'By Service', [s['service__name'] for s in sv_qs], [s['cnt'] for s in sv_qs], 'J20')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Custom_Patient_Report_{date.today()}.xlsx"'
    return response


# ─────────────────────────────────────────────────────────
#  Custom Financial Report Generator
# ─────────────────────────────────────────────────────────

def _apply_custom_financial_filters(request):
    """Parse financial report filters and return filtered querysets + applied dict."""
    from clinic_settings.models import EnabledModule
    enabled = EnabledModule.get_enabled_modules()
    applied = {}

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        applied['Date From'] = date_from
    if date_to:
        applied['Date To'] = date_to

    invoices = Invoice.objects.none()
    payments = Payment.objects.none()
    appointments_qs = Appointment.objects.none()

    # ── Billing filters ──
    if 'billing' in enabled:
        invoices = Invoice.objects.all()
        payments = Payment.objects.all()
        if date_from:
            invoices = invoices.filter(created_at__date__gte=date_from)
            payments = payments.filter(payment_date__date__gte=date_from)
        if date_to:
            invoices = invoices.filter(created_at__date__lte=date_to)
            payments = payments.filter(payment_date__date__lte=date_to)
        inv_status = request.GET.get('inv_status')
        if inv_status:
            invoices = invoices.filter(status=inv_status)
            applied['Invoice Status'] = inv_status.title()
        pay_method = request.GET.get('payment_method')
        if pay_method:
            payments = payments.filter(payment_method=pay_method)
            applied['Payment Method'] = pay_method.replace('_', ' ').title()
        amt_min = request.GET.get('amount_min')
        if amt_min:
            invoices = invoices.filter(total_amount__gte=amt_min)
            applied['Min Amount'] = amt_min
        amt_max = request.GET.get('amount_max')
        if amt_max:
            invoices = invoices.filter(total_amount__lte=amt_max)
            applied['Max Amount'] = amt_max

    # ── Appointment filters ──
    if 'appointments' in enabled:
        appointments_qs = Appointment.objects.all()
        if date_from:
            appointments_qs = appointments_qs.filter(appointment_date__gte=date_from)
        if date_to:
            appointments_qs = appointments_qs.filter(appointment_date__lte=date_to)
        svc_id = request.GET.get('service')
        if svc_id:
            appointments_qs = appointments_qs.filter(service_id=svc_id)
            try:
                applied['Service'] = Service.objects.get(pk=svc_id).name
            except Service.DoesNotExist:
                pass
        svc_cat = request.GET.get('service_category')
        if svc_cat:
            appointments_qs = appointments_qs.filter(service__category=svc_cat)
            applied['Service Category'] = svc_cat.title()
        prov = request.GET.get('provider')
        if prov:
            appointments_qs = appointments_qs.filter(provider_id=prov)
            try:
                applied['Provider'] = User.objects.get(pk=prov).get_full_name()
            except User.DoesNotExist:
                pass
        a_status = request.GET.get('appt_status')
        if a_status:
            appointments_qs = appointments_qs.filter(status=a_status)
            applied['Appt Status'] = a_status.replace('_', ' ').title()

    return invoices, payments, appointments_qs, applied, enabled


@login_required
def custom_financial_report(request):
    """Render custom financial report results page."""
    invoices, payments, appointments_qs, applied, enabled = _apply_custom_financial_filters(request)

    summary = {}
    inv_data = []
    if 'billing' in enabled:
        inv_agg = invoices.aggregate(total=Sum('total_amount'), cnt=Count('id'))
        pay_agg = payments.filter(status='completed').aggregate(total=Sum('amount'), cnt=Count('id'))
        summary['total_invoices'] = inv_agg['cnt']
        summary['total_billed'] = float(inv_agg['total'] or 0)
        summary['total_payments'] = pay_agg['cnt']
        summary['total_collected'] = float(pay_agg['total'] or 0)
        summary['outstanding'] = summary['total_billed'] - summary['total_collected']
        for inv in invoices.select_related('patient').order_by('-created_at')[:500]:
            paid = inv.get_total_paid()
            inv_data.append({
                'invoice': inv,
                'paid': paid,
                'balance': inv.total_amount - paid,
            })

    appt_data = []
    if 'appointments' in enabled:
        summary['total_appts'] = appointments_qs.count()
        summary['completed_appts'] = appointments_qs.filter(status='completed').count()
        for a in appointments_qs.select_related('patient', 'service', 'provider').order_by('-appointment_date')[:500]:
            appt_data.append(a)

    context = {
        'inv_data': inv_data,
        'appt_data': appt_data,
        'summary': summary,
        'filters_applied': applied,
        'enabled_modules': enabled,
        'query_string': request.GET.urlencode(),
    }
    return render(request, 'reports/custom_financial_results.html', context)


@login_required
def download_custom_financial_report(request):
    """Generate Excel workbook for custom filtered financial report."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList

    invoices, payments, appointments_qs, applied, enabled = _apply_custom_financial_filters(request)

    wb = Workbook()
    tf = Font(name='Calibri', size=14, bold=True, color='1B4F72')
    sf = Font(name='Calibri', size=11, bold=True, color='2E86C1')
    hf = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')
    alt_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
    bdr = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    money = '#,##0'

    def wh(ws, row, cols):
        for c, v in enumerate(cols, 1):
            cl = ws.cell(row=row, column=c, value=v)
            cl.font = hf; cl.fill = hfill; cl.alignment = Alignment(horizontal='center'); cl.border = bdr
        return row + 1

    def wr(ws, row, vals, fmt=None):
        for c, v in enumerate(vals, 1):
            cl = ws.cell(row=row, column=c, value=v)
            cl.border = bdr
            if fmt and c in fmt:
                cl.number_format = fmt[c]
            if row % 2 == 0:
                cl.fill = alt_fill
        return row + 1

    def aw(ws, n):
        for c in range(1, n + 1):
            mx = 12
            for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
                for cell in row:
                    if cell.value:
                        mx = max(mx, len(str(cell.value)) + 2)
            ws.column_dimensions[get_column_letter(c)].width = min(mx, 40)

    def add_pie(ws, title, labels, values, anchor):
        if not labels or not values or not any(values):
            return
        base = ws.max_row + 3
        for i, (l, v) in enumerate(zip(labels, values)):
            ws.cell(row=base + i, column=10, value=str(l))
            ws.cell(row=base + i, column=11, value=v)
        chart = PieChart(); chart.title = title; chart.style = 10; chart.width = 13; chart.height = 9
        chart.add_data(Reference(ws, min_col=11, min_row=base, max_row=base + len(labels) - 1))
        chart.set_categories(Reference(ws, min_col=10, min_row=base, max_row=base + len(labels) - 1))
        chart.dataLabels = DataLabelList(); chart.dataLabels.showPercent = True; chart.dataLabels.showCatName = True
        ws.add_chart(chart, anchor)

    # ── Summary Sheet ──
    ws = wb.active; ws.title = 'Summary'
    ws.cell(row=1, column=1, value='Custom Financial Report').font = tf
    ws.cell(row=2, column=1, value=f'Generated: {date.today()}').font = Font(name='Calibri', size=10, italic=True)
    r = 4
    if applied:
        ws.cell(row=r, column=1, value='Applied Filters').font = sf; r += 1
        r = wh(ws, r, ['Filter', 'Value'])
        for k, v in applied.items():
            r = wr(ws, r, [k, v])
        r += 1
    ws.cell(row=r, column=1, value='Summary').font = sf; r += 1
    r = wh(ws, r, ['Metric', 'Value'])
    if 'billing' in enabled:
        inv_agg = invoices.aggregate(total=Sum('total_amount'), cnt=Count('id'))
        pay_agg = payments.filter(status='completed').aggregate(total=Sum('amount'), cnt=Count('id'))
        r = wr(ws, r, ['Total Invoices', inv_agg['cnt']])
        r = wr(ws, r, ['Total Billed', float(inv_agg['total'] or 0)], fmt={2: money})
        r = wr(ws, r, ['Total Payments', pay_agg['cnt']])
        r = wr(ws, r, ['Total Collected', float(pay_agg['total'] or 0)], fmt={2: money})
    if 'appointments' in enabled:
        r = wr(ws, r, ['Total Appointments', appointments_qs.count()])
        r = wr(ws, r, ['Completed', appointments_qs.filter(status='completed').count()])
    aw(ws, 2)

    # ── Invoice Sheet ──
    if 'billing' in enabled:
        ws2 = wb.create_sheet('Invoices')
        ws2.cell(row=1, column=1, value='Invoice Details').font = tf
        r = 3
        r = wh(ws2, r, ['Invoice #', 'Patient', 'Status', 'Total Amount', 'Paid', 'Balance', 'Issue Date'])
        invs_base = invoices.select_related('patient').order_by('-created_at')
        for inv in list(invs_base[:1000]):
            paid = inv.get_total_paid()
            r = wr(ws2, r, [
                inv.invoice_number, inv.patient.get_full_name() if inv.patient else '',
                inv.get_status_display(), float(inv.total_amount), float(paid),
                float(inv.total_amount - paid),
                inv.issue_date.strftime('%Y-%m-%d') if inv.issue_date else '',
            ], fmt={4: money, 5: money, 6: money})
        aw(ws2, 7)
        st_qs = invoices.values('status').annotate(cnt=Count('id')).order_by('-cnt')
        add_pie(ws2, 'Invoices by Status', [s['status'].title() for s in st_qs], [s['cnt'] for s in st_qs], 'I3')

        # ── Payments Sheet ──
        ws3 = wb.create_sheet('Payments')
        ws3.cell(row=1, column=1, value='Payment Details').font = tf
        r = 3
        r = wh(ws3, r, ['Payment ID', 'Patient', 'Invoice', 'Amount', 'Method', 'Status', 'Date'])
        for pay in list(payments.select_related('patient', 'invoice').order_by('-payment_date')[:1000]):
            r = wr(ws3, r, [
                pay.payment_id, pay.patient.get_full_name() if pay.patient else '',
                pay.invoice.invoice_number if pay.invoice else 'N/A',
                float(pay.amount), pay.get_payment_method_display(),
                pay.get_status_display(),
                pay.payment_date.strftime('%Y-%m-%d') if pay.payment_date else '',
            ], fmt={4: money})
        aw(ws3, 7)
        pm_qs = payments.values('payment_method').annotate(cnt=Count('id')).order_by('-cnt')
        add_pie(ws3, 'Payment Methods', [p['payment_method'].replace('_', ' ').title() for p in pm_qs], [p['cnt'] for p in pm_qs], 'I3')

    # ── Appointments Sheet ──
    if 'appointments' in enabled:
        ws4 = wb.create_sheet('Appointments')
        ws4.cell(row=1, column=1, value='Appointment Details').font = tf
        r = 3
        r = wh(ws4, r, ['Patient', 'Service', 'Category', 'Provider', 'Date', 'Time', 'Status'])
        appts_base = appointments_qs.select_related('patient', 'service', 'provider').order_by('-appointment_date')
        for a in list(appts_base[:1000]):
            r = wr(ws4, r, [
                a.patient.get_full_name(), a.service.name, a.service.get_category_display(),
                a.provider.get_full_name() if a.provider else 'N/A',
                a.appointment_date.strftime('%Y-%m-%d'), a.appointment_time.strftime('%H:%M'),
                a.get_status_display(),
            ])
        aw(ws4, 7)
        st_qs = appointments_qs.values('status').annotate(cnt=Count('id')).order_by('-cnt')
        add_pie(ws4, 'Status', [s['status'].replace('_', ' ').title() for s in st_qs], [s['cnt'] for s in st_qs], 'I3')

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Custom_Financial_Report_{date.today()}.xlsx"'
    return response


# ─────────────────────────────────────────────────────────
#  Custom Appointment Report Generator
# ─────────────────────────────────────────────────────────

def _apply_custom_appointment_filters(request):
    """Parse appointment report filters and return filtered queryset + applied dict."""
    applied = {}
    qs = Appointment.objects.all()
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        qs = qs.filter(appointment_date__gte=date_from); applied['Date From'] = date_from
    if date_to:
        qs = qs.filter(appointment_date__lte=date_to); applied['Date To'] = date_to
    svc_id = request.GET.get('service')
    if svc_id:
        qs = qs.filter(service_id=svc_id)
        try:
            applied['Service'] = Service.objects.get(pk=svc_id).name
        except Service.DoesNotExist:
            pass
    svc_cat = request.GET.get('service_category')
    if svc_cat:
        qs = qs.filter(service__category=svc_cat); applied['Category'] = svc_cat.title()
    prov = request.GET.get('provider')
    if prov:
        qs = qs.filter(provider_id=prov)
        try:
            applied['Provider'] = User.objects.get(pk=prov).get_full_name()
        except User.DoesNotExist:
            pass
    status = request.GET.get('appt_status')
    if status:
        qs = qs.filter(status=status); applied['Status'] = status.replace('_', ' ').title()
    patient_name = request.GET.get('patient_name')
    if patient_name:
        qs = qs.filter(Q(patient__first_name__icontains=patient_name) | Q(patient__last_name__icontains=patient_name))
        applied['Patient Name'] = patient_name
    return qs, applied


@login_required
def custom_appointment_report(request):
    """Render custom appointment report results page."""
    qs, applied = _apply_custom_appointment_filters(request)
    total = qs.count()
    completed = qs.filter(status='completed').count()
    cancelled = qs.filter(status='cancelled').count()
    no_show = qs.filter(status='no_show').count()
    scheduled = qs.filter(status='scheduled').count()

    by_service = list(qs.values('service__name').annotate(cnt=Count('id')).order_by('-cnt')[:10])
    by_provider = list(qs.values('provider__first_name', 'provider__last_name').annotate(cnt=Count('id')).order_by('-cnt')[:10])

    appt_list = qs.select_related('patient', 'service', 'provider').order_by('-appointment_date')[:500]

    context = {
        'appt_list': appt_list,
        'total': total, 'completed': completed, 'cancelled': cancelled,
        'no_show': no_show, 'scheduled': scheduled,
        'completion_rate': round(completed / total * 100, 1) if total else 0,
        'by_service': by_service,
        'by_provider': by_provider,
        'filters_applied': applied,
        'query_string': request.GET.urlencode(),
    }
    return render(request, 'reports/custom_appointment_results.html', context)


@login_required
def download_custom_appointment_report(request):
    """Generate Excel workbook for custom filtered appointment report."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList

    qs, applied = _apply_custom_appointment_filters(request)

    wb = Workbook()
    tf = Font(name='Calibri', size=14, bold=True, color='1B4F72')
    sf = Font(name='Calibri', size=11, bold=True, color='2E86C1')
    hf = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')
    alt_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
    bdr = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def wh(ws, row, cols):
        for c, v in enumerate(cols, 1):
            cl = ws.cell(row=row, column=c, value=v)
            cl.font = hf; cl.fill = hfill; cl.alignment = Alignment(horizontal='center'); cl.border = bdr
        return row + 1

    def wr(ws, row, vals):
        for c, v in enumerate(vals, 1):
            cl = ws.cell(row=row, column=c, value=v); cl.border = bdr
            if row % 2 == 0:
                cl.fill = alt_fill
        return row + 1

    def aw(ws, n):
        for c in range(1, n + 1):
            mx = 12
            for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
                for cell in row:
                    if cell.value:
                        mx = max(mx, len(str(cell.value)) + 2)
            ws.column_dimensions[get_column_letter(c)].width = min(mx, 40)

    def add_pie(ws, title, labels, values, anchor):
        if not labels or not values or not any(values):
            return
        base = ws.max_row + 3
        for i, (l, v) in enumerate(zip(labels, values)):
            ws.cell(row=base + i, column=10, value=str(l)); ws.cell(row=base + i, column=11, value=v)
        chart = PieChart(); chart.title = title; chart.style = 10; chart.width = 13; chart.height = 9
        chart.add_data(Reference(ws, min_col=11, min_row=base, max_row=base + len(labels) - 1))
        chart.set_categories(Reference(ws, min_col=10, min_row=base, max_row=base + len(labels) - 1))
        chart.dataLabels = DataLabelList(); chart.dataLabels.showPercent = True; chart.dataLabels.showCatName = True
        ws.add_chart(chart, anchor)

    # Summary
    ws = wb.active; ws.title = 'Summary'
    ws.cell(row=1, column=1, value='Custom Appointment Report').font = tf
    ws.cell(row=2, column=1, value=f'Generated: {date.today()}').font = Font(name='Calibri', size=10, italic=True)
    r = 4
    if applied:
        ws.cell(row=r, column=1, value='Applied Filters').font = sf; r += 1
        r = wh(ws, r, ['Filter', 'Value'])
        for k, v in applied.items():
            r = wr(ws, r, [k, v])
        r += 1
    total = qs.count()
    ws.cell(row=r, column=1, value='Summary').font = sf; r += 1
    r = wh(ws, r, ['Metric', 'Value'])
    r = wr(ws, r, ['Total Appointments', total])
    r = wr(ws, r, ['Completed', qs.filter(status='completed').count()])
    r = wr(ws, r, ['Scheduled', qs.filter(status='scheduled').count()])
    r = wr(ws, r, ['Cancelled', qs.filter(status='cancelled').count()])
    r = wr(ws, r, ['No Show', qs.filter(status='no_show').count()])
    aw(ws, 2)

    status_labels = ['Completed', 'Scheduled', 'Cancelled', 'No Show']
    status_vals = [qs.filter(status='completed').count(), qs.filter(status='scheduled').count(),
                   qs.filter(status='cancelled').count(), qs.filter(status='no_show').count()]
    add_pie(ws, 'Status Distribution', status_labels, status_vals, 'E4')

    # Details
    ws2 = wb.create_sheet('Appointment Details')
    ws2.cell(row=1, column=1, value='Appointment Details').font = tf
    r = 3
    r = wh(ws2, r, ['Patient ID', 'Patient Name', 'Service', 'Category', 'Provider', 'Date', 'Time', 'Status'])
    appts_base = qs.select_related('patient', 'service', 'provider').order_by('-appointment_date')
    for a in list(appts_base[:1000]):
        r = wr(ws2, r, [
            a.patient.patient_id, a.patient.get_full_name(),
            a.service.name, a.service.get_category_display(),
            a.provider.get_full_name() if a.provider else 'N/A',
            a.appointment_date.strftime('%Y-%m-%d'), a.appointment_time.strftime('%H:%M'),
            a.get_status_display(),
        ])
    aw(ws2, 8)
    svc_qs = qs.values('service__name').annotate(cnt=Count('id')).order_by('-cnt')
    add_pie(ws2, 'By Service', [s['service__name'] for s in svc_qs], [s['cnt'] for s in svc_qs], 'J3')

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Custom_Appointment_Report_{date.today()}.xlsx"'
    return response


# ─────────────────────────────────────────────────────────
#  Custom Department (Physio / Nutrition / Clinical) Report
# ─────────────────────────────────────────────────────────

def _apply_custom_department_filters(request, department=None):
    """Parse department assessment filters and return filtered queryset + applied dict."""
    from patients.models import Assessment
    applied = {}
    qs = Assessment.objects.all()
    if department:
        qs = qs.filter(department=department)
        applied['Department'] = department.title()
    else:
        dept = request.GET.get('department')
        if dept:
            qs = qs.filter(department=dept); applied['Department'] = dept.title()
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        qs = qs.filter(assessment_date__gte=date_from); applied['Date From'] = date_from
    if date_to:
        qs = qs.filter(assessment_date__lte=date_to); applied['Date To'] = date_to
    a_type = request.GET.get('assessment_type')
    if a_type:
        qs = qs.filter(assessment_type=a_type); applied['Assessment Type'] = a_type.replace('_', ' ').title()
    assessor = request.GET.get('assessor')
    if assessor:
        qs = qs.filter(assessed_by_id=assessor)
        try:
            applied['Assessed By'] = User.objects.get(pk=assessor).get_full_name()
        except User.DoesNotExist:
            pass
    diag = request.GET.get('diagnosis')
    if diag:
        qs = qs.filter(diagnosis__icontains=diag); applied['Diagnosis Contains'] = diag
    follow_up = request.GET.get('follow_up')
    if follow_up == 'yes':
        qs = qs.filter(follow_up_required=True); applied['Follow-up Required'] = 'Yes'
    elif follow_up == 'no':
        qs = qs.filter(follow_up_required=False); applied['Follow-up Required'] = 'No'
    patient_name = request.GET.get('patient_name')
    if patient_name:
        qs = qs.filter(Q(patient__first_name__icontains=patient_name) | Q(patient__last_name__icontains=patient_name))
        applied['Patient Name'] = patient_name
    return qs, applied


@login_required
def custom_department_report(request):
    """Render custom department assessment report results page."""
    dept_param = request.GET.get('dept_type', '')
    fixed_dept = dept_param if dept_param in ('physiotherapy', 'nutrition') else None
    qs, applied = _apply_custom_department_filters(request, department=fixed_dept)
    total = qs.count()
    first_visits = qs.filter(assessment_type='first_visit').count()
    follow_ups = qs.filter(assessment_type='follow_up').count()
    follow_up_required = qs.filter(follow_up_required=True).count()

    by_dept = list(qs.values('department').annotate(cnt=Count('id')).order_by('-cnt'))
    by_assessor = list(qs.values('assessed_by__first_name', 'assessed_by__last_name').annotate(cnt=Count('id')).order_by('-cnt')[:10])

    assess_list = qs.select_related('patient', 'assessed_by').order_by('-assessment_date')[:500]

    context = {
        'assess_list': assess_list,
        'total': total, 'first_visits': first_visits, 'follow_ups': follow_ups,
        'follow_up_required': follow_up_required,
        'by_dept': by_dept, 'by_assessor': by_assessor,
        'filters_applied': applied,
        'query_string': request.GET.urlencode(),
        'dept_type': fixed_dept or 'all',
    }
    return render(request, 'reports/custom_department_results.html', context)


@login_required
def download_custom_department_report(request):
    """Generate Excel workbook for custom filtered department report."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList

    dept_param = request.GET.get('dept_type', '')
    fixed_dept = dept_param if dept_param in ('physiotherapy', 'nutrition') else None
    qs, applied = _apply_custom_department_filters(request, department=fixed_dept)

    wb = Workbook()
    tf = Font(name='Calibri', size=14, bold=True, color='1B4F72')
    sf = Font(name='Calibri', size=11, bold=True, color='2E86C1')
    hf = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')
    alt_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
    bdr = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def wh(ws, row, cols):
        for c, v in enumerate(cols, 1):
            cl = ws.cell(row=row, column=c, value=v)
            cl.font = hf; cl.fill = hfill; cl.alignment = Alignment(horizontal='center'); cl.border = bdr
        return row + 1

    def wr(ws, row, vals):
        for c, v in enumerate(vals, 1):
            cl = ws.cell(row=row, column=c, value=v); cl.border = bdr
            if row % 2 == 0:
                cl.fill = alt_fill
        return row + 1

    def aw(ws, n):
        for c in range(1, n + 1):
            mx = 12
            for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
                for cell in row:
                    if cell.value:
                        mx = max(mx, len(str(cell.value)) + 2)
            ws.column_dimensions[get_column_letter(c)].width = min(mx, 40)

    def add_pie(ws, title, labels, values, anchor):
        if not labels or not values or not any(values):
            return
        base = ws.max_row + 3
        for i, (l, v) in enumerate(zip(labels, values)):
            ws.cell(row=base + i, column=10, value=str(l)); ws.cell(row=base + i, column=11, value=v)
        chart = PieChart(); chart.title = title; chart.style = 10; chart.width = 13; chart.height = 9
        chart.add_data(Reference(ws, min_col=11, min_row=base, max_row=base + len(labels) - 1))
        chart.set_categories(Reference(ws, min_col=10, min_row=base, max_row=base + len(labels) - 1))
        chart.dataLabels = DataLabelList(); chart.dataLabels.showPercent = True; chart.dataLabels.showCatName = True
        ws.add_chart(chart, anchor)

    label = (fixed_dept or 'Department').title()
    ws = wb.active; ws.title = 'Summary'
    ws.cell(row=1, column=1, value=f'Custom {label} Report').font = tf
    ws.cell(row=2, column=1, value=f'Generated: {date.today()}').font = Font(name='Calibri', size=10, italic=True)
    r = 4
    if applied:
        ws.cell(row=r, column=1, value='Applied Filters').font = sf; r += 1
        r = wh(ws, r, ['Filter', 'Value'])
        for k, v in applied.items():
            r = wr(ws, r, [k, v])
        r += 1
    total = qs.count()
    ws.cell(row=r, column=1, value='Summary').font = sf; r += 1
    r = wh(ws, r, ['Metric', 'Value'])
    r = wr(ws, r, ['Total Assessments', total])
    r = wr(ws, r, ['First Visits', qs.filter(assessment_type='first_visit').count()])
    r = wr(ws, r, ['Follow-ups', qs.filter(assessment_type='follow_up').count()])
    r = wr(ws, r, ['Follow-up Required', qs.filter(follow_up_required=True).count()])
    aw(ws, 2)

    dept_qs = qs.values('department').annotate(cnt=Count('id')).order_by('-cnt')
    add_pie(ws, 'By Department', [d['department'].title() for d in dept_qs], [d['cnt'] for d in dept_qs], 'E4')

    # Details
    ws2 = wb.create_sheet('Assessment Details')
    ws2.cell(row=1, column=1, value='Assessment Details').font = tf
    r = 3
    r = wh(ws2, r, ['Patient ID', 'Patient Name', 'Department', 'Type', 'Assessed By', 'Date', 'Diagnosis', 'Follow-up'])
    for a in list(qs.select_related('patient', 'assessed_by').order_by('-assessment_date')[:1000]):
        r = wr(ws2, r, [
            a.patient.patient_id if a.patient else '', a.patient.get_full_name() if a.patient else '',
            a.get_department_display() if hasattr(a, 'get_department_display') else a.department,
            a.get_assessment_type_display() if hasattr(a, 'get_assessment_type_display') else a.assessment_type,
            a.assessed_by.get_full_name() if a.assessed_by else 'N/A',
            a.assessment_date.strftime('%Y-%m-%d') if a.assessment_date else '',
            (a.diagnosis or '')[:80],
            'Yes' if a.follow_up_required else 'No',
        ])
    aw(ws2, 8)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Custom_{label}_Report_{date.today()}.xlsx"'
    return response


# ============================================================================
# Pharmacy Reports
# ============================================================================

@login_required
def pharmacy_reports(request):
    """Pharmacy reports with inventory, sales, and stock analytics."""
    from pharmacy.models import Medication, Batch, StockMovement, Prescription, Supplier, Category
    from django.db.models import Sum, Count, F, ExpressionWrapper, DecimalField

    period = request.GET.get('period', 'this_month')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    end_date = date.today()

    if period == 'custom' and date_from and date_to:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    elif period == 'last_month':
        start_date = (end_date.replace(day=1) - timedelta(days=1)).replace(day=1)
        end_date = end_date.replace(day=1) - timedelta(days=1)
    elif period == 'last_7_days':
        start_date = end_date - timedelta(days=7)
    elif period == 'last_30_days':
        start_date = end_date - timedelta(days=30)
    elif period == 'this_year':
        start_date = end_date.replace(month=1, day=1)
    else:
        start_date = end_date.replace(day=1)

    # KPI totals
    total_medications = Medication.objects.filter(is_active=True).count()
    total_batches = Batch.objects.filter(is_active=True, expiry_date__gt=timezone.now()).count()
    total_suppliers = Supplier.objects.filter(is_active=True).count()
    total_categories = Category.objects.count()

    # Inventory value
    inventory_value = Batch.objects.filter(
        is_active=True, expiry_date__gt=timezone.now()
    ).aggregate(
        total=Sum(ExpressionWrapper(
            F('quantity_remaining') * F('medication__unit_price'),
            output_field=DecimalField()
        ))
    )['total'] or 0

    # Low stock
    low_stock_items = Medication.objects.filter(is_active=True).annotate(
        total_stock=Sum('batches__quantity_remaining', filter=Q(batches__is_active=True))
    ).filter(
        Q(total_stock__lte=F('reorder_level')) | Q(total_stock__isnull=True)
    )
    low_stock_count = low_stock_items.count()

    # Expiring within 90 days
    expiry_threshold = timezone.now() + timedelta(days=90)
    expiring_soon = Batch.objects.filter(
        is_active=True, expiry_date__gt=timezone.now(),
        expiry_date__lte=expiry_threshold, quantity_remaining__gt=0
    ).select_related('medication').order_by('expiry_date')[:10]

    # Sales in period (StockMovement with SALE reference)
    sales_qs = StockMovement.objects.filter(
        movement_type='out', reference__icontains='SALE',
        created_at__date__gte=start_date, created_at__date__lte=end_date
    )
    total_sales = sales_qs.count()
    total_revenue = sales_qs.annotate(
        rev=F('quantity') * F('batch__selling_price')
    ).aggregate(total=Sum('rev'))['total'] or 0

    # Top selling medications
    top_selling = sales_qs.values('batch__medication__name').annotate(
        total_qty=Sum('quantity'),
        total_rev=Sum(F('quantity') * F('batch__selling_price')),
        sale_count=Count('id')
    ).order_by('-total_qty')[:10]

    # Sales trend (last 7 days)
    sales_trend_labels = []
    sales_trend_data = []
    for i in range(7):
        day = end_date - timedelta(days=6 - i)
        day_sales = StockMovement.objects.filter(
            movement_type='out', reference__icontains='SALE', created_at__date=day
        ).annotate(rev=F('quantity') * F('batch__selling_price')).aggregate(total=Sum('rev'))['total'] or 0
        sales_trend_labels.append(day.strftime('%m/%d'))
        sales_trend_data.append(float(day_sales))

    # Prescriptions in period
    prescriptions_total = Prescription.objects.filter(
        prescribed_date__date__gte=start_date, prescribed_date__date__lte=end_date
    ).count()
    prescriptions_pending = Prescription.objects.filter(status='pending').count()
    prescriptions_dispensed = Prescription.objects.filter(
        status='dispensed',
        prescribed_date__date__gte=start_date, prescribed_date__date__lte=end_date
    ).count()

    # Category breakdown
    category_data = Medication.objects.filter(is_active=True).values('category__name').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    cat_labels = [c['category__name'] or 'Uncategorized' for c in category_data]
    cat_counts = [c['count'] for c in category_data]

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'period': period,
        'total_medications': total_medications,
        'total_batches': total_batches,
        'total_suppliers': total_suppliers,
        'total_categories': total_categories,
        'inventory_value': inventory_value,
        'low_stock_count': low_stock_count,
        'low_stock_items': low_stock_items[:10],
        'expiring_soon': expiring_soon,
        'total_sales': total_sales,
        'total_revenue': total_revenue,
        'top_selling': top_selling,
        'sales_trend_labels': json.dumps(sales_trend_labels),
        'sales_trend_data': json.dumps(sales_trend_data),
        'prescriptions_total': prescriptions_total,
        'prescriptions_pending': prescriptions_pending,
        'prescriptions_dispensed': prescriptions_dispensed,
        'cat_labels': json.dumps(cat_labels),
        'cat_counts': json.dumps(cat_counts),
    }
    return render(request, 'reports/pharmacy_reports.html', context)


# ============================================================================
# Laboratory Reports
# ============================================================================

@login_required
def laboratory_reports(request):
    """Laboratory reports with test request analytics."""
    from laboratory.models import LabTest, LabTestRequest, LabTestResult

    period = request.GET.get('period', 'this_month')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    end_date = date.today()

    if period == 'custom' and date_from and date_to:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    elif period == 'last_month':
        start_date = (end_date.replace(day=1) - timedelta(days=1)).replace(day=1)
        end_date = end_date.replace(day=1) - timedelta(days=1)
    elif period == 'last_7_days':
        start_date = end_date - timedelta(days=7)
    elif period == 'last_30_days':
        start_date = end_date - timedelta(days=30)
    elif period == 'this_year':
        start_date = end_date.replace(month=1, day=1)
    else:
        start_date = end_date.replace(day=1)

    # KPI totals
    total_tests = LabTest.objects.filter(is_active=True).count()

    requests_qs = LabTestRequest.objects.filter(
        date_requested__date__gte=start_date,
        date_requested__date__lte=end_date
    )
    total_requests = requests_qs.count()
    completed_requests = requests_qs.filter(status='completed').count()
    pending_requests = requests_qs.filter(status__in=['requested', 'sample_collected', 'in_progress']).count()
    cancelled_requests = requests_qs.filter(status='cancelled').count()

    # Revenue from lab tests
    lab_revenue = requests_qs.filter(status='completed').aggregate(
        total=Sum('test__price')
    )['total'] or 0

    # Requests by status
    status_breakdown = requests_qs.values('status').annotate(count=Count('id')).order_by('-count')
    status_labels = [s['status'].replace('_', ' ').title() for s in status_breakdown]
    status_counts = [s['count'] for s in status_breakdown]

    # Requests by category
    category_breakdown = requests_qs.values('test__category__name').annotate(
        count=Count('id')
    ).order_by('-count')
    lab_cat_labels = [c['test__category__name'] or 'Other' for c in category_breakdown]
    lab_cat_counts = [c['count'] for c in category_breakdown]

    # Most requested tests
    popular_tests = requests_qs.values('test__name', 'test__code').annotate(
        count=Count('id'),
        revenue=Sum('test__price')
    ).order_by('-count')[:10]

    # Requests trend (last 7 days)
    trend_labels = []
    trend_data = []
    for i in range(7):
        day = end_date - timedelta(days=6 - i)
        day_count = LabTestRequest.objects.filter(date_requested__date=day).count()
        trend_labels.append(day.strftime('%m/%d'))
        trend_data.append(day_count)

    # Priority breakdown
    priority_breakdown = requests_qs.values('priority').annotate(count=Count('id'))
    priority_labels = [p['priority'].title() for p in priority_breakdown]
    priority_counts = [p['count'] for p in priority_breakdown]

    # Recent requests
    recent_requests = LabTestRequest.objects.select_related(
        'patient', 'test', 'requested_by'
    ).order_by('-date_requested')[:10]

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'period': period,
        'total_tests': total_tests,
        'total_requests': total_requests,
        'completed_requests': completed_requests,
        'pending_requests': pending_requests,
        'cancelled_requests': cancelled_requests,
        'lab_revenue': lab_revenue,
        'status_labels': json.dumps(status_labels),
        'status_counts': json.dumps(status_counts),
        'lab_cat_labels': json.dumps(lab_cat_labels),
        'lab_cat_counts': json.dumps(lab_cat_counts),
        'popular_tests': popular_tests,
        'trend_labels': json.dumps(trend_labels),
        'trend_data': json.dumps(trend_data),
        'priority_labels': json.dumps(priority_labels),
        'priority_counts': json.dumps(priority_counts),
        'recent_requests': recent_requests,
    }
    return render(request, 'reports/laboratory_reports.html', context)


# ============================================================================
# Budget & Expenses Reports
# ============================================================================

@login_required
def budget_expense_reports(request):
    """Budget and expense reports with spending analytics."""
    from budget.models import Budget as BudgetModel, Expense, ExpenseCategory, BudgetItem

    period = request.GET.get('period', 'this_month')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    end_date = date.today()

    if period == 'custom' and date_from and date_to:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    elif period == 'last_month':
        start_date = (end_date.replace(day=1) - timedelta(days=1)).replace(day=1)
        end_date = end_date.replace(day=1) - timedelta(days=1)
    elif period == 'last_7_days':
        start_date = end_date - timedelta(days=7)
    elif period == 'last_30_days':
        start_date = end_date - timedelta(days=30)
    elif period == 'this_year':
        start_date = end_date.replace(month=1, day=1)
    else:
        start_date = end_date.replace(day=1)

    # KPI totals
    total_budgets = BudgetModel.objects.filter(status='active').count()
    total_categories = ExpenseCategory.objects.filter(is_active=True).count()

    expenses_qs = Expense.objects.filter(
        expense_date__gte=start_date, expense_date__lte=end_date
    )
    total_expenses = expenses_qs.count()
    total_expense_amount = expenses_qs.filter(status__in=['approved', 'paid']).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')
    pending_expenses = expenses_qs.filter(status='pending').count()
    approved_expenses = expenses_qs.filter(status__in=['approved', 'paid']).count()

    # Active budgets summary
    active_budgets = BudgetModel.objects.filter(status='active').order_by('-start_date')[:5]
    budget_summaries = []
    for b in active_budgets:
        spent = b.get_spent_amount()
        util = b.get_utilization_percentage()
        budget_summaries.append({
            'name': b.name,
            'total': b.total_amount,
            'spent': spent,
            'remaining': b.total_amount - spent,
            'utilization': round(float(util), 1),
            'period': f"{b.start_date.strftime('%b %d')} - {b.end_date.strftime('%b %d, %Y')}",
        })

    # Expense by category
    category_breakdown = expenses_qs.filter(status__in=['approved', 'paid']).values(
        'category__name'
    ).annotate(total=Sum('amount'), count=Count('id')).order_by('-total')
    exp_cat_labels = [c['category__name'] for c in category_breakdown]
    exp_cat_amounts = [float(c['total']) for c in category_breakdown]

    # Expense by payment method
    method_breakdown = expenses_qs.filter(status__in=['approved', 'paid']).values(
        'payment_method'
    ).annotate(total=Sum('amount'), count=Count('id')).order_by('-total')
    method_labels = [m['payment_method'].replace('_', ' ').title() for m in method_breakdown]
    method_amounts = [float(m['total']) for m in method_breakdown]

    # Expense trend (last 7 days)
    exp_trend_labels = []
    exp_trend_data = []
    for i in range(7):
        day = end_date - timedelta(days=6 - i)
        day_total = Expense.objects.filter(
            expense_date=day, status__in=['approved', 'paid']
        ).aggregate(total=Sum('amount'))['total'] or 0
        exp_trend_labels.append(day.strftime('%m/%d'))
        exp_trend_data.append(float(day_total))

    # Recent expenses
    recent_expenses = Expense.objects.select_related(
        'category', 'submitted_by'
    ).order_by('-expense_date', '-created_at')[:10]

    # Top vendors
    top_vendors = expenses_qs.filter(
        status__in=['approved', 'paid']
    ).exclude(vendor_name='').values('vendor_name').annotate(
        total=Sum('amount'), count=Count('id')
    ).order_by('-total')[:10]

    context = {
        'start_date': start_date,
        'end_date': end_date,
        'period': period,
        'total_budgets': total_budgets,
        'total_categories': total_categories,
        'total_expenses': total_expenses,
        'total_expense_amount': total_expense_amount,
        'pending_expenses': pending_expenses,
        'approved_expenses': approved_expenses,
        'budget_summaries': budget_summaries,
        'exp_cat_labels': json.dumps(exp_cat_labels),
        'exp_cat_amounts': json.dumps(exp_cat_amounts),
        'method_labels': json.dumps(method_labels),
        'method_amounts': json.dumps(method_amounts),
        'exp_trend_labels': json.dumps(exp_trend_labels),
        'exp_trend_data': json.dumps(exp_trend_data),
        'recent_expenses': recent_expenses,
        'top_vendors': top_vendors,
    }
    return render(request, 'reports/budget_expense_reports.html', context)
