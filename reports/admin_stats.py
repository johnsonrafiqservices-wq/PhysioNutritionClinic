"""
Admin Statistics Views — comprehensive data aggregation for all clinic modules.
Provides dashboard overview + per-module detail pages with Chart.js data and CSV/Excel export.
"""
import csv
import io
import json
from datetime import date, timedelta, datetime
from decimal import Decimal

from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Count, Sum, Avg, Q, F, Min, Max
from django.db.models.functions import TruncMonth, TruncWeek, TruncDate, ExtractHour
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone

# ──────────────────────────────────────────────
#  Helpers
# ──────────────────────────────────────────────

def _parse_dates(request):
    """Parse start_date / end_date from GET, default to current month."""
    fmt = '%Y-%m-%d'
    try:
        start = datetime.strptime(request.GET.get('start_date', ''), fmt).date()
    except (ValueError, TypeError):
        start = date.today().replace(day=1)
    try:
        end = datetime.strptime(request.GET.get('end_date', ''), fmt).date()
    except (ValueError, TypeError):
        end = date.today()
    return start, end


def _prev_period(start, end):
    """Return equivalent previous period dates."""
    delta = (end - start).days
    return start - timedelta(days=delta + 1), start - timedelta(days=1)


def _pct(part, total):
    if total == 0:
        return 0
    return round(part / total * 100, 1)


def _change_pct(current, previous):
    if previous == 0:
        return 100 if current > 0 else 0
    return round((current - previous) / previous * 100, 1)


def _json(obj):
    """Serialize for template injection."""
    return json.dumps(obj, default=str)


# ──────────────────────────────────────────────
#  Dashboard Overview
# ──────────────────────────────────────────────

@staff_member_required
def stats_dashboard(request):
    from patients.models import Patient
    from appointments.models import Appointment, Service
    from billing.models import Invoice, Payment
    from pharmacy.models import Medication, Batch, Prescription
    from laboratory.models import LabTestRequest, LabTestResult
    from staff_management.models import Staff, Attendance
    from budget.models import Expense, Budget

    start, end = _parse_dates(request)
    prev_start, prev_end = _prev_period(start, end)

    # ── Patients ──
    total_patients = Patient.objects.filter(is_active=True).count()
    new_patients = Patient.objects.filter(registration_date__range=[start, end]).count()
    prev_new_patients = Patient.objects.filter(registration_date__range=[prev_start, prev_end]).count()
    patients_change = _change_pct(new_patients, prev_new_patients)

    # Monthly patient registration trend (last 12 months)
    twelve_months_ago = date.today() - timedelta(days=365)
    patient_trend = list(
        Patient.objects.filter(registration_date__gte=twelve_months_ago)
        .annotate(month=TruncMonth('registration_date'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    patient_trend_labels = [d['month'].strftime('%b %Y') for d in patient_trend]
    patient_trend_data = [d['count'] for d in patient_trend]

    # Gender distribution
    gender_dist = list(Patient.objects.filter(is_active=True).values('gender').annotate(count=Count('id')))
    gender_labels = [g['gender'] or 'Unknown' for g in gender_dist]
    gender_data = [g['count'] for g in gender_dist]

    # ── Appointments ──
    total_appts = Appointment.objects.filter(appointment_date__range=[start, end]).count()
    completed_appts = Appointment.objects.filter(appointment_date__range=[start, end], status='completed').count()
    cancelled_appts = Appointment.objects.filter(appointment_date__range=[start, end], status='cancelled').count()
    no_show_appts = Appointment.objects.filter(appointment_date__range=[start, end], status='no_show').count()
    prev_total_appts = Appointment.objects.filter(appointment_date__range=[prev_start, prev_end]).count()
    appts_change = _change_pct(total_appts, prev_total_appts)
    completion_rate = _pct(completed_appts, total_appts)

    appt_status_labels = ['Completed', 'Cancelled', 'No Show', 'Other']
    other_appts = total_appts - completed_appts - cancelled_appts - no_show_appts
    appt_status_data = [completed_appts, cancelled_appts, no_show_appts, other_appts]

    # Daily appointments trend (last 30 days)
    thirty_days_ago = date.today() - timedelta(days=30)
    daily_appts = list(
        Appointment.objects.filter(appointment_date__gte=thirty_days_ago)
        .annotate(day=TruncDate('appointment_date'))
        .values('day')
        .annotate(count=Count('id'))
        .order_by('day')
    )
    daily_appt_labels = [d['day'].strftime('%m/%d') for d in daily_appts]
    daily_appt_data = [d['count'] for d in daily_appts]

    # Top services
    top_services = list(
        Service.objects.annotate(appt_count=Count('appointment'))
        .filter(appt_count__gt=0)
        .order_by('-appt_count')[:8]
        .values('name', 'appt_count')
    )
    svc_labels = [s['name'] for s in top_services]
    svc_data = [s['appt_count'] for s in top_services]

    # ── Billing / Revenue ──
    total_revenue = Payment.objects.filter(
        payment_date__date__range=[start, end], status='completed'
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0')
    prev_revenue = Payment.objects.filter(
        payment_date__date__range=[prev_start, prev_end], status='completed'
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0')
    revenue_change = _change_pct(float(total_revenue), float(prev_revenue))

    total_invoices = Invoice.objects.filter(created_at__date__range=[start, end]).count()
    outstanding = Invoice.objects.filter(status__in=['sent', 'overdue']).aggregate(s=Sum('total_amount'))['s'] or 0
    paid_invoices = Invoice.objects.filter(created_at__date__range=[start, end], status='paid').count()
    collection_rate = _pct(paid_invoices, total_invoices)

    # Revenue trend (last 12 months)
    revenue_trend = list(
        Payment.objects.filter(status='completed', payment_date__date__gte=twelve_months_ago)
        .annotate(month=TruncMonth('payment_date'))
        .values('month')
        .annotate(total=Sum('amount'))
        .order_by('month')
    )
    rev_trend_labels = [d['month'].strftime('%b %Y') for d in revenue_trend]
    rev_trend_data = [float(d['total']) for d in revenue_trend]

    # Payment methods
    pay_methods = list(
        Payment.objects.filter(status='completed', payment_date__date__range=[start, end])
        .values('payment_method')
        .annotate(total=Sum('amount'))
        .order_by('-total')
    )
    pm_labels = [p['payment_method'].replace('_', ' ').title() for p in pay_methods]
    pm_data = [float(p['total']) for p in pay_methods]

    # ── Pharmacy ──
    total_meds = Medication.objects.count()
    low_stock = Medication.objects.filter(current_stock__lte=F('reorder_level')).count()
    expired_batches = Batch.objects.filter(expiry_date__lt=date.today(), status='active').count()
    expiring_soon = Batch.objects.filter(
        expiry_date__range=[date.today(), date.today() + timedelta(days=90)], status='active'
    ).count()
    prescriptions_count = Prescription.objects.filter(date_prescribed__range=[start, end]).count()

    # ── Laboratory ──
    lab_requests = LabTestRequest.objects.filter(requested_date__date__range=[start, end]).count()
    lab_completed = LabTestRequest.objects.filter(requested_date__date__range=[start, end], status='completed').count()
    lab_pending = LabTestRequest.objects.filter(requested_date__date__range=[start, end], status__in=['requested', 'in_progress']).count()
    lab_completion_rate = _pct(lab_completed, lab_requests)

    # ── Staff ──
    total_staff = Staff.objects.filter(is_active=True).count()
    attendance_today = Attendance.objects.filter(date=date.today(), status='present').count()

    # ── Budget ──
    total_expenses = Expense.objects.filter(
        date__range=[start, end], status='approved'
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0')
    prev_expenses = Expense.objects.filter(
        date__range=[prev_start, prev_end], status='approved'
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0')
    expenses_change = _change_pct(float(total_expenses), float(prev_expenses))
    net_income = float(total_revenue) - float(total_expenses)

    # Expense categories
    exp_cats = list(
        Expense.objects.filter(date__range=[start, end], status='approved')
        .values('category__name')
        .annotate(total=Sum('amount'))
        .order_by('-total')[:8]
    )
    exp_cat_labels = [e['category__name'] for e in exp_cats]
    exp_cat_data = [float(e['total']) for e in exp_cats]

    # Revenue vs Expenses monthly
    exp_trend = list(
        Expense.objects.filter(status='approved', date__gte=twelve_months_ago)
        .annotate(month=TruncMonth('date'))
        .values('month')
        .annotate(total=Sum('amount'))
        .order_by('month')
    )
    # Align with revenue trend months
    exp_by_month = {d['month'].strftime('%b %Y'): float(d['total']) for d in exp_trend}
    rev_vs_exp_labels = rev_trend_labels[:]
    rev_vs_exp_rev = rev_trend_data[:]
    rev_vs_exp_exp = [exp_by_month.get(l, 0) for l in rev_vs_exp_labels]

    context = {
        'title': 'Statistics Dashboard',
        'start_date': start, 'end_date': end,
        # Patients
        'total_patients': total_patients, 'new_patients': new_patients, 'patients_change': patients_change,
        'patient_trend_labels': _json(patient_trend_labels), 'patient_trend_data': _json(patient_trend_data),
        'gender_labels': _json(gender_labels), 'gender_data': _json(gender_data),
        # Appointments
        'total_appts': total_appts, 'completed_appts': completed_appts, 'cancelled_appts': cancelled_appts,
        'no_show_appts': no_show_appts, 'appts_change': appts_change, 'completion_rate': completion_rate,
        'appt_status_labels': _json(appt_status_labels), 'appt_status_data': _json(appt_status_data),
        'daily_appt_labels': _json(daily_appt_labels), 'daily_appt_data': _json(daily_appt_data),
        'svc_labels': _json(svc_labels), 'svc_data': _json(svc_data),
        # Revenue
        'total_revenue': total_revenue, 'prev_revenue': prev_revenue, 'revenue_change': revenue_change,
        'total_invoices': total_invoices, 'outstanding': outstanding, 'collection_rate': collection_rate,
        'rev_trend_labels': _json(rev_trend_labels), 'rev_trend_data': _json(rev_trend_data),
        'pm_labels': _json(pm_labels), 'pm_data': _json(pm_data),
        # Revenue vs Expenses
        'rev_vs_exp_labels': _json(rev_vs_exp_labels),
        'rev_vs_exp_rev': _json(rev_vs_exp_rev),
        'rev_vs_exp_exp': _json(rev_vs_exp_exp),
        'net_income': net_income,
        # Pharmacy
        'total_meds': total_meds, 'low_stock': low_stock, 'expired_batches': expired_batches,
        'expiring_soon': expiring_soon, 'prescriptions_count': prescriptions_count,
        # Lab
        'lab_requests': lab_requests, 'lab_completed': lab_completed, 'lab_pending': lab_pending,
        'lab_completion_rate': lab_completion_rate,
        # Staff
        'total_staff': total_staff, 'attendance_today': attendance_today,
        # Budget
        'total_expenses': total_expenses, 'expenses_change': expenses_change,
        'exp_cat_labels': _json(exp_cat_labels), 'exp_cat_data': _json(exp_cat_data),
    }
    return render(request, 'admin/stats/dashboard.html', context)


# ──────────────────────────────────────────────
#  Patient Statistics Detail
# ──────────────────────────────────────────────

@staff_member_required
def stats_patients(request):
    from patients.models import Patient, Assessment, Triage

    start, end = _parse_dates(request)
    prev_start, prev_end = _prev_period(start, end)

    total = Patient.objects.filter(is_active=True).count()
    new = Patient.objects.filter(registration_date__range=[start, end]).count()
    prev_new = Patient.objects.filter(registration_date__range=[prev_start, prev_end]).count()
    change = _change_pct(new, prev_new)

    # Gender
    gender = list(Patient.objects.filter(is_active=True).values('gender').annotate(c=Count('id')))
    gender_labels = [g['gender'] or 'Unknown' for g in gender]
    gender_data = [g['c'] for g in gender]

    # Monthly registration trend (12 months)
    twelve = date.today() - timedelta(days=365)
    trend = list(
        Patient.objects.filter(registration_date__gte=twelve)
        .annotate(month=TruncMonth('registration_date')).values('month')
        .annotate(c=Count('id')).order_by('month')
    )
    trend_labels = [d['month'].strftime('%b %Y') for d in trend]
    trend_data = [d['c'] for d in trend]

    # Insurance distribution
    insurance = list(
        Patient.objects.filter(is_active=True)
        .values('insurance_provider').annotate(c=Count('id')).order_by('-c')[:10]
    )
    ins_labels = [i['insurance_provider'] or 'None' for i in insurance]
    ins_data = [i['c'] for i in insurance]

    # Age distribution
    patients_qs = Patient.objects.filter(is_active=True)
    age_buckets = {'0-17': 0, '18-30': 0, '31-45': 0, '46-60': 0, '61+': 0}
    for p in patients_qs:
        try:
            age = p.get_age() if hasattr(p, 'get_age') else None
            if age is None:
                continue
            age = int(age)
            if age <= 17:
                age_buckets['0-17'] += 1
            elif age <= 30:
                age_buckets['18-30'] += 1
            elif age <= 45:
                age_buckets['31-45'] += 1
            elif age <= 60:
                age_buckets['46-60'] += 1
            else:
                age_buckets['61+'] += 1
        except (ValueError, TypeError, AttributeError):
            continue
    age_labels = list(age_buckets.keys())
    age_data = list(age_buckets.values())

    # Assessments
    total_assessments = Assessment.objects.filter(assessment_date__range=[start, end]).count()
    by_dept = list(
        Assessment.objects.filter(assessment_date__range=[start, end])
        .values('department').annotate(c=Count('id')).order_by('-c')
    )
    dept_labels = [d['department'].title() for d in by_dept]
    dept_data = [d['c'] for d in by_dept]

    # Triage priority distribution
    triage_data_qs = list(
        Triage.objects.filter(triage_date__date__range=[start, end])
        .values('priority').annotate(c=Count('id')).order_by('priority')
    )
    triage_labels = [t['priority'] for t in triage_data_qs]
    triage_data = [t['c'] for t in triage_data_qs]

    # Recent patients
    recent_patients = Patient.objects.filter(is_active=True).order_by('-registration_date')[:20]

    context = {
        'title': 'Patient Statistics',
        'start_date': start, 'end_date': end,
        'total': total, 'new': new, 'change': change,
        'gender_labels': _json(gender_labels), 'gender_data': _json(gender_data),
        'trend_labels': _json(trend_labels), 'trend_data': _json(trend_data),
        'ins_labels': _json(ins_labels), 'ins_data': _json(ins_data),
        'age_labels': _json(age_labels), 'age_data': _json(age_data),
        'total_assessments': total_assessments,
        'dept_labels': _json(dept_labels), 'dept_data': _json(dept_data),
        'triage_labels': _json(triage_labels), 'triage_data': _json(triage_data),
        'recent_patients': recent_patients,
    }
    return render(request, 'admin/stats/patients.html', context)


# ──────────────────────────────────────────────
#  Appointment Statistics Detail
# ──────────────────────────────────────────────

@staff_member_required
def stats_appointments(request):
    from appointments.models import Appointment, Service

    start, end = _parse_dates(request)
    prev_start, prev_end = _prev_period(start, end)

    qs = Appointment.objects.filter(appointment_date__range=[start, end])
    total = qs.count()
    completed = qs.filter(status='completed').count()
    cancelled = qs.filter(status='cancelled').count()
    no_show = qs.filter(status='no_show').count()
    scheduled = qs.filter(status='scheduled').count()
    prev_total = Appointment.objects.filter(appointment_date__range=[prev_start, prev_end]).count()
    change = _change_pct(total, prev_total)
    comp_rate = _pct(completed, total)
    cancel_rate = _pct(cancelled, total)

    # Status breakdown
    status_labels = ['Completed', 'Scheduled', 'Cancelled', 'No Show']
    status_data = [completed, scheduled, cancelled, no_show]

    # Daily trend (30 days)
    thirty = date.today() - timedelta(days=30)
    daily = list(
        Appointment.objects.filter(appointment_date__gte=thirty)
        .annotate(day=TruncDate('appointment_date')).values('day')
        .annotate(c=Count('id')).order_by('day')
    )
    daily_labels = [d['day'].strftime('%m/%d') for d in daily]
    daily_data = [d['c'] for d in daily]

    # Monthly trend (12 months)
    twelve = date.today() - timedelta(days=365)
    monthly = list(
        Appointment.objects.filter(appointment_date__gte=twelve)
        .annotate(month=TruncMonth('appointment_date')).values('month')
        .annotate(c=Count('id')).order_by('month')
    )
    monthly_labels = [d['month'].strftime('%b %Y') for d in monthly]
    monthly_data = [d['c'] for d in monthly]

    # By service
    by_svc = list(
        qs.values('service__name').annotate(c=Count('id')).order_by('-c')[:10]
    )
    svc_labels = [s['service__name'] or 'N/A' for s in by_svc]
    svc_data = [s['c'] for s in by_svc]

    # By provider
    by_provider = list(
        qs.values('provider__first_name', 'provider__last_name')
        .annotate(total=Count('id'), comp=Count('id', filter=Q(status='completed')))
        .order_by('-total')[:10]
    )

    # By hour of day
    by_hour = list(
        qs.annotate(hour=ExtractHour('appointment_time'))
        .values('hour').annotate(c=Count('id')).order_by('hour')
    )
    hour_labels = [f"{h['hour']}:00" for h in by_hour]
    hour_data = [h['c'] for h in by_hour]

    # By day of week
    from django.db.models.functions import ExtractWeekDay
    by_dow = list(
        qs.annotate(dow=ExtractWeekDay('appointment_date'))
        .values('dow').annotate(c=Count('id')).order_by('dow')
    )
    dow_map = {1: 'Sun', 2: 'Mon', 3: 'Tue', 4: 'Wed', 5: 'Thu', 6: 'Fri', 7: 'Sat'}
    dow_labels = [dow_map.get(d['dow'], '?') for d in by_dow]
    dow_data = [d['c'] for d in by_dow]

    context = {
        'title': 'Appointment Statistics',
        'start_date': start, 'end_date': end,
        'total': total, 'completed': completed, 'cancelled': cancelled,
        'no_show': no_show, 'scheduled': scheduled,
        'change': change, 'comp_rate': comp_rate, 'cancel_rate': cancel_rate,
        'status_labels': _json(status_labels), 'status_data': _json(status_data),
        'daily_labels': _json(daily_labels), 'daily_data': _json(daily_data),
        'monthly_labels': _json(monthly_labels), 'monthly_data': _json(monthly_data),
        'svc_labels': _json(svc_labels), 'svc_data': _json(svc_data),
        'by_provider': by_provider,
        'hour_labels': _json(hour_labels), 'hour_data': _json(hour_data),
        'dow_labels': _json(dow_labels), 'dow_data': _json(dow_data),
    }
    return render(request, 'admin/stats/appointments.html', context)


# ──────────────────────────────────────────────
#  Billing / Revenue Statistics Detail
# ──────────────────────────────────────────────

@staff_member_required
def stats_billing(request):
    from billing.models import Invoice, Payment, InsuranceClaim

    start, end = _parse_dates(request)
    prev_start, prev_end = _prev_period(start, end)

    # Revenue
    revenue = Payment.objects.filter(payment_date__date__range=[start, end], status='completed').aggregate(s=Sum('amount'))['s'] or 0
    prev_rev = Payment.objects.filter(payment_date__date__range=[prev_start, prev_end], status='completed').aggregate(s=Sum('amount'))['s'] or 0
    rev_change = _change_pct(float(revenue), float(prev_rev))

    # Invoices
    invoices_created = Invoice.objects.filter(created_at__date__range=[start, end]).count()
    invoices_paid = Invoice.objects.filter(created_at__date__range=[start, end], status='paid').count()
    invoices_overdue = Invoice.objects.filter(status='overdue').count()
    outstanding = Invoice.objects.filter(status__in=['sent', 'overdue']).aggregate(s=Sum('total_amount'))['s'] or 0
    avg_invoice = Invoice.objects.filter(created_at__date__range=[start, end]).aggregate(a=Avg('total_amount'))['a'] or 0
    collection_rate = _pct(invoices_paid, invoices_created)

    # Monthly revenue trend
    twelve = date.today() - timedelta(days=365)
    rev_trend = list(
        Payment.objects.filter(status='completed', payment_date__date__gte=twelve)
        .annotate(month=TruncMonth('payment_date')).values('month')
        .annotate(total=Sum('amount')).order_by('month')
    )
    rev_labels = [d['month'].strftime('%b %Y') for d in rev_trend]
    rev_data = [float(d['total']) for d in rev_trend]

    # Daily revenue (30 days)
    thirty = date.today() - timedelta(days=30)
    daily_rev = list(
        Payment.objects.filter(status='completed', payment_date__date__gte=thirty)
        .annotate(day=TruncDate('payment_date')).values('day')
        .annotate(total=Sum('amount')).order_by('day')
    )
    daily_rev_labels = [d['day'].strftime('%m/%d') for d in daily_rev]
    daily_rev_data = [float(d['total']) for d in daily_rev]

    # Payment methods
    methods = list(
        Payment.objects.filter(status='completed', payment_date__date__range=[start, end])
        .values('payment_method').annotate(total=Sum('amount'), count=Count('id')).order_by('-total')
    )
    pm_labels = [m['payment_method'].replace('_', ' ').title() for m in methods]
    pm_data = [float(m['total']) for m in methods]
    pm_counts = [m['count'] for m in methods]

    # Invoice status
    inv_statuses = list(
        Invoice.objects.filter(created_at__date__range=[start, end])
        .values('status').annotate(c=Count('id')).order_by('-c')
    )
    inv_st_labels = [s['status'].title() for s in inv_statuses]
    inv_st_data = [s['c'] for s in inv_statuses]

    # Outstanding aging
    today = date.today()
    o30 = Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__gte=today - timedelta(days=30), due_date__lt=today).aggregate(s=Sum('total_amount'))['s'] or 0
    o60 = Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__gte=today - timedelta(days=60), due_date__lt=today - timedelta(days=30)).aggregate(s=Sum('total_amount'))['s'] or 0
    o90 = Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__gte=today - timedelta(days=90), due_date__lt=today - timedelta(days=60)).aggregate(s=Sum('total_amount'))['s'] or 0
    o90p = Invoice.objects.filter(status__in=['sent', 'overdue'], due_date__lt=today - timedelta(days=90)).aggregate(s=Sum('total_amount'))['s'] or 0
    aging_labels = ['0-30 days', '31-60 days', '61-90 days', '90+ days']
    aging_data = [float(o30), float(o60), float(o90), float(o90p)]

    # Insurance claims
    claims_total = InsuranceClaim.objects.filter(submission_date__range=[start, end]).count()
    claims_approved = InsuranceClaim.objects.filter(submission_date__range=[start, end], status='approved').count()
    claims_pending = InsuranceClaim.objects.filter(submission_date__range=[start, end], status__in=['submitted', 'pending']).count()
    claims_denied = InsuranceClaim.objects.filter(submission_date__range=[start, end], status='denied').count()

    # Recent payments
    recent_payments = Payment.objects.filter(status='completed').select_related('invoice__patient').order_by('-payment_date')[:15]

    context = {
        'title': 'Billing & Revenue Statistics',
        'start_date': start, 'end_date': end,
        'revenue': revenue, 'prev_rev': prev_rev, 'rev_change': rev_change,
        'invoices_created': invoices_created, 'invoices_paid': invoices_paid,
        'invoices_overdue': invoices_overdue, 'outstanding': outstanding,
        'avg_invoice': avg_invoice, 'collection_rate': collection_rate,
        'rev_labels': _json(rev_labels), 'rev_data': _json(rev_data),
        'daily_rev_labels': _json(daily_rev_labels), 'daily_rev_data': _json(daily_rev_data),
        'pm_labels': _json(pm_labels), 'pm_data': _json(pm_data), 'pm_counts': pm_counts,
        'inv_st_labels': _json(inv_st_labels), 'inv_st_data': _json(inv_st_data),
        'aging_labels': _json(aging_labels), 'aging_data': _json(aging_data),
        'claims_total': claims_total, 'claims_approved': claims_approved,
        'claims_pending': claims_pending, 'claims_denied': claims_denied,
        'recent_payments': recent_payments,
    }
    return render(request, 'admin/stats/billing.html', context)


# ──────────────────────────────────────────────
#  Pharmacy Statistics Detail
# ──────────────────────────────────────────────

@staff_member_required
def stats_pharmacy(request):
    from pharmacy.models import Medication, Batch, Prescription, PurchaseOrder, StockMovement, Category

    start, end = _parse_dates(request)

    total_meds = Medication.objects.count()
    active_meds = Medication.objects.filter(is_active=True).count()
    low_stock = Medication.objects.filter(current_stock__lte=F('reorder_level')).count()
    out_of_stock = Medication.objects.filter(current_stock=0).count()

    # Batch info
    total_batches = Batch.objects.filter(status='active').count()
    expired = Batch.objects.filter(expiry_date__lt=date.today(), status='active').count()
    expiring_30 = Batch.objects.filter(expiry_date__range=[date.today(), date.today() + timedelta(days=30)], status='active').count()
    expiring_90 = Batch.objects.filter(expiry_date__range=[date.today(), date.today() + timedelta(days=90)], status='active').count()

    # Stock value
    stock_cost = sum(
        float(m.cost_price or 0) * m.current_stock
        for m in Medication.objects.filter(is_active=True)
    )
    stock_sell = sum(
        float(m.selling_price or 0) * m.current_stock
        for m in Medication.objects.filter(is_active=True)
    )

    # Prescriptions
    prescriptions = Prescription.objects.filter(date_prescribed__range=[start, end]).count()
    dispensed = Prescription.objects.filter(date_prescribed__range=[start, end], status='dispensed').count()
    pending_rx = Prescription.objects.filter(date_prescribed__range=[start, end], status='pending').count()

    # Purchase orders
    po_count = PurchaseOrder.objects.filter(order_date__range=[start, end]).count()
    po_total = PurchaseOrder.objects.filter(order_date__range=[start, end]).aggregate(s=Sum('total_amount'))['s'] or 0

    # Stock movements
    stock_in = StockMovement.objects.filter(date__date__range=[start, end], movement_type='in').aggregate(s=Sum('quantity'))['s'] or 0
    stock_out = StockMovement.objects.filter(date__date__range=[start, end], movement_type='out').aggregate(s=Sum('quantity'))['s'] or 0

    # By category
    by_cat = list(
        Medication.objects.filter(is_active=True).values('category__name')
        .annotate(c=Count('id')).order_by('-c')
    )
    cat_labels = [c['category__name'] or 'Uncategorized' for c in by_cat]
    cat_data = [c['c'] for c in by_cat]

    # Stock movement trend (monthly)
    twelve = date.today() - timedelta(days=365)
    move_trend = list(
        StockMovement.objects.filter(date__date__gte=twelve)
        .annotate(month=TruncMonth('date')).values('month', 'movement_type')
        .annotate(total=Sum('quantity')).order_by('month')
    )
    move_months = sorted(set(d['month'].strftime('%b %Y') for d in move_trend))
    move_in = {d['month'].strftime('%b %Y'): d['total'] for d in move_trend if d['movement_type'] == 'in'}
    move_out = {d['month'].strftime('%b %Y'): d['total'] for d in move_trend if d['movement_type'] == 'out'}
    move_in_data = [move_in.get(m, 0) for m in move_months]
    move_out_data = [move_out.get(m, 0) for m in move_months]

    # Low stock items list
    low_stock_items = list(
        Medication.objects.filter(current_stock__lte=F('reorder_level'), is_active=True)
        .values('name', 'current_stock', 'reorder_level')[:20]
    )

    # Expiring items
    expiring_items = list(
        Batch.objects.filter(expiry_date__range=[date.today(), date.today() + timedelta(days=90)], status='active')
        .select_related('medication')
        .order_by('expiry_date')[:20]
    )

    context = {
        'title': 'Pharmacy Statistics',
        'start_date': start, 'end_date': end,
        'total_meds': total_meds, 'active_meds': active_meds,
        'low_stock': low_stock, 'out_of_stock': out_of_stock,
        'total_batches': total_batches, 'expired': expired,
        'expiring_30': expiring_30, 'expiring_90': expiring_90,
        'stock_cost': stock_cost, 'stock_sell': stock_sell,
        'prescriptions': prescriptions, 'dispensed': dispensed, 'pending_rx': pending_rx,
        'po_count': po_count, 'po_total': po_total,
        'stock_in': stock_in, 'stock_out': stock_out,
        'cat_labels': _json(cat_labels), 'cat_data': _json(cat_data),
        'move_months': _json(move_months), 'move_in_data': _json(move_in_data), 'move_out_data': _json(move_out_data),
        'low_stock_items': low_stock_items, 'expiring_items': expiring_items,
    }
    return render(request, 'admin/stats/pharmacy.html', context)


# ──────────────────────────────────────────────
#  Laboratory Statistics Detail
# ──────────────────────────────────────────────

@staff_member_required
def stats_laboratory(request):
    from laboratory.models import LabTestRequest, LabTestResult, LabTest, TestCategory

    start, end = _parse_dates(request)
    prev_start, prev_end = _prev_period(start, end)

    qs = LabTestRequest.objects.filter(requested_date__date__range=[start, end])
    total = qs.count()
    completed = qs.filter(status='completed').count()
    pending = qs.filter(status__in=['requested', 'in_progress']).count()
    cancelled = qs.filter(status='cancelled').count()
    prev_total = LabTestRequest.objects.filter(requested_date__date__range=[prev_start, prev_end]).count()
    change = _change_pct(total, prev_total)
    comp_rate = _pct(completed, total)

    # Results
    results_qs = LabTestResult.objects.filter(result_date__date__range=[start, end])
    total_results = results_qs.count()
    abnormal_results = results_qs.filter(is_abnormal=True).count()
    abnormal_rate = _pct(abnormal_results, total_results)

    # Status breakdown
    status_labels = ['Completed', 'Pending', 'Cancelled']
    status_data = [completed, pending, cancelled]

    # Monthly trend
    twelve = date.today() - timedelta(days=365)
    monthly = list(
        LabTestRequest.objects.filter(requested_date__date__gte=twelve)
        .annotate(month=TruncMonth('requested_date')).values('month')
        .annotate(c=Count('id')).order_by('month')
    )
    monthly_labels = [d['month'].strftime('%b %Y') for d in monthly]
    monthly_data = [d['c'] for d in monthly]

    # Top tests
    top_tests = list(
        qs.values('test__name').annotate(c=Count('id')).order_by('-c')[:10]
    )
    test_labels = [t['test__name'] for t in top_tests]
    test_data = [t['c'] for t in top_tests]

    # By priority
    by_priority = list(
        qs.values('priority').annotate(c=Count('id')).order_by('-c')
    )
    pri_labels = [p['priority'].title() for p in by_priority]
    pri_data = [p['c'] for p in by_priority]

    # By category
    by_cat = list(
        qs.values('test__category__name').annotate(c=Count('id')).order_by('-c')
    )
    cat_labels = [c['test__category__name'] or 'Uncategorized' for c in by_cat]
    cat_data = [c['c'] for c in by_cat]

    # Turnaround time (average hours from request to completion)
    completed_requests = qs.filter(status='completed', completed_date__isnull=False)
    avg_tat = None
    if completed_requests.exists():
        from django.db.models import ExpressionWrapper, DurationField
        tats = []
        for req in completed_requests[:100]:
            if req.completed_date and req.requested_date:
                delta = req.completed_date - req.requested_date
                tats.append(delta.total_seconds() / 3600)
        avg_tat = round(sum(tats) / len(tats), 1) if tats else None

    context = {
        'title': 'Laboratory Statistics',
        'start_date': start, 'end_date': end,
        'total': total, 'completed': completed, 'pending': pending, 'cancelled': cancelled,
        'change': change, 'comp_rate': comp_rate,
        'total_results': total_results, 'abnormal_results': abnormal_results, 'abnormal_rate': abnormal_rate,
        'status_labels': _json(status_labels), 'status_data': _json(status_data),
        'monthly_labels': _json(monthly_labels), 'monthly_data': _json(monthly_data),
        'test_labels': _json(test_labels), 'test_data': _json(test_data),
        'pri_labels': _json(pri_labels), 'pri_data': _json(pri_data),
        'cat_labels': _json(cat_labels), 'cat_data': _json(cat_data),
        'avg_tat': avg_tat,
    }
    return render(request, 'admin/stats/laboratory.html', context)


# ──────────────────────────────────────────────
#  Staff Statistics Detail
# ──────────────────────────────────────────────

@staff_member_required
def stats_staff(request):
    from staff_management.models import Staff, Department, Attendance, LeaveRequest, DutyRoster

    start, end = _parse_dates(request)

    total_staff = Staff.objects.filter(is_active=True).count()
    by_status = list(
        Staff.objects.filter(is_active=True).values('employment_status').annotate(c=Count('id')).order_by('-c')
    )
    status_labels = [s['employment_status'].replace('_', ' ').title() for s in by_status]
    status_data = [s['c'] for s in by_status]

    # By department
    by_dept = list(
        Staff.objects.filter(is_active=True).values('department__name').annotate(c=Count('id')).order_by('-c')
    )
    dept_labels = [d['department__name'] or 'Unassigned' for d in by_dept]
    dept_data = [d['c'] for d in by_dept]

    # Attendance
    att_qs = Attendance.objects.filter(date__range=[start, end])
    total_att = att_qs.count()
    present = att_qs.filter(status='present').count()
    absent = att_qs.filter(status='absent').count()
    late = att_qs.filter(status='late').count()
    att_rate = _pct(present, total_att)

    att_labels = ['Present', 'Absent', 'Late']
    att_data = [present, absent, late]

    # Monthly attendance trend
    twelve = date.today() - timedelta(days=365)
    att_trend = list(
        Attendance.objects.filter(date__gte=twelve)
        .annotate(month=TruncMonth('date')).values('month')
        .annotate(
            present=Count('id', filter=Q(status='present')),
            total=Count('id')
        ).order_by('month')
    )
    att_trend_labels = [d['month'].strftime('%b %Y') for d in att_trend]
    att_trend_present = [d['present'] for d in att_trend]
    att_trend_total = [d['total'] for d in att_trend]

    # Leave requests
    leave_total = LeaveRequest.objects.filter(start_date__range=[start, end]).count()
    leave_approved = LeaveRequest.objects.filter(start_date__range=[start, end], status='approved').count()
    leave_pending = LeaveRequest.objects.filter(start_date__range=[start, end], status='pending').count()
    leave_rejected = LeaveRequest.objects.filter(start_date__range=[start, end], status='rejected').count()

    by_leave_type = list(
        LeaveRequest.objects.filter(start_date__range=[start, end])
        .values('leave_type').annotate(c=Count('id')).order_by('-c')
    )
    leave_type_labels = [l['leave_type'].replace('_', ' ').title() for l in by_leave_type]
    leave_type_data = [l['c'] for l in by_leave_type]

    context = {
        'title': 'Staff Statistics',
        'start_date': start, 'end_date': end,
        'total_staff': total_staff,
        'status_labels': _json(status_labels), 'status_data': _json(status_data),
        'dept_labels': _json(dept_labels), 'dept_data': _json(dept_data),
        'total_att': total_att, 'present': present, 'absent': absent, 'late': late, 'att_rate': att_rate,
        'att_labels': _json(att_labels), 'att_data': _json(att_data),
        'att_trend_labels': _json(att_trend_labels),
        'att_trend_present': _json(att_trend_present), 'att_trend_total': _json(att_trend_total),
        'leave_total': leave_total, 'leave_approved': leave_approved,
        'leave_pending': leave_pending, 'leave_rejected': leave_rejected,
        'leave_type_labels': _json(leave_type_labels), 'leave_type_data': _json(leave_type_data),
    }
    return render(request, 'admin/stats/staff.html', context)


# ──────────────────────────────────────────────
#  Export Views
# ──────────────────────────────────────────────

@staff_member_required
def export_stats_csv(request, module):
    """Export statistics for a module as CSV."""
    start, end = _parse_dates(request)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{module}_statistics_{start}_{end}.csv"'
    writer = csv.writer(response)

    if module == 'patients':
        from patients.models import Patient
        writer.writerow(['Patient ID', 'First Name', 'Last Name', 'Gender', 'Phone', 'Email', 'Registration Date', 'Insurance'])
        for p in Patient.objects.filter(is_active=True).order_by('-registration_date'):
            writer.writerow([p.patient_id, p.first_name, p.last_name, p.gender, p.phone, p.email, p.registration_date, p.insurance_provider or 'None'])

    elif module == 'appointments':
        from appointments.models import Appointment
        writer.writerow(['Date', 'Time', 'Patient', 'Service', 'Provider', 'Status'])
        for a in Appointment.objects.filter(appointment_date__range=[start, end]).select_related('patient', 'service', 'provider').order_by('-appointment_date'):
            writer.writerow([a.appointment_date, a.appointment_time, a.patient.get_full_name() if a.patient else '', a.service.name if a.service else '', a.provider.get_full_name() if a.provider else '', a.get_status_display()])

    elif module == 'billing':
        from billing.models import Payment
        writer.writerow(['Date', 'Patient', 'Invoice #', 'Amount', 'Method', 'Status'])
        for p in Payment.objects.filter(payment_date__date__range=[start, end]).select_related('invoice__patient').order_by('-payment_date'):
            patient_name = p.invoice.patient.get_full_name() if p.invoice and p.invoice.patient else ''
            inv_num = p.invoice.invoice_number if p.invoice else ''
            writer.writerow([p.payment_date.strftime('%Y-%m-%d'), patient_name, inv_num, float(p.amount), p.get_payment_method_display(), p.status])

    elif module == 'pharmacy':
        from pharmacy.models import Medication
        writer.writerow(['Name', 'Category', 'Form', 'Current Stock', 'Reorder Level', 'Cost Price', 'Selling Price', 'Active'])
        for m in Medication.objects.all().select_related('category').order_by('name'):
            writer.writerow([m.name, m.category.name if m.category else '', m.dosage_form, m.current_stock, m.reorder_level, float(m.cost_price or 0), float(m.selling_price or 0), m.is_active])

    elif module == 'laboratory':
        from laboratory.models import LabTestRequest
        writer.writerow(['Date', 'Patient', 'Test', 'Priority', 'Status', 'Requested By'])
        for r in LabTestRequest.objects.filter(requested_date__date__range=[start, end]).select_related('patient', 'test', 'requested_by').order_by('-requested_date'):
            writer.writerow([r.requested_date.strftime('%Y-%m-%d'), r.patient.get_full_name() if r.patient else '', r.test.name if r.test else '', r.priority, r.get_status_display(), r.requested_by.get_full_name() if r.requested_by else ''])

    elif module == 'staff':
        from staff_management.models import Staff
        writer.writerow(['Name', 'Department', 'Position', 'Employment Status', 'Join Date', 'Active'])
        for s in Staff.objects.filter(is_active=True).select_related('department', 'user').order_by('user__last_name'):
            writer.writerow([s.user.get_full_name(), s.department.name if s.department else '', s.position, s.get_employment_status_display(), s.join_date, s.is_active])

    elif module == 'overview':
        from patients.models import Patient
        from appointments.models import Appointment
        from billing.models import Payment, Invoice
        from pharmacy.models import Medication
        from laboratory.models import LabTestRequest
        from staff_management.models import Staff
        from budget.models import Expense

        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Period', f'{start} to {end}'])
        writer.writerow([])
        writer.writerow(['--- Patients ---'])
        writer.writerow(['Total Active Patients', Patient.objects.filter(is_active=True).count()])
        writer.writerow(['New Patients (Period)', Patient.objects.filter(registration_date__range=[start, end]).count()])
        writer.writerow([])
        writer.writerow(['--- Appointments ---'])
        appts = Appointment.objects.filter(appointment_date__range=[start, end])
        writer.writerow(['Total Appointments', appts.count()])
        writer.writerow(['Completed', appts.filter(status='completed').count()])
        writer.writerow(['Cancelled', appts.filter(status='cancelled').count()])
        writer.writerow([])
        writer.writerow(['--- Revenue ---'])
        writer.writerow(['Total Revenue', float(Payment.objects.filter(payment_date__date__range=[start, end], status='completed').aggregate(s=Sum('amount'))['s'] or 0)])
        writer.writerow(['Invoices Created', Invoice.objects.filter(created_at__date__range=[start, end]).count()])
        writer.writerow(['Outstanding', float(Invoice.objects.filter(status__in=['sent', 'overdue']).aggregate(s=Sum('total_amount'))['s'] or 0)])
        writer.writerow([])
        writer.writerow(['--- Pharmacy ---'])
        writer.writerow(['Total Medications', Medication.objects.count()])
        writer.writerow(['Low Stock', Medication.objects.filter(current_stock__lte=F('reorder_level')).count()])
        writer.writerow([])
        writer.writerow(['--- Laboratory ---'])
        lab = LabTestRequest.objects.filter(requested_date__date__range=[start, end])
        writer.writerow(['Total Requests', lab.count()])
        writer.writerow(['Completed', lab.filter(status='completed').count()])
        writer.writerow([])
        writer.writerow(['--- Staff ---'])
        writer.writerow(['Active Staff', Staff.objects.filter(is_active=True).count()])
        writer.writerow([])
        writer.writerow(['--- Expenses ---'])
        writer.writerow(['Total Expenses', float(Expense.objects.filter(date__range=[start, end], status='approved').aggregate(s=Sum('amount'))['s'] or 0)])

    return response


@staff_member_required
def export_stats_excel(request, module):
    """Export statistics for a module as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    start, end = _parse_dates(request)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = module.title()

    title_font = Font(name='Arial', size=14, bold=True, color='1B5E96')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1B5E96', end_color='1B5E96', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.cell(row=1, column=1, value=f'{module.title()} Statistics Report').font = title_font
    ws.cell(row=2, column=1, value=f'Period: {start} to {end}').font = Font(name='Arial', size=10, italic=True)

    def write_header(row, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        return row + 1

    def write_row(row, values):
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = thin_border
        return row + 1

    r = 4

    if module == 'patients':
        from patients.models import Patient
        r = write_header(r, ['Patient ID', 'First Name', 'Last Name', 'Gender', 'Phone', 'Email', 'Registration Date', 'Insurance'])
        for p in Patient.objects.filter(is_active=True).order_by('-registration_date'):
            r = write_row(r, [p.patient_id, p.first_name, p.last_name, p.gender, p.phone, p.email, str(p.registration_date), p.insurance_provider or 'None'])

    elif module == 'appointments':
        from appointments.models import Appointment
        r = write_header(r, ['Date', 'Time', 'Patient', 'Service', 'Provider', 'Status'])
        for a in Appointment.objects.filter(appointment_date__range=[start, end]).select_related('patient', 'service', 'provider').order_by('-appointment_date'):
            r = write_row(r, [str(a.appointment_date), str(a.appointment_time), a.patient.get_full_name() if a.patient else '', a.service.name if a.service else '', a.provider.get_full_name() if a.provider else '', a.get_status_display()])

    elif module == 'billing':
        from billing.models import Payment
        r = write_header(r, ['Date', 'Patient', 'Invoice #', 'Amount', 'Method', 'Status'])
        for p in Payment.objects.filter(payment_date__date__range=[start, end]).select_related('invoice__patient').order_by('-payment_date'):
            patient_name = p.invoice.patient.get_full_name() if p.invoice and p.invoice.patient else ''
            inv_num = p.invoice.invoice_number if p.invoice else ''
            r = write_row(r, [p.payment_date.strftime('%Y-%m-%d'), patient_name, inv_num, float(p.amount), p.get_payment_method_display(), p.status])

    elif module == 'pharmacy':
        from pharmacy.models import Medication
        r = write_header(r, ['Name', 'Category', 'Form', 'Stock', 'Reorder Level', 'Cost', 'Selling Price'])
        for m in Medication.objects.all().select_related('category').order_by('name'):
            r = write_row(r, [m.name, m.category.name if m.category else '', m.dosage_form, m.current_stock, m.reorder_level, float(m.cost_price or 0), float(m.selling_price or 0)])

    elif module == 'laboratory':
        from laboratory.models import LabTestRequest
        r = write_header(r, ['Date', 'Patient', 'Test', 'Priority', 'Status', 'Requested By'])
        for req in LabTestRequest.objects.filter(requested_date__date__range=[start, end]).select_related('patient', 'test', 'requested_by').order_by('-requested_date'):
            r = write_row(r, [req.requested_date.strftime('%Y-%m-%d'), req.patient.get_full_name() if req.patient else '', req.test.name if req.test else '', req.priority, req.get_status_display(), req.requested_by.get_full_name() if req.requested_by else ''])

    elif module == 'staff':
        from staff_management.models import Staff
        r = write_header(r, ['Name', 'Department', 'Position', 'Employment Status', 'Join Date'])
        for s in Staff.objects.filter(is_active=True).select_related('department', 'user').order_by('user__last_name'):
            r = write_row(r, [s.user.get_full_name(), s.department.name if s.department else '', s.position, s.get_employment_status_display(), str(s.join_date)])

    # Auto-width columns
    from openpyxl.utils import get_column_letter
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{module}_statistics_{start}_{end}.xlsx"'
    return response
